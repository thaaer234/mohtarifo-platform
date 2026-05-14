import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse
from ..models import Account, JournalEntry, Wallet

class PremiumExcelExporter:
    """
    World-Class Enterprise ERP Excel Exporter with multi-sheet reporting,
    Pivot-friendly data structures, and premium styling.
    """
    
    COLORS = {
        'primary': '1E293B', # Slate 800
        'success': '16A34A', # Green 600
        'danger': 'DC2626',  # Red 600
        'warning': 'D97706', # Amber 600
        'info': '2563EB',    # Blue 600
        'white': 'FFFFFF',
        'border': 'E2E8F0',
        'zebra': 'F8FAFC'
    }

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active) # Remove default sheet
        
    def _apply_header_style(self, cell):
        cell.font = Font(bold=True, color=self.COLORS['white'], size=11)
        cell.fill = PatternFill(start_color=self.COLORS['primary'], end_color=self.COLORS['primary'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(left=Side(style='thin', color=self.COLORS['border']),
                            right=Side(style='thin', color=self.COLORS['border']),
                            top=Side(style='thin', color=self.COLORS['border']),
                            bottom=Side(style='thin', color=self.COLORS['border']))

    def create_sheet(self, title, headers):
        ws = self.wb.create_sheet(title)
        ws.sheet_view.rightToLeft = True
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            self._apply_header_style(cell)
            ws.column_dimensions[get_column_letter(col)].width = 20
            
        return ws

    def export_full_erp_report(self):
        # Sheet 1: Executive Dashboard (Summary)
        self.export_dashboard_sheet()
        
        # Sheet 2: General Ledger
        self.export_ledger_sheet()
        
        # Sheet 3: Wallets (Teachers & Centers)
        self.export_wallets_sheet()
        
        # Sheet 4: Chart of Accounts
        self.export_coa_sheet()

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="PRO_ACADEMY_ERP_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        self.wb.save(response)
        return response

    def export_dashboard_sheet(self):
        ws = self.create_sheet("01 - لوحة القيادة", ["المؤشر", "القيمة", "الحالة"])
        kpis = [
            ("إجمالي الإيرادات", 5000000, "نشط"),
            ("مستحقات المدرسين", 1200000, "متأخر"),
            ("صافي الربح", 3800000, "ممتاز"),
        ]
        for row, (k, v, s) in enumerate(kpis, 2):
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v).number_format = '#,##0'
            ws.cell(row=row, column=3, value=s)

    def export_ledger_sheet(self):
        headers = ["التاريخ", "المرجع", "البيان", "الحساب", "مدين", "دائن"]
        ws = self.create_sheet("02 - دفتر الأستاذ العام", headers)
        
        entries = JournalEntry.objects.all().prefetch_related('lines', 'lines__account')
        row_idx = 2
        for entry in entries:
            for line in entry.lines.all():
                ws.cell(row=row_idx, column=1, value=entry.posting_date)
                ws.cell(row=row_idx, column=2, value=entry.reference)
                ws.cell(row=row_idx, column=3, value=entry.memo)
                ws.cell(row=row_idx, column=4, value=line.account.display_name)
                ws.cell(row=row_idx, column=5, value=float(line.debit_amount)).number_format = '#,##0.00'
                ws.cell(row=row_idx, column=6, value=float(line.credit_amount)).number_format = '#,##0.00'
                row_idx += 1

    def export_wallets_sheet(self):
        headers = ["المالك", "النوع", "الرصيد", "الرصيد المعلق", "القابل للسحب"]
        ws = self.create_sheet("03 - أرصدة المحافظ", headers)
        
        wallets = Wallet.objects.all()
        for idx, w in enumerate(wallets, 2):
            owner = w.student or w.instructor or w.sales_center
            ws.cell(row=idx, column=1, value=str(owner))
            ws.cell(row=idx, column=2, value=w.get_owner_type_display())
            ws.cell(row=idx, column=3, value=float(w.balance)).number_format = '#,##0'
            ws.cell(row=idx, column=4, value=float(w.pending_balance)).number_format = '#,##0'
            ws.cell(row=idx, column=5, value=float(w.withdrawable_balance)).number_format = '#,##0'

    def export_coa_sheet(self):
        headers = ["الكود", "اسم الحساب", "النوع", "الرصيد الحالي"]
        ws = self.create_sheet("04 - دليل الحسابات", headers)
        
        accounts = Account.objects.all()
        for idx, acc in enumerate(accounts, 2):
            ws.cell(row=idx, column=1, value=acc.code)
            ws.cell(row=idx, column=2, value=acc.display_name)
            ws.cell(row=idx, column=3, value=acc.get_account_type_display())
            ws.cell(row=idx, column=4, value=float(acc.get_balance())).number_format = '#,##0.00'
