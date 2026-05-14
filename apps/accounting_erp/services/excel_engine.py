import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime
from django.http import HttpResponse

class HighQualityExcelExporter:
    """
    محرك تصدير إكسيل عالي الجودة متوافق مع متطلبات المؤسسات (Enterprise Quality).
    يدعم التصميم الملون، الجداول المنظمة، وتنسيق العملات.
    """
    
    # هوية المؤسسة (أخضر تعليمي، داكن، كهرماني)
    PRIMARY_COLOR = "047857"   # الأخضر (Educational Green)
    SECONDARY_COLOR = "111827" # الداكن (Dark Ink)
    ACCENT_COLOR = "F59E0B"    # الكهرماني (Warm Amber)
    HEADER_TEXT_COLOR = "FFFFFF"

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.setup_styles()

    def setup_styles(self):
        # Header Style
        self.header_font = Font(name='Arial', size=11, bold=True, color=self.HEADER_TEXT_COLOR)
        self.header_fill = PatternFill(start_color=self.PRIMARY_COLOR, end_color=self.PRIMARY_COLOR, fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Data Style
        self.data_font = Font(name='Arial', size=10)
        self.border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='thin', color='D1D5DB')
        )
        self.rtl_alignment = Alignment(horizontal='right', vertical='center')
        self.number_format = '#,##0.00'

    def export_trial_balance(self, data, period_display="All Periods"):
        ws = self.wb.active
        ws.title = "ميزان المراجعة"
        ws.sheet_view.rightToLeft = True

        # 1. Report Title
        ws.merge_cells('A1:E1')
        ws['A1'] = "ميزان المراجعة - Trial Balance"
        ws['A1'].font = Font(size=14, bold=True, color=self.PRIMARY_COLOR)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:E2')
        ws['A2'] = f"الفترة: {period_display}"
        ws['A2'].alignment = Alignment(horizontal='center')

        # 2. Headers
        headers = ["كود الحساب", "اسم الحساب", "الرصيد المدين", "الرصيد الدائن", "الصافي"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=text)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # 3. Data
        row = 5
        for item in data:
            ws.cell(row=row, column=1, value=item['code']).border = self.border
            ws.cell(row=row, column=2, value=item['name']).border = self.border
            
            dr_cell = ws.cell(row=row, column=3, value=float(item['debit']))
            dr_cell.number_format = self.number_format
            dr_cell.border = self.border
            
            cr_cell = ws.cell(row=row, column=4, value=float(item['credit']))
            cr_cell.number_format = self.number_format
            cr_cell.border = self.border
            
            net_cell = ws.cell(row=row, column=5, value=float(item['net']))
            net_cell.number_format = self.number_format
            net_cell.border = self.border
            net_cell.font = Font(bold=True)
            
            row += 1

        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 5

        # Create Response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Trial_Balance_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        self.wb.save(response)
        return response
