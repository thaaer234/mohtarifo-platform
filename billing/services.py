import csv
import secrets
from io import TextIOWrapper

from django import forms
from django.utils import timezone

from .models import AccessCode


def create_codes_from_upload(batch, uploaded_file, free_codes=True):
    rows = read_student_rows(uploaded_file)
    created = 0
    for row in rows:
        student_name = (row.get("name") or row.get("الاسم") or row.get("student_name") or "").strip()
        student_phone = (row.get("phone") or row.get("الهاتف") or row.get("رقم الهاتف") or row.get("mobile") or "").strip()
        if not student_name and not student_phone:
            continue
        AccessCode.objects.create(
            code=unique_code(batch),
            access_type="course",
            course=batch.course,
            batch=batch,
            institute=batch.institute,
            sales_center=batch.sales_center,
            assigned_student_name=student_name,
            assigned_student_phone=student_phone,
            sale_status="free" if free_codes else "reserved",
            is_free_code=free_codes,
            max_redemptions=1,
            valid_until=timezone.now() + timezone.timedelta(days=180),
            notes=f"تم توليده من دفعة {batch.name}",
        )
        created += 1
    if created:
        batch.allocated_count = batch.codes.count()
        batch.free_count = batch.codes.filter(is_free_code=True).count()
        batch.save(update_fields=["allocated_count", "free_count"])
    return created


def create_codes_for_batch(batch, quantity, free_codes=False):
    created = 0
    access_type = "package" if batch.package_id else "course"
    for _index in range(quantity):
        AccessCode.objects.create(
            code=unique_code(batch),
            access_type=access_type,
            course=batch.course,
            package=batch.package,
            batch=batch,
            institute=batch.institute,
            sales_center=batch.sales_center,
            sale_status="free" if free_codes else "available",
            is_free_code=free_codes,
            max_redemptions=1,
            valid_until=timezone.now() + timezone.timedelta(days=180),
            notes=f"تم توليده من دفعة {batch.name}",
        )
        created += 1
    if created:
        batch.allocated_count = batch.codes.count()
        batch.free_count = batch.codes.filter(is_free_code=True).count()
        batch.save(update_fields=["allocated_count", "free_count"])
    return created


def read_student_rows(uploaded_file):
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        wrapper = TextIOWrapper(uploaded_file.file, encoding="utf-8-sig")
        return list(csv.DictReader(wrapper))
    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise forms.ValidationError("قراءة XLSX تحتاج تثبيت openpyxl.") from exc
        workbook = load_workbook(uploaded_file, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value or "").strip() for value in rows[0]]
        return [{headers[index]: str(value or "").strip() for index, value in enumerate(row)} for row in rows[1:]]
    raise forms.ValidationError("الملف يجب أن يكون CSV أو XLSX.")


def unique_code(batch):
    prefix = (batch.code_prefix or "M").upper().replace("-", "")
    # Generate a more random string to prevent guessing
    # Format: PREFIX-XXXX-XXXX
    while True:
        # secrets.token_urlsafe(6) generates 8 chars
        suffix = secrets.token_urlsafe(6).upper().replace("_", "X").replace("-", "Y")
        code = f"{prefix}-{suffix[:4]}-{suffix[4:8]}"
        if not AccessCode.objects.filter(code=code).exists():
            return code
