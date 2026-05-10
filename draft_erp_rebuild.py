import os
import django
import sys

sys.path.append(os.getcwd())
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
django.setup()

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.utils import timezone
from django.db.models import Sum, Count

from billing.models import AccessCode, SalesCenter
from learning.models import Course
from accounts.models import StudentProfile, InstructorProfile

def _apply_premium_base(ws):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False

def _draw_table(ws, sr, er, sc, ec, zebra=True):
    head_f = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    head_font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    row_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_b = Side(style="thin", color="E2E8F0")
    uni_border = Border(left=thin_b, right=thin_b, top=thin_b, bottom=thin_b)
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = uni_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r == sr:
                cell.fill = head_f
                cell.font = head_font
            else:
                cell.font = Font(size=9, name="Segoe UI", color="334155")
                if zebra and (r % 2 == 0):
                    cell.fill = row_even

def _draw_card(ws, row, col, title, val_formula, sub_label, hex_theme="233646"):
    head_f = PatternFill(start_color=hex_theme, end_color=hex_theme, fill_type="solid")
    body_f = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    foot_f = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin = Side(style="thin", color="D1D5DB")
    heavy_b = Side(style="medium", color=hex_theme)
    
    c1 = ws.cell(row=row, column=col)
    c1.value = title
    c1.font = Font(bold=True, color="FFFFFF", size=11, name="Segoe UI")
    c1.fill = head_f
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = Border(top=thin, left=thin, right=thin)

    c2 = ws.cell(row=row+1, column=col)
    c2.value = val_formula
    c2.font = Font(bold=True, color="1F2937", size=24, name="Segoe UI")
    c2.fill = body_f
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.number_format = '#,##0'
    c2.border = Border(left=thin, right=thin)

    c3 = ws.cell(row=row+2, column=col)
    c3.value = sub_label
    c3.font = Font(italic=True, color="6B7280", size=9, name="Segoe UI")
    c3.fill = foot_f
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border = Border(left=thin, right=thin, bottom=heavy_b)

def build_erp():
    wb = Workbook()
    wb.remove(wb.active)

    # Sheets setup
    sheet_names = [
        "00 - لوحة القيادة التنفيذية",
        "01 - الإعدادات المالية",
        "02 - محرك الحسابات المالي",
        "03 - تقرير المدرسين",
        "04 - تقرير المراكز",
        "05 - تقرير الطلاب",
        "06 - ربحية الدورات",
        "07 - تتبع الأقساط",
        "08 - الديون المتأخرة",
        "09 - الاستحقاقات",
        "10 - قائمة الدخل",
        "11 - التدفق النقدي",
        "12 - مؤشرات الأداء",
        "13 - تنبيهات مالية",
        "99 - البيانات الخام"
    ]
    sheets = {}
    for sn in sheet_names:
        ws = wb.create_sheet(sn)
        _apply_premium_base(ws)
        sheets[sn] = ws

    # ---------------------------------------------------------
    # 01 - FINANCIAL SETTINGS
    # ---------------------------------------------------------
    stg = sheets["01 - الإعدادات المالية"]
    stg.column_dimensions['B'].width = 35
    stg.column_dimensions['C'].width = 20
    
    stg.append(["", "⚙️ الإعدادات العامة (الافتراضية)", "القيمة"])
    stg.append(["", "حصة المدرس الافتراضية (%)", 0.40])      # C2
    stg.append(["", "حصة المركز الافتراضية (%)", 0.15])      # C3
    stg.append(["", "حصة المنصة الافتراضية (%)", 0.45])      # C4
    stg.append(["", "مصاريف تشغيل / تسويق (%)", 0.05])       # C5
    stg.append(["", "نسبة الضريبة المقتطعة (%)", 0.00])       # C6
    
    # Format settings
    for r in range(2, 7):
        stg.cell(row=r, column=3).number_format = '0%'
    _draw_table(stg, 1, 6, 2, 3)

    # Teacher Overrides Table
    stg.append([])
    stg.append(["", "👨‍🏫 استثناءات نسب المدرسين", "النسبة المخصصة"])
    stg.append(["", "المدرس", "نسبة الحصة"]) # Row 9
    
    t_row = 10
    for t in InstructorProfile.objects.select_related("user"):
        nm = t.user.get_full_name() or t.user.username
        stg.append(["", nm, ""]) # Leave blank for user override
        stg.cell(row=t_row, column=3).number_format = '0%'
        t_row += 1
    stg.append(["", "استثناء 1", 0.50])
    stg.append(["", "استثناء 2", 0.60])
    _draw_table(stg, 8, max(t_row-1, 11), 2, 3)
    
    # Center Overrides Table
    c_start = max(t_row, 12) + 2
    stg.cell(row=c_start, column=2, value="🏢 استثناءات نسب المراكز")
    stg.cell(row=c_start+1, column=2, value="المركز")
    stg.cell(row=c_start+1, column=3, value="نسبة الحصة")
    
    c_row = c_start + 2
    for c in SalesCenter.objects.all():
        stg.cell(row=c_row, column=2, value=c.name)
        stg.cell(row=c_row, column=3, value="")
        stg.cell(row=c_row, column=3).number_format = '0%'
        c_row += 1
    stg.cell(row=c_row, column=2, value="مركز الأمل")
    stg.cell(row=c_row, column=3, value=0.20)
    _draw_table(stg, c_start, c_row, 2, 3)

    # Course Overrides Table
    cr_start = c_row + 2
    stg.cell(row=cr_start, column=2, value="📚 استثناءات المواد (أقوى أولوية)")
    stg.cell(row=cr_start+1, column=2, value="المادة")
    stg.cell(row=cr_start+1, column=3, value="حصة المدرس للمادة")
    cr_row = cr_start + 2
    for c in Course.objects.all():
        stg.cell(row=cr_row, column=2, value=c.title)
        stg.cell(row=cr_row, column=3, value="")
        stg.cell(row=cr_row, column=3).number_format = '0%'
        cr_row += 1
    _draw_table(stg, cr_start, cr_row-1, 2, 3)

    # ---------------------------------------------------------
    # 99 - RAW DATA
    # ---------------------------------------------------------
    raw = sheets["99 - البيانات الخام"]
    raw.append(["رقم المعاملة", "التاريخ", "الطالب", "المادة", "المدرس", "المركز", "المبلغ المدفوع", "الحالة"])
    
    sales = AccessCode.objects.filter(sale_status="sold").select_related("course", "course__instructor", "sales_center").order_by("-sold_at")
    raw_r = 2
    for s in sales.iterator(chunk_size=1000):
        dt = s.sold_at.strftime("%Y-%m-%d") if s.sold_at else "2026-01-01"
        st_nm = s.assigned_student_name or "مجهول"
        crs_nm = s.course.title if s.course else "عام"
        tch_nm = s.course.instructor.get_full_name() if (s.course and s.course.instructor) else "-"
        cnt_nm = s.sales_center.name if s.sales_center else "منصة"
        amt = int((s.sold_price_cents or 0)/100)
        
        raw.append([s.id, dt, st_nm, crs_nm, tch_nm, cnt_nm, amt, "مكتمل"])
        raw_r += 1
    _draw_table(raw, 1, max(2, raw_r-1), 1, 8)

    # ---------------------------------------------------------
    # 02 - FINANCIAL ENGINE (MASTER COMPUTATION)
    # ---------------------------------------------------------
    eng = sheets["02 - محرك الحسابات المالي"]
    eng_cols = ["ID", "التاريخ", "الطالب", "المادة", "المدرس", "المركز", "المبلغ المدفوع", 
                "نسبة المدرس", "قيمة المدرس", "نسبة المركز", "قيمة المركز", 
                "نسبة التشغيل", "قيمة التشغيل", "صافي المنصة"]
    eng.append(eng_cols)
    for i in range(1, 15): eng.column_dimensions[get_column_letter(i)].width = 15
    
    # Formulas that do priority lookups!
    # Course Override: VLOOKUP(D2, Settings!B$cr_start:C$cr_row, 2, FALSE)
    # Teacher Override: VLOOKUP(E2, Settings!B$10:C$t_row, 2, FALSE)
    # Default: Settings!C2
    
    c_rng = f"'01 - الإعدادات المالية'!B${cr_start+2}:C${cr_row-1}"
    t_rng = f"'01 - الإعدادات المالية'!B$10:C${t_row}"
    cnt_rng = f"'01 - الإعدادات المالية'!B${c_start+2}:C${c_row}"
    
    def_t_pct = "'01 - الإعدادات المالية'!C$2"
    def_c_pct = "'01 - الإعدادات المالية'!C$3"
    def_m_pct = "'01 - الإعدادات المالية'!C$5" # Marketing/Ops
    
    for r in range(2, max(3, raw_r)):
        # Copy from raw
        eng.append([
            f"='99 - البيانات الخام'!A{r}", f"='99 - البيانات الخام'!B{r}", f"='99 - البيانات الخام'!C{r}",
            f"='99 - البيانات الخام'!D{r}", f"='99 - البيانات الخام'!E{r}", f"='99 - البيانات الخام'!F{r}",
            f"='99 - البيانات الخام'!G{r}"
        ])
        
        # Teacher %: IF Course override exists use it, ELSE IF Teacher override exists use it, ELSE Default
        # In Excel: IFERROR(VLOOKUP(Course, CRng, 2, FALSE), IFERROR(VLOOKUP(Teacher, TRng, 2, FALSE), Default))
        # But VLOOKUP on empty returns 0. Let's use IF(ISNUMBER(VLOOKUP), VLOOKUP, ...)
        # A simpler way: IF(IFERROR(VLOOKUP(D2, CRng, 2, FALSE), "")<>"", VLOOKUP(D2, CRng, 2, FALSE), IF(IFERROR(VLOOKUP(E2, TRng, 2, FALSE), "")<>"", VLOOKUP(E2, TRng, 2, FALSE), Def))
        t_pct_f = f'=IFERROR(1/(1/VLOOKUP(D{r},{c_rng},2,FALSE)), IFERROR(1/(1/VLOOKUP(E{r},{t_rng},2,FALSE)), {def_t_pct}))'
        
        # Center %: 
        c_pct_f = f'=IF(F{r}="منصة", 0, IFERROR(1/(1/VLOOKUP(F{r},{cnt_rng},2,FALSE)), {def_c_pct}))'
        
        # Ops %
        o_pct_f = f'={def_m_pct}'
        
        eng.cell(row=r, column=8, value=t_pct_f).number_format = '0%'
        eng.cell(row=r, column=9, value=f'=G{r}*H{r}') # Teacher val
        eng.cell(row=r, column=10, value=c_pct_f).number_format = '0%'
        eng.cell(row=r, column=11, value=f'=G{r}*J{r}') # Center val
        eng.cell(row=r, column=12, value=o_pct_f).number_format = '0%'
        eng.cell(row=r, column=13, value=f'=G{r}*L{r}') # Ops val
        
        # Net Platform = Gross - Teacher - Center - Ops
        eng.cell(row=r, column=14, value=f'=G{r}-I{r}-K{r}-M{r}') 
        
    _draw_table(eng, 1, max(2, raw_r-1), 1, 14)

    # ---------------------------------------------------------
    # 03 - TEACHER EARNINGS
    # ---------------------------------------------------------
    tch = sheets["03 - تقرير المدرسين"]
    tch.append(["المدرس", "إجمالي المبيعات", "مستحقات المدرس", "المسدد", "الرصيد المتبقي"])
    for i in range(1, 6): tch.column_dimensions[get_column_letter(i)].width = 20
    t_out = 2
    for t in InstructorProfile.objects.select_related("user"):
        nm = t.user.get_full_name() or t.user.username
        tch.append([
            nm,
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!E:E, A{t_out}, \'02 - محرك الحسابات المالي\'!G:G)',
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!E:E, A{t_out}, \'02 - محرك الحسابات المالي\'!I:I)',
            0, # Paid
            f'=C{t_out}-D{t_out}'
        ])
        t_out += 1
    _draw_table(tch, 1, max(2, t_out-1), 1, 5)

    # ---------------------------------------------------------
    # 04 - CENTER EARNINGS
    # ---------------------------------------------------------
    cnt = sheets["04 - تقرير المراكز"]
    cnt.append(["المركز", "إجمالي المبيعات عبر المركز", "مستحقات المركز", "المسدد", "الرصيد"])
    for i in range(1, 6): cnt.column_dimensions[get_column_letter(i)].width = 20
    c_out = 2
    for c in SalesCenter.objects.all():
        cnt.append([
            c.name,
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!F:F, A{c_out}, \'02 - محرك الحسابات المالي\'!G:G)',
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!F:F, A{c_out}, \'02 - محرك الحسابات المالي\'!K:K)',
            0,
            f'=C{c_out}-D{c_out}'
        ])
        c_out += 1
    _draw_table(cnt, 1, max(2, c_out-1), 1, 5)

    # ---------------------------------------------------------
    # 06 - COURSE PROFITABILITY
    # ---------------------------------------------------------
    cpf = sheets["06 - ربحية الدورات"]
    cpf.append(["المادة", "إجمالي المبيعات", "حصة المدرس", "حصة المراكز", "تكاليف تشغيلية", "صافي ربح المنصة", "هامش الربح"])
    for i in range(1, 8): cpf.column_dimensions[get_column_letter(i)].width = 20
    cf_out = 2
    for c in Course.objects.all():
        cpf.append([
            c.title,
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!D:D, A{cf_out}, \'02 - محرك الحسابات المالي\'!G:G)',
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!D:D, A{cf_out}, \'02 - محرك الحسابات المالي\'!I:I)',
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!D:D, A{cf_out}, \'02 - محرك الحسابات المالي\'!K:K)',
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!D:D, A{cf_out}, \'02 - محرك الحسابات المالي\'!M:M)',
            f'=SUMIF(\'02 - محرك الحسابات المالي\'!D:D, A{cf_out}, \'02 - محرك الحسابات المالي\'!N:N)',
            f'=IFERROR(F{cf_out}/B{cf_out}, 0)'
        ])
        cpf.cell(row=cf_out, column=7).number_format = '0%'
        cf_out += 1
    _draw_table(cpf, 1, max(2, cf_out-1), 1, 7)

    # ---------------------------------------------------------
    # 00 - EXECUTIVE DASHBOARD
    # ---------------------------------------------------------
    dsh = sheets["00 - لوحة القيادة التنفيذية"]
    for char in ['B','C','D','E', 'F']: dsh.column_dimensions[char].width = 25
    dsh.merge_cells('B2:F2')
    dsh['B2'] = "محترفو التعليم | 🚀 ERP FINANCIAL INTELLIGENCE"
    dsh['B2'].font = Font(bold=True, size=18, color="FFFFFF", name="Segoe UI")
    dsh['B2'].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    dsh['B2'].alignment = Alignment(horizontal="center", vertical="center")
    
    _draw_card(dsh, 4, 2, "إجمالي الإيرادات (Gross)", "=SUM('02 - محرك الحسابات المالي'!G:G)", "مبيعات الكلية", "059669")
    _draw_card(dsh, 4, 3, "مستحقات المدرسين", "=SUM('02 - محرك الحسابات المالي'!I:I)", "التزامات مستحقة", "DC2626")
    _draw_card(dsh, 4, 4, "عمولات المراكز", "=SUM('02 - محرك الحسابات المالي'!K:K)", "أرباح الفروع", "EA580C")
    _draw_card(dsh, 4, 5, "مصاريف تشغيلية", "=SUM('02 - محرك الحسابات المالي'!M:M)", "استضافة وتسويق", "64748B")
    _draw_card(dsh, 4, 6, "صافي ربح المنصة (Net)", "=SUM('02 - محرك الحسابات المالي'!N:N)", "الربح الفعلي", "2563EB")
    
    # Financial KPI Table
    dsh.append([])
    dsh.append([])
    dsh.append([])
    dsh.cell(row=8, column=2, value="📈 مؤشرات الأداء الحيوية (KPIs)")
    dsh.cell(row=8, column=2).font = Font(bold=True, size=14, color="1E293B")
    
    dsh.cell(row=10, column=2, value="متوسط نسبة ربح المنصة")
    dsh.cell(row=10, column=3, value="=IFERROR(F5/B5, 0)")
    dsh.cell(row=10, column=3).number_format = '0%'
    
    dsh.cell(row=11, column=2, value="أفضل مادة مبيعاً")
    dsh.cell(row=11, column=3, value="=INDEX('06 - ربحية الدورات'!A2:A100, MATCH(MAX('06 - ربحية الدورات'!B2:B100), '06 - ربحية الدورات'!B2:B100, 0))")

    dsh.cell(row=12, column=2, value="أفضل مدرس مبيعاً")
    dsh.cell(row=12, column=3, value="=INDEX('03 - تقرير المدرسين'!A2:A100, MATCH(MAX('03 - تقرير المدرسين'!B2:B100), '03 - تقرير المدرسين'!B2:B100, 0))")
    
    _draw_table(dsh, 10, 12, 2, 3)

    # P&L
    pnl = sheets["10 - قائمة الدخل"]
    pnl.column_dimensions['B'].width = 35
    pnl.column_dimensions['C'].width = 20
    pnl.append(["", "قائمة الدخل الشاملة (P&L)", "القيمة"])
    pnl.append(["", "إجمالي الإيرادات (Gross Revenue)", "=SUM('02 - محرك الحسابات المالي'!G:G)"])
    pnl.append(["", "(-) الخصومات والمنح", 0])
    pnl.append(["", "صافي الإيرادات", "=C2-C3"])
    pnl.append(["", "(-) مستحقات المدرسين (COGS 1)", "=SUM('02 - محرك الحسابات المالي'!I:I)"])
    pnl.append(["", "(-) عمولات المراكز (COGS 2)", "=SUM('02 - محرك الحسابات المالي'!K:K)"])
    pnl.append(["", "إجمالي الربح (Gross Profit)", "=C4-C5-C6"])
    pnl.append(["", "(-) المصاريف التشغيلية (Opex)", "=SUM('02 - محرك الحسابات المالي'!M:M)"])
    pnl.append(["", "صافي الربح قبل الضريبة (EBT)", "=C7-C8"])
    pnl.append(["", "(-) الضريبة المقتطعة", "='01 - الإعدادات المالية'!C6 * C9"])
    pnl.append(["", "صافي الدخل (Net Income)", "=C9-C10"])
    _draw_table(pnl, 1, 11, 2, 3)

    wb.save("DRAFT_REBUILT_ERP.xlsx")
    print("Successfully built the new percentage-driven dynamic ERP Excel.")

build_erp()
