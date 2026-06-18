from functools import wraps
from collections import defaultdict
import secrets
import base64
import os
from io import BytesIO
import qrcode
from xml.sax.saxutils import escape
import threading
import hashlib
import time

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.cache import cache
from django.core import management
from django.core.exceptions import PermissionDenied, ValidationError
from django.db import models, transaction
from django.db.models import Avg, Count, Sum
from django.core import signing
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.core.paginator import Paginator
from openpyxl import Workbook
from django.views.decorators.http import require_POST

from accounts.models import InstructorProfile, StudentProfile
from analytics.models import TopicPerformance
from billing.devices import activate_user_device, device_fingerprint, device_seed, set_device_cookie
from billing.models import AccessCode, AccessCodeBatch, AccessCodePrintLog, AccessGrant, CoursePackage, Institute, Payment, SalesCenter, Subscription, UserDevice
from billing.services import create_codes_for_batch, create_codes_from_upload, unique_code
from exams.models import Attempt, Exam, Question
from learning.models import Course, CourseProgress, Lesson, LessonAttendance, LessonProgress, OnlineLessonSession
from apps.accounting_erp.models import Wallet, WalletTransaction, JournalEntry


from .forms import (
    CatalogSectionForm,
    CourseCardMediaForm,
    CourseCodeBatchForm,
    CourseCodeSaleForm,
    CourseCreateForm,
    CourseEditForm,
    CourseLessonUploadForm,
    CourseUnitQuickForm,
    CoursePackageForm,
    CourseStudentImportForm,
    PackageCodeGenerateForm,
    PackageCodeBatchForm,
    PackageCodeSaleForm,
    InstructorAddForm,
    InstructorEditForm,
    InstituteForm,
    PartnerBatchForm,
    PartnerInstituteImportForm,
    RedeemCodeForm,
    SalesCenterForm,
    StudentRegistrationForm,
    PasswordResetRequestForm,
    SetNewPasswordForm,
)
from .models import CatalogSection, StudentNotification, WhatsAppTemplate
from .seo import _site_url
from .security import sanitize_plain_text, validate_syrian_mobile
from .whatsapp_utils import get_whatsapp_status, logout_whatsapp, send_whatsapp_message, guess_gender_from_name, parse_gender_grammar, is_2fa_disabled
from .otp_service import send_otp, verify_otp, OTP_EXPIRY_SECONDS, OTP_RESEND_COOLDOWN_SECONDS, OTP_LENGTH


def _is_admin_user(user):
    return user.is_authenticated and user.is_active and user.is_superuser


def _is_active_instructor(user):
    profile = getattr(user, "instructor_profile", None)
    return user.is_authenticated and user.is_active and profile is not None and profile.status == "active"


def _role_required(test_func, login_url="dashboard:login"):
    def decorator(view_func):
        @wraps(view_func)
        def wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect(f"{reverse(login_url)}?next={request.get_full_path()}")
            if not test_func(request.user):
                raise PermissionDenied
            return view_func(request, *args, **kwargs)

        return wrapped

    return decorator


admin_required = _role_required(_is_admin_user)
instructor_required = _role_required(_is_active_instructor)


def home(request):
    from analytics.services import TrackingService
    TrackingService.log_landing_visit(request)
    
    if not request.user.is_authenticated:
        return redirect("dashboard:landing")
    if _is_admin_user(request.user):
        return admin_dashboard(request)
    if _is_active_instructor(request.user):
        return instructor_dashboard(request)
    return student_dashboard(request)


def landing_page(request):
    from analytics.services import TrackingService
    TrackingService.log_landing_visit(request)
    instructor_id = request.GET.get("instructor")
    catalog_tabs = _catalog_tabs()
    requested_kind = request.GET.get("kind")
    requested_track = request.GET.get("track")
    selected_tab = None
    if requested_kind and requested_track:
        selected_tab = next(
            (tab for tab in catalog_tabs if tab["kind"] == requested_kind and tab["track"] == requested_track),
            None,
        )
    if not selected_tab and catalog_tabs:
        selected_tab = catalog_tabs[0]
    selected_kind = selected_tab["kind"] if selected_tab else ""
    selected_track = selected_tab["track"] if selected_tab else ""
    
    courses_query = Course.objects.filter(status="published")
    
    instructor_name = None
    if instructor_id:
        courses_query = courses_query.filter(instructor_id=instructor_id)
        instructor = User.objects.filter(id=instructor_id).first()
        if instructor:
            instructor_name = instructor.get_full_name() or instructor.username
    elif selected_tab:
        if selected_track in {"scientific", "literary"}:
            courses_query = courses_query.filter(kind=selected_kind).filter(
                models.Q(academic_track=selected_track) | models.Q(academic_track="general")
            )
        else:
            courses_query = courses_query.filter(kind=selected_kind, academic_track=selected_track)
    else:
        courses_query = courses_query.none()
        
    courses = (
        courses_query
        .select_related("subject", "instructor", "instructor__instructor_profile")
        .annotate(lessons_total=Count("units__lessons", distinct=True))
        .order_by("subject__name", "title")
    )
    
    return render(
        request,
        "dashboard/landing.html",
        {
            "courses": courses,
            "selected_kind": selected_kind,
            "selected_track": selected_track,
            "catalog_tabs": catalog_tabs,
            "filtered_instructor_name": instructor_name,
        },
    )


def shop_page(request):
    selected_kind = request.GET.get("kind", "")
    selected_track = request.GET.get("track", "")
    courses_query = Course.objects.filter(status="published").select_related("subject", "instructor", "instructor__instructor_profile")
    if selected_kind:
        courses_query = courses_query.filter(kind=selected_kind)
    if selected_track:
        if selected_track in {"scientific", "literary"}:
            courses_query = courses_query.filter(models.Q(academic_track=selected_track) | models.Q(academic_track="general"))
        else:
            courses_query = courses_query.filter(academic_track=selected_track)
    courses = (
        courses_query
        .annotate(
            lessons_total=Count("units__lessons", distinct=True),
            sales_centers_total=Count("access_codes__sales_center", distinct=True),
        )
        .order_by("academic_track", "kind", "subject__name", "title")
    )
    sales_centers = SalesCenter.objects.filter(is_active=True).exclude(name="حساب شام كاش (الإدارة)").select_related("institute").order_by("name")
    return render(
        request,
        "dashboard/shop.html",
        {
            "courses": courses,
            "sales_centers": sales_centers,
            "catalog_tabs": _catalog_tabs(),
            "selected_kind": selected_kind,
            "selected_track": selected_track,
        },
    )


def device_logged_out_page(request):
    return render(request, "dashboard/device_logged_out.html")


def public_course_detail(request, course_id):
    course = get_object_or_404(
        Course.objects.filter(status="published")
        .select_related("subject", "instructor", "instructor__instructor_profile")
        .prefetch_related("units__lessons"),
        id=course_id,
    )
    first_lesson = (
        Lesson.objects.filter(unit__course=course, is_free_preview=True, video_url__gt="")
        .select_related("unit", "unit__course")
        .order_by("unit__sort_order", "sort_order", "id")
        .first()
    )
    has_access = False
    if request.user.is_authenticated:
        has_access = _active_access_grants(request.user).filter(course=course).exists()
    sales_centers = (
        SalesCenter.objects.filter(access_codes__course=course, is_active=True)
        .exclude(name="حساب شام كاش (الإدارة)")
        .select_related("institute")
        .distinct()
        .order_by("name")
    )
    return render(
        request,
        "dashboard/public_course_detail.html",
        {
            "course": course,
            "first_lesson": first_lesson,
            "has_access": has_access,
            "sales_centers": sales_centers,
            "video_embed_url": _video_embed_url(first_lesson.video_url) if first_lesson and (has_access or first_lesson.is_free_preview) else "",
        },
    )


def thaaer_review(request):
    courses = (
        Course.objects.filter(status="published")
        .select_related("subject", "instructor", "instructor__instructor_profile")
        .annotate(lessons_total=Count("units__lessons", distinct=True))
        .order_by("subject__name", "title")[:8]
    )
    sample_course = courses[0] if courses else None
    sample_lessons = []
    if sample_course:
        sample_lessons = Lesson.objects.filter(unit__course=sample_course).select_related("unit").order_by("unit__sort_order", "sort_order")[:12]
    context = {
        "courses": courses,
        "sample_course": sample_course,
        "sample_lessons": sample_lessons,
        "stats": {
            "courses": Course.objects.count(),
            "lessons": Lesson.objects.count(),
            "codes": AccessCode.objects.count(),
            "batches": AccessCodeBatch.objects.count(),
            "institutes": Institute.objects.count(),
            "sales_centers": SalesCenter.objects.count(),
            "students": User.objects.filter(is_staff=False).count(),
            "devices": UserDevice.objects.count(),
        },
        "screens": [
            ("الكتالوج العام", "/landing/", "كروت المكثفات قبل تسجيل الدخول."),
            ("صفحة المكثفة العامة", f"/course/{sample_course.id}/" if sample_course else "/course/<id>/", "معلومات المكثفة، الدروس المقفلة، ومعاينة الفيديو."),
            ("تسجيل الدخول", "/login/", "دخول الطالب برقم الهاتف وكلمة المرور."),
            ("إنشاء حساب", "/register/", "اسم الطالب، رقم الهاتف، الفرع، المحافظة."),
            ("لوحة الطالب", "/student/", "تفعيل كود، عرض المكثفات المفتوحة، الفيديوهات، الإشعارات."),
            ("صفحة مكثفة الطالب", f"/student/courses/{sample_course.id}/" if sample_course else "/student/courses/<id>/", "جلسات الفيديو المتاحة بعد التفعيل."),
            ("مشغل الجلسة", "/student/lessons/<id>/", "مشاهدة الفيديو وتسجيل حضور الجلسة."),
            ("لوحة الإدارة الداخلية", "/admin-dashboard/", "ملخص الطلاب والمكثفات والأكواد."),
            ("Django Admin", "/admin/", "إدارة المواد، الدروس، المعاهد، مراكز البيع، ودفعات الأكواد."),
        ],
    }
    return render(request, "dashboard/thaaer.html", context)


def pwa_manifest(_request):
    return JsonResponse(
        {
            "name": "محترفو التعليم",
            "short_name": "محترفون",
            "description": "منصة تعليمية عربية للمكثفات والدروس والامتحانات.",
            "start_url": "/landing/",
            "scope": "/",
            "display": "standalone",
            "orientation": "portrait-primary",
            "dir": "rtl",
            "lang": "ar",
            "background_color": "#f5f5f5",
            "theme_color": "#8f6f3e",
            "categories": ["education", "productivity"],
            "icons": [
                {
                    "src": "/static/dashboard/icons/icon-192.svg",
                    "sizes": "any",
                    "type": "image/svg+xml",
                    "purpose": "any maskable",
                },
                {
                    "src": "/static/dashboard/icons/icon-192.png",
                    "sizes": "192x192",
                    "type": "image/png",
                    "purpose": "any",
                },
                {
                    "src": "/static/dashboard/icons/icon-512.png",
                    "sizes": "512x512",
                    "type": "image/png",
                    "purpose": "any maskable",
                },
            ],
        }
    )


def service_worker(_request):
    script = """
const CACHE_NAME = "mohtarifo-platform-v2";
const APP_SHELL = [
  "/landing/",
  "/login/",
  "/register/",
  "/static/dashboard/styles.css",
  "/static/dashboard/icons/icon.svg",
  "/static/dashboard/icons/icon-192.png",
  "/static/dashboard/icons/icon-512.png"
];

self.addEventListener("install", (event) => {
  event.waitUntil(caches.open(CACHE_NAME).then((cache) => cache.addAll(APP_SHELL)));
  self.skipWaiting();
});

self.addEventListener("activate", (event) => {
  event.waitUntil(
    caches.keys().then((keys) =>
      Promise.all(keys.filter((key) => key !== CACHE_NAME).map((key) => caches.delete(key)))
    )
  );
  self.clients.claim();
});

self.addEventListener("fetch", (event) => {
  if (event.request.method !== "GET") {
    return;
  }

  const url = new URL(event.request.url);
  const isSameOrigin = url.origin === self.location.origin;
  const isSensitiveRoute = isSameOrigin && (
    url.pathname.startsWith("/student/") ||
    url.pathname.startsWith("/api/") ||
    url.pathname === "/device-logged-out/" ||
    url.pathname === "/device-logged-out" ||
    url.pathname === "/login/" ||
    url.pathname === "/logout/"
  );

  if (isSensitiveRoute || event.request.mode === "navigate") {
    event.respondWith(fetch(event.request));
    return;
  }

  event.respondWith(
    fetch(event.request)
      .then((response) => {
        const copy = response.clone();
        caches.open(CACHE_NAME).then((cache) => cache.put(event.request, copy));
        return response;
      })
      .catch(() => caches.match(event.request).then((cached) => cached || caches.match("/landing/")))
  );
});
"""
    response = HttpResponse(script, content_type="application/javascript")
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    return response


def favicon_view(request):
    import os
    from django.conf import settings
    from django.http import FileResponse, Http404
    
    # Try multiple locations in order of priority
    possible_paths = [
        os.path.join(settings.BASE_DIR, 'static', 'images', 'favicon.ico'),
        os.path.join(settings.BASE_DIR, 'static', 'favicon.ico'),
        os.path.join(settings.BASE_DIR, 'static', 'dashboard', 'icons', 'icon-192.ico'),
    ]
    
    for path in possible_paths:
        if os.path.exists(path):
            # Detect actual content type (some files might be PNGs renamed to .ico)
            with open(path, 'rb') as f:
                header = f.read(4)
                content_type = 'image/png' if header == b'\x89PNG' else 'image/x-icon'
            return FileResponse(open(path, 'rb'), content_type=content_type)
    raise Http404()


def favicon_png_view(request):
    import os
    from django.conf import settings
    from django.http import FileResponse, Http404
    # Check our new standard directory
    path = os.path.join(settings.BASE_DIR, 'static', 'images', 'favicon.png')
    if os.path.exists(path):
        return FileResponse(open(path, 'rb'), content_type='image/png')
    # Fallback
    path = os.path.join(settings.BASE_DIR, 'static', 'favicon.png')
    if os.path.exists(path):
        return FileResponse(open(path, 'rb'), content_type='image/png')
    raise Http404()


def robots_txt(request):
    site_url = _site_url(request).rstrip("/")
    lines = [
        "User-agent: *",
        "Allow: /",
        "Allow: /favicon.ico",
        "Allow: /favicon.png",
        "Allow: /static/",
        "",
        "Disallow: /admin/",
        "Disallow: /api/",
        "",
        f"Sitemap: {site_url}/sitemap.xml",
    ]
    # Standard robots.txt prefers a trailing newline and clear separators
    content = "\n".join(lines) + "\n"
    return HttpResponse(content, content_type="text/plain; charset=utf-8")


def sitemap_xml(request):
    site_url = _site_url(request)
    urls = [
        (reverse("dashboard:landing"), "daily", "1.0"),
        (reverse("dashboard:departments_list"), "weekly", "0.8"),
        (reverse("dashboard:instructors_list"), "weekly", "0.8"),
        (reverse("dashboard:about"), "monthly", "0.6"),
        (reverse("dashboard:contact"), "monthly", "0.6"),
        (reverse("dashboard:faq"), "monthly", "0.6"),
        (reverse("dashboard:privacy"), "yearly", "0.3"),
        (reverse("dashboard:terms"), "yearly", "0.3"),
    ]

    courses = Course.objects.filter(status="published").only("id", "updated_at").order_by("-updated_at")
    for course in courses:
        urls.append((reverse("dashboard:public_course_detail", args=[course.id]), "weekly", "0.9", course.updated_at))

    instructors = (
        User.objects.filter(courses__status="published")
        .distinct()
        .only("id")
        .order_by("id")
    )
    for instructor in instructors:
        urls.append((reverse("dashboard:landing") + f"?instructor={instructor.id}", "weekly", "0.7"))

    items = []
    for entry in urls:
        path, changefreq, priority, *lastmod = entry
        lastmod_tag = ""
        if lastmod and lastmod[0]:
            lastmod_tag = f"<lastmod>{lastmod[0].date().isoformat()}</lastmod>"
        items.append(
            "<url>"
            f"<loc>{escape(site_url + path)}</loc>"
            f"{lastmod_tag}"
            f"<changefreq>{changefreq}</changefreq>"
            f"<priority>{priority}</priority>"
            "</url>"
        )

    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">'
        + "".join(items)
        + "</urlset>"
    )
    return HttpResponse(xml, content_type="application/xml; charset=utf-8")


def _client_ip(request):
    forwarded_for = request.META.get("HTTP_X_FORWARDED_FOR", "")
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR", "")


def _login_attempt_key(request):
    username = (request.POST.get("username") or "").strip().lower()
    return f"login_attempts:{_client_ip(request)}:{username}"


def _login_is_rate_limited(request):
    return int(cache.get(_login_attempt_key(request), 0)) >= settings.LOGIN_RATE_LIMIT_ATTEMPTS


def _record_failed_login(request):
    key = _login_attempt_key(request)
    try:
        cache.incr(key)
    except ValueError:
        cache.set(key, 1, settings.LOGIN_RATE_LIMIT_TIMEOUT_SECONDS)
    else:
        cache.touch(key, settings.LOGIN_RATE_LIMIT_TIMEOUT_SECONDS)


def _get_user_phone(user):
    from accounts.auth_utils import get_instructor_login_phone

    return get_instructor_login_phone(user)


def login_view(request):
    from analytics.services import TrackingService
    TrackingService.log_landing_visit(request)
    
    if request.user.is_authenticated:
        return redirect("dashboard:home")

    login_error = None
    login_username_value = ""

    if request.method == "POST":
        raw_username = request.POST.get("username", "").strip()
        raw_password = request.POST.get("password", "")
        login_username_value = raw_username

        from accounts.auth_utils import is_instructor_account, resolve_user_for_login, verify_instructor_password

        resolved_user = resolve_user_for_login(raw_username)

        if resolved_user and is_instructor_account(resolved_user):
            if _login_is_rate_limited(request):
                messages.error(request, "لقد تجاوزت الحد المسموح به من محاولات تسجيل الدخول الفاشلة. يرجى المحاولة لاحقاً.")
                return render(request, "registration/login.html", {"form": AuthenticationForm()}, status=429)

            if verify_instructor_password(resolved_user, raw_password):
                from .whatsapp_utils import is_2fa_disabled
                if is_2fa_disabled():
                    login(request, resolved_user)
                    return redirect("dashboard:home")

                phone = _get_user_phone(resolved_user)
                request.session["otp_login_user_id"] = resolved_user.id
                request.session["otp_login_phone"] = phone
                otp_result = send_otp(phone, purpose="login", request=request)
                if not otp_result["success"]:
                    messages.error(request, otp_result["message"])
                    return render(request, "registration/login.html", {"form": AuthenticationForm()})
                cache.delete(_login_attempt_key(request))
                return redirect("dashboard:verify_login_otp")

            _record_failed_login(request)
            login_error = "كلمة المرور غير صحيحة. جرّب رقم هاتفك أو تواصل مع الإدارة لإعادة التعيين."
            return render(
                request,
                "registration/login.html",
                {"form": AuthenticationForm(), "login_error": login_error, "login_username": raw_username},
            )

        if resolved_user:
            post_data = request.POST.copy()
            post_data["username"] = resolved_user.username
            form = AuthenticationForm(request, data=post_data)
        else:
            form = AuthenticationForm(request, data=request.POST or None)
    else:
        form = AuthenticationForm(request, data=None)
    if request.method == "POST" and _login_is_rate_limited(request):
        messages.error(request, "لقد تجاوزت الحد المسموح به من محاولات تسجيل الدخول الفاشلة. يرجى المحاولة لاحقاً.")
        return render(request, "registration/login.html", {"form": form}, status=429)

    if request.method == "POST" and form.is_valid():
        user = form.get_user()
        
        # Resolve the phone number associated with the user
        phone = _get_user_phone(user)
        
        from .whatsapp_utils import is_2fa_disabled
        if is_2fa_disabled():
            login(request, user)
            response = redirect("dashboard:home")
            if not user.is_staff:
                activate_user_device(request, user, response)
            return response

        request.session["otp_login_user_id"] = user.id
        request.session["otp_login_phone"] = phone
        
        # Send OTP
        otp_result = send_otp(phone, purpose="login", request=request)
        if not otp_result["success"]:
            messages.error(request, otp_result["message"])
            return render(request, "registration/login.html", {"form": form})
        
        cache.delete(_login_attempt_key(request))
        return redirect("dashboard:verify_login_otp")
    
    if request.method == "POST":
        _record_failed_login(request)
        if not login_error:
            if not resolve_user_for_login(request.POST.get("username", "").strip()):
                login_error = "لا يوجد حساب بهذا الاسم أو الرقم."
            else:
                login_error = "كلمة المرور غير صحيحة."

    return render(
        request,
        "registration/login.html",
        {
            "form": form,
            "login_error": login_error,
            "login_username": login_username_value,
        },
    )


def verify_login_otp(request):
    """OTP verification page for login."""
    user_id = request.session.get("otp_login_user_id")
    phone = request.session.get("otp_login_phone")
    
    if not user_id or not phone:
        return redirect("dashboard:login")
    
    error_message = None
    success_message = None
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "resend_otp":
            otp_result = send_otp(phone, purpose="login", request=request)
            if otp_result["success"]:
                success_message = otp_result["message"]
            else:
                error_message = otp_result["message"]
        
        elif action == "verify_otp":
            submitted_code = request.POST.get("otp_code", "").strip()
            result = verify_otp(phone, submitted_code, purpose="login")
            
            if result["valid"]:
                # OTP verified - log user in
                user = User.objects.filter(id=user_id).first()
                if user:
                    login(request, user)
                    # Clean up session
                    request.session.pop("otp_login_user_id", None)
                    request.session.pop("otp_login_phone", None)
                    response = redirect("dashboard:home")
                    if not user.is_staff:
                        activate_user_device(request, user, response)
                    return response
                else:
                    error_message = "حدث خطأ. حاول تسجيل الدخول مجدداً."
                    return redirect("dashboard:login")
            else:
                error_message = result["message"]
    
    # Mask phone for display (show first 3 and last 2)
    phone_display = phone[:3] + "•" * (len(phone) - 5) + phone[-2:] if len(phone) > 5 else phone
    
    return render(request, "registration/verify_otp.html", {
        "phone_display": phone_display,
        "error_message": error_message,
        "success_message": success_message,
        "otp_expiry_seconds": OTP_EXPIRY_SECONDS,
        "resend_cooldown": OTP_RESEND_COOLDOWN_SECONDS,
        "otp_length": OTP_LENGTH,
        "back_url": "/login/",
    })


def register_view(request):
    from analytics.services import TrackingService
    TrackingService.log_landing_visit(request)
    
    if request.user.is_authenticated:
        return redirect("dashboard:home")

    form = StudentRegistrationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        # Store form data in session and send OTP
        phone = form.cleaned_data["username"]
        reg_data = {
            "first_name": form.cleaned_data["first_name"],
            "username": phone,
            "track": form.cleaned_data["track"],
            "governorate": form.cleaned_data["governorate"],
            "gender": form.cleaned_data["gender"],
            "email": form.cleaned_data.get("email", ""),
            "password1": form.cleaned_data["password1"],
        }
        request.session["otp_register_data"] = reg_data
        
        from .whatsapp_utils import is_2fa_disabled
        if is_2fa_disabled():
            from django.db import transaction
            import threading
            from .whatsapp_utils import parse_gender_grammar, send_whatsapp_message
            from django.contrib.auth.models import User
            from accounts.models import StudentProfile
            
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=reg_data["username"],
                        password=reg_data["password1"],
                        first_name=reg_data["first_name"],
                        last_name="",
                        email=reg_data.get("email", ""),
                    )
                    student_gender = reg_data["gender"]
                    StudentProfile.objects.update_or_create(
                        user=user,
                        defaults={
                            "grade": "الثالث الثانوي",
                            "track": reg_data["track"],
                            "governorate": reg_data["governorate"],
                            "phone": reg_data["username"],
                            "gender": student_gender,
                        },
                    )
                    
                    # Send WhatsApp welcome message
                    student_phone = reg_data["username"]
                    student_name = user.get_full_name() or user.username
                    
                    welcome_template = (
                        f"مرحباً بك {student_name} في منصة محترفو التعليم! ✨🎓\n\n"
                        "يسعدنا جداً {انضمامك|انضمامكِ} إلى عائلتنا التعليمية المتميزة. تم إنشاء حسابك الشخصي بنجاح وأصبح جاهزاً تماماً للاستخدام. 🚀🌟\n\n"
                        "نتمنى لك مسيرة مليئة بالتفوق والإنجاز، ونؤكد لك بأن كادر المنصة دائماً {بجانبك|بجانبكِ} ويسعى لتقديم أفضل تجربة تعليمية تليق بطموحاتك العالية. 💼💫"
                    )
                    
                    welcome_text = parse_gender_grammar(welcome_template, student_gender)
                    
                    transaction.on_commit(lambda: threading.Thread(
                        target=send_whatsapp_message,
                        args=(student_phone, welcome_text),
                        daemon=True
                    ).start())
                
                # Clean up session
                request.session.pop("otp_register_data", None)
                
                login(request, user)
                messages.success(request, "تم إنشاء حسابك بنجاح. يمكنك الآن إضافة كود المادة.")
                response = redirect("dashboard:student_dashboard")
                activate_user_device(request, user, response)
                return response
            except Exception as e:
                messages.error(request, f"حدث خطأ أثناء إنشاء الحساب: {str(e)}")
                return render(request, "registration/register.html", {"form": form})

        # Send OTP
        otp_result = send_otp(phone, purpose="register", request=request)
        if not otp_result["success"]:
            messages.error(request, otp_result["message"])
            return render(request, "registration/register.html", {"form": form})
        
        return redirect("dashboard:verify_register_otp")

    return render(request, "registration/register.html", {"form": form})


def verify_register_otp(request):
    """OTP verification page for registration."""
    reg_data = request.session.get("otp_register_data")
    
    if not reg_data:
        return redirect("dashboard:register")
    
    phone = reg_data["username"]
    error_message = None
    success_message = None
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "resend_otp":
            otp_result = send_otp(phone, purpose="register", request=request)
            if otp_result["success"]:
                success_message = otp_result["message"]
            else:
                error_message = otp_result["message"]
        
        elif action == "verify_otp":
            submitted_code = request.POST.get("otp_code", "").strip()
            result = verify_otp(phone, submitted_code, purpose="register")
            
            if result["valid"]:
                # OTP verified - create the account
                from django.db import transaction
                
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=reg_data["username"],
                        password=reg_data["password1"],
                        first_name=reg_data["first_name"],
                        last_name="",
                        email=reg_data.get("email", ""),
                    )
                    student_gender = reg_data["gender"]
                    StudentProfile.objects.update_or_create(
                        user=user,
                        defaults={
                            "grade": "الثالث الثانوي",
                            "track": reg_data["track"],
                            "governorate": reg_data["governorate"],
                            "phone": reg_data["username"],
                            "gender": student_gender,
                        },
                    )
                    
                    # Send WhatsApp welcome message
                    student_phone = reg_data["username"]
                    student_name = user.get_full_name() or user.username
                    
                    welcome_template = (
                        f"مرحباً بك {student_name} في منصة محترفو التعليم! ✨🎓\n\n"
                        "يسعدنا جداً {انضمامك|انضمامكِ} إلى عائلتنا التعليمية المتميزة. تم إنشاء حسابك الشخصي بنجاح وأصبح جاهزاً تماماً للاستخدام. 🚀🌟\n\n"
                        "نتمنى لك مسيرة مليئة بالتفوق والإنجاز، ونؤكد لك بأن كادر المنصة دائماً {بجانبك|بجانبكِ} ويسعى لتقديم أفضل تجربة تعليمية تليق بطموحاتك العالية. 💼💫"
                    )
                    
                    welcome_text = parse_gender_grammar(welcome_template, student_gender)
                    
                    transaction.on_commit(lambda: threading.Thread(
                        target=send_whatsapp_message,
                        args=(student_phone, welcome_text),
                        daemon=True
                    ).start())
                
                # Clean up session
                request.session.pop("otp_register_data", None)
                
                login(request, user)
                messages.success(request, "تم إنشاء حسابك بنجاح. يمكنك الآن إضافة كود المادة.")
                response = redirect("dashboard:student_dashboard")
                activate_user_device(request, user, response)
                return response
            else:
                error_message = result["message"]
    
    # Mask phone for display
    phone_display = phone[:3] + "•" * (len(phone) - 5) + phone[-2:] if len(phone) > 5 else phone
    
    return render(request, "registration/verify_otp.html", {
        "phone_display": phone_display,
        "error_message": error_message,
        "success_message": success_message,
        "otp_expiry_seconds": OTP_EXPIRY_SECONDS,
        "resend_cooldown": OTP_RESEND_COOLDOWN_SECONDS,
        "otp_length": OTP_LENGTH,
        "back_url": "/register/",
    })


def password_reset_request_view(request):
    """View to start the password reset process by providing a phone number."""
    if request.user.is_authenticated:
        return redirect("dashboard:home")
        
    form = PasswordResetRequestForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        phone = form.cleaned_data["phone"]
        
        # Verify user exists
        user = User.objects.filter(username=phone).first()
        if not user:
            # For security, we don't always want to confirm if a user exists, 
            # but in this student platform, we can show a helpful error.
            messages.error(request, "هذا الرقم غير مسجل لدينا.")
            return render(request, "registration/password_reset_request.html", {"form": form})
        
        # Store in session and send OTP
        request.session["otp_reset_phone"] = phone
        otp_result = send_otp(phone, purpose="reset_password", request=request)
        
        if otp_result["success"]:
            return redirect("dashboard:password_reset_verify")
        else:
            messages.error(request, otp_result["message"])
            
    return render(request, "registration/password_reset_request.html", {"form": form})


def password_reset_verify_view(request):
    """OTP verification page for password reset."""
    phone = request.session.get("otp_reset_phone")
    if not phone:
        return redirect("dashboard:password_reset_request")
    
    error_message = None
    success_message = None
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "resend_otp":
            otp_result = send_otp(phone, purpose="reset_password", request=request)
            if otp_result["success"]:
                success_message = otp_result["message"]
            else:
                error_message = otp_result["message"]
        
        elif action == "verify_otp":
            submitted_code = request.POST.get("otp_code", "").strip()
            result = verify_otp(phone, submitted_code, purpose="reset_password")
            
            if result["valid"]:
                # OTP verified - allow password change
                request.session["otp_reset_verified"] = True
                return redirect("dashboard:password_reset_complete")
            else:
                error_message = result["message"]
    
    # Mask phone for display
    phone_display = phone[:3] + "•" * (len(phone) - 5) + phone[-2:] if len(phone) > 5 else phone
    
    return render(request, "registration/verify_otp.html", {
        "phone_display": phone_display,
        "error_message": error_message,
        "success_message": success_message,
        "otp_expiry_seconds": OTP_EXPIRY_SECONDS,
        "resend_cooldown": OTP_RESEND_COOLDOWN_SECONDS,
        "otp_length": OTP_LENGTH,
        "back_url": "/login/", # Back to login or reset request
    })


def password_reset_complete_view(request):
    """Final stage of password reset: enter new password."""
    phone = request.session.get("otp_reset_phone")
    verified = request.session.get("otp_reset_verified")
    
    if not phone or not verified:
        messages.error(request, "الرجاء التحقق من الرمز أولاً.")
        return redirect("dashboard:password_reset_request")
        
    form = SetNewPasswordForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = User.objects.filter(username=phone).first()
        if user:
            user.set_password(form.cleaned_data["password1"])
            user.save()
            
            # Send Security Notification via WhatsApp
            ua = request.META.get('HTTP_USER_AGENT', '')
            device = "جهاز غير معروف"
            if "Windows" in ua: device = "Windows PC"
            elif "Android" in ua: device = "Android Phone"
            elif "iPhone" in ua: device = "iPhone"
            elif "iPad" in ua: device = "iPad"
            elif "Macintosh" in ua: device = "MacBook/iMac"
            
            browser = ""
            if "Edg/" in ua: browser = "Edge"
            elif "Chrome/" in ua: browser = "Chrome"
            elif "Firefox/" in ua: browser = "Firefox"
            elif "Safari/" in ua: browser = "Safari"
            
            device_full = f"{device} ({browser})" if browser else device
            timestamp = timezone.now().strftime("%Y-%m-%d %H:%M")
            
            security_msg = (
                f"🛡️ إشعار أمان: تم تغيير كلمة مرور حسابك بنجاح.\n\n"
                f"📱 الجهاز: *{device_full}*\n"
                f"⏰ الوقت: *{timestamp}*\n\n"
                f"⚠️ إذا لم تقم بهذا الإجراء، يرجى التواصل مع الدعم الفني فوراً لحماية حسابك."
            )
            
            threading.Thread(
                target=send_whatsapp_message,
                args=(phone, security_msg),
                daemon=True
            ).start()
            
            # Clean up session
            request.session.pop("otp_reset_phone", None)
            request.session.pop("otp_reset_verified", None)
            
            messages.success(request, "تم تغيير كلمة المرور بنجاح. يمكنك الآن تسجيل الدخول.")
            return redirect("dashboard:login")
        else:
            messages.error(request, "حدث خطأ غير متوقع.")
            return redirect("dashboard:password_reset_request")
            
    return render(request, "registration/password_reset_complete.html", {"form": form})


@admin_required
def admin_dashboard(request):
    from analytics.services import TrackingService
    TrackingService.log_landing_visit(request)

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "clear_system_logs":
            import os
            for f in ["landing_error_log.txt", "profile_error_log.txt"]:
                if os.path.exists(f):
                    os.remove(f)
            messages.success(request, "تم مسح سجلات الأخطاء بنجاح.")
            return redirect("dashboard:admin_dashboard")
        elif action == "toggle_2fa":
            from .whatsapp_utils import set_2fa_disabled
            disable_otp = request.POST.get("disable_2fa") == "true"
            success = set_2fa_disabled(disable_otp)
            if success:
                if disable_otp:
                    messages.success(request, "🔓 تم إيقاف التحقق الثنائي بنجاح. يمكن للطلاب تسجيل الدخول وإنشاء حسابات دون إرسال رمز تحقق.")
                else:
                    messages.success(request, "🔒 تم تفعيل التحقق الثنائي بنجاح. سيتم طلب رمز تحقق عبر واتساب عند تسجيل الدخول أو إنشاء حساب.")
            else:
                messages.error(request, "❌ فشل حفظ الإعدادات. يرجى التحقق من صلاحيات المجلدات على الخادم.")
            return redirect("dashboard:admin_dashboard")

    
    from analytics.models import LandingVisit
    from django.db.models import Count
    
    # Fetch analytical data for the overview
    total_visits = LandingVisit.objects.count()
    print("DEBUG ADMIN DASHBOARD - TOTAL VISITS COMPUTED AS:", total_visits)
    logged_in_visits = LandingVisit.objects.filter(user__isnull=False).count()
    anon_visits = total_visits - logged_in_visits
    
    # Aggregate by device types
    device_counts = LandingVisit.objects.values('device_type').annotate(total=Count('id')).order_by('-total')
    
    # Get last few specific actions to view
    recent_traffic = LandingVisit.objects.select_related('user').order_by('-visited_at')[:10]
    courses = (
        Course.objects.select_related("subject", "instructor", "instructor__instructor_profile")
        .annotate(
            lessons_total=Count("units__lessons", distinct=True),
            codes_total=Count("access_codes", distinct=True),
            grants_total=Count("access_grants", distinct=True),
        )
        .order_by("-published_at", "title")[:12]
    )
    recent_batches = AccessCodeBatch.objects.select_related("course", "institute", "sales_center").order_by("-created_at")[:8]
    sales_centers = SalesCenter.objects.select_related("institute").filter(is_active=True).order_by("name")[:8]
    # Run system diagnostics
    from .diagnostics import SystemDiagnostics
    health_checks = SystemDiagnostics.run_all_checks()

    # Fetch system logs for debugging

    system_logs = []
    try:
        import os
        log_files = [
            ("Tracking Errors", "landing_error_log.txt"),
            ("Profile Errors", "profile_error_log.txt"),
        ]
        for label, filename in log_files:
            if os.path.exists(filename):
                with open(filename, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                    if content:
                        system_logs.append({"label": label, "content": content[-2000:]}) # Last 2000 chars
    except Exception:
        pass

    context = {
        "instructor_profile": getattr(request.user, 'instructor_profile', None),
        "students_count": User.objects.filter(is_staff=False).count(),
        "courses_count": Course.objects.count(),
        "lessons_count": Lesson.objects.count(),
        "questions_count": Question.objects.count(),
        "exams_count": Exam.objects.count(),
        "attempts_count": Attempt.objects.count(),
        "codes_count": AccessCode.objects.count(),
        "redeemed_codes_count": AccessCode.objects.filter(redeemed_count__gt=0).count(),
        "sold_codes_count": AccessCode.objects.filter(sale_status="sold").count(),
        "free_codes_count": AccessCode.objects.filter(is_free_code=True).count(),
        "institutes_count": Institute.objects.count(),
        "sales_centers_count": SalesCenter.objects.count(),
        "active_subscriptions": Subscription.objects.filter(status="active").count(),
        "paid_payments": Payment.objects.filter(status="paid").count(),
        "managed_courses": courses,
        "courses": courses,
        "recent_batches": recent_batches,
        "sales_centers": sales_centers,
        "traffic_stats": {
            "total": total_visits,
            "logged_in": logged_in_visits,
            "anon": anon_visits,
            "devices": device_counts,
            "recent": recent_traffic
        },
        "system_logs": system_logs,
        "health_checks": health_checks,
        "otp_2fa_disabled": is_2fa_disabled()
    }


    return render(request, "dashboard/admin_dashboard.html", context)


@admin_required
def admin_course_create(request):
    initial = {}
    if request.method == "GET":
        initial["kind"] = request.GET.get("kind") or None
        initial["academic_track"] = request.GET.get("track") or None
        initial = {key: value for key, value in initial.items() if value}
    form = CourseCreateForm(request.POST or None, request.FILES or None, initial=initial)
    if request.method == "POST" and form.is_valid():
        course = form.save()
        messages.success(request, "تم إنشاء المحتوى. الآن كمل الجلسات والأكواد والتفاصيل من لوحة الدورة.")
        return redirect("dashboard:admin_course_control", course_id=course.id)
    selected_tab = None
    for tab in _catalog_tabs():
        if tab["kind"] == initial.get("kind") and tab["track"] == initial.get("academic_track"):
            selected_tab = tab
            break
    return render(request, "dashboard/admin_course_create.html", {"form": form, "selected_tab": selected_tab})


@admin_required
def admin_catalog_manager(request):
    if request.method == "POST":
        action = request.POST.get("action")
        section_id = request.POST.get("section_id")
        section = CatalogSection.objects.filter(id=section_id).first() if section_id else None
        if action in {"create_section", "update_section"}:
            form = CatalogSectionForm(
                request.POST,
                instance=section,
                prefix="new" if action == "create_section" else "section",
            )
            if form.is_valid():
                saved = form.save()
                messages.success(request, f"تم حفظ فلتر {saved.label}.")
                return redirect(f"{reverse('dashboard:admin_catalog_manager')}?kind={saved.kind}&track={saved.track}")
        elif action == "delete_section" and section:
            courses_count = Course.objects.filter(kind=section.kind, academic_track=section.track).count()
            label = section.label
            if courses_count:
                section.is_visible = False
                section.save(update_fields=["is_visible", "updated_at"])
                messages.warning(request, f"تم إخفاء فلتر {label} من واجهة الطالب لأنه يحتوي على دورات.")
            else:
                section.delete()
                messages.success(request, f"تم حذف فلتر {label}.")
            return redirect("dashboard:admin_catalog_manager")

    base_tabs = _catalog_tabs(include_hidden=True)
    selected_kind = request.GET.get("kind") or (base_tabs[0]["kind"] if base_tabs else "")
    selected_track = request.GET.get("track") or (base_tabs[0]["track"] if base_tabs else "")
    tabs = []
    for tab in base_tabs:
        if tab["track"] in {"scientific", "literary"}:
            courses_qs = Course.objects.filter(kind=tab["kind"]).filter(
                models.Q(academic_track=tab["track"]) | models.Q(academic_track="general")
            )
        else:
            courses_qs = Course.objects.filter(kind=tab["kind"], academic_track=tab["track"])
            
        tab = {
            **tab,
            "courses_count": courses_qs.count(),
            "published_count": courses_qs.filter(status="published").count(),
            "draft_count": courses_qs.filter(status="draft").count(),
            "is_active": tab["kind"] == selected_kind and tab["track"] == selected_track,
        }
        tabs.append(tab)

    selected_tab = next((tab for tab in tabs if tab["is_active"]), tabs[0] if tabs else None)
    selected_section = CatalogSection.objects.filter(id=selected_tab.get("id")).first() if selected_tab else None
    section_form = CatalogSectionForm(instance=selected_section, prefix="section")
    new_section_form = CatalogSectionForm(prefix="new", initial={"sort_order": (CatalogSection.objects.count() + 1) * 10})
    courses = Course.objects.none()
    if selected_tab:
        if selected_tab["track"] in {"scientific", "literary"}:
            courses = (
                Course.objects.filter(kind=selected_tab["kind"]).filter(
                    models.Q(academic_track=selected_tab["track"]) | models.Q(academic_track="general")
                )
                .select_related("subject", "instructor", "instructor__instructor_profile")
                .annotate(
                    lessons_total=Count("units__lessons", distinct=True),
                    codes_total=Count("access_codes", distinct=True),
                    students_total=Count("access_grants", distinct=True),
                )
                .order_by("subject__name", "title")
            )
        else:
            courses = (
                Course.objects.filter(kind=selected_tab["kind"], academic_track=selected_tab["track"])
                .select_related("subject", "instructor", "instructor__instructor_profile")
                .annotate(
                    lessons_total=Count("units__lessons", distinct=True),
                    codes_total=Count("access_codes", distinct=True),
                    students_total=Count("access_grants", distinct=True),
                )
                .order_by("subject__name", "title")
            )
    return render(
        request,
        "dashboard/admin_catalog_manager.html",
        {
            "catalog_tabs": tabs,
            "selected_tab": selected_tab,
            "section_form": section_form,
            "new_section_form": new_section_form,
            "courses": courses,
        },
    )


@admin_required
def admin_partners(request):
    institute_form = InstituteForm(prefix="institute")
    center_form = SalesCenterForm(prefix="center")
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_institute":
            institute_form = InstituteForm(request.POST, prefix="institute")
            if institute_form.is_valid():
                institute_form.save()
                messages.success(request, "تم إنشاء المعهد.")
                return redirect("dashboard:admin_partners")
        elif action == "create_center":
            center_form = SalesCenterForm(request.POST, prefix="center")
            if center_form.is_valid():
                center_form.save()
                messages.success(request, "تم إنشاء مركز البيع.")
                return redirect("dashboard:admin_partners")

    institutes = (
        Institute.objects.annotate(
            centers_total=Count("sales_centers", distinct=True),
            codes_total=Count("access_codes", distinct=True),
            activated_total=Count("access_codes__grants", distinct=True),
        )
        .order_by("name")
    )
    centers = (
        SalesCenter.objects.select_related("institute")
        .annotate(codes_total=Count("access_codes", distinct=True), activated_total=Count("access_codes__grants", distinct=True))
        .order_by("name")
    )
    return render(
        request,
        "dashboard/admin_partners.html",
        {
            "institute_form": institute_form,
            "center_form": center_form,
            "institutes": institutes,
            "centers": centers,
        },
    )


@admin_required
def admin_students(request):
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', 'all')
    
    students_query = User.objects.filter(is_staff=False).select_related('student_profile')
    
    if query:
        students_query = students_query.filter(
            models.Q(username__icontains=query) |
            models.Q(first_name__icontains=query) |
            models.Q(last_name__icontains=query) |
            models.Q(student_profile__phone__icontains=query)
        )
        
    if status_filter == 'active':
        students_query = students_query.filter(is_active=True)
    elif status_filter == 'inactive':
        students_query = students_query.filter(is_active=False)
        
    students_query = students_query.annotate(
        courses_total=Count("access_grants", distinct=True), 
        devices_total=Count("devices", distinct=True)
    ).order_by("-date_joined")
    
    # Pagination
    paginator = Paginator(students_query, 50)
    page_number = request.GET.get('page')
    students = paginator.get_page(page_number)
    
    # Stats
    total_students = User.objects.filter(is_staff=False).count()
    active_students = User.objects.filter(is_staff=False, is_active=True).count()
    new_students_today = User.objects.filter(is_staff=False, date_joined__date=timezone.now().date()).count()
    total_devices = UserDevice.objects.count()

    grants = AccessGrant.objects.select_related("user", "course", "access_code").order_by("-created_at")[:20]
    devices = UserDevice.objects.select_related("user").order_by("-last_seen_at")[:20]

    # Governorate statistics
    gov_stats = []
    from accounts.models import StudentProfile
    profiles_grouped = StudentProfile.objects.values('governorate').annotate(count=Count('id')).order_by('-count')
    for p in profiles_grouped:
        gov_name = p['governorate'].strip() if p['governorate'] else ""
        if not gov_name:
            gov_name = "غير محدد"
        # Combine if "غير محدد" appears multiple times due to spaces
        existing = next((item for item in gov_stats if item['name'] == gov_name), None)
        if existing:
            existing['count'] += p['count']
        else:
            gov_stats.append({
                'name': gov_name,
                'count': p['count']
            })
    # Sort again by count descending
    gov_stats.sort(key=lambda x: x['count'], reverse=True)

    return render(
        request,
        "dashboard/admin_students.html",
        {
            "students": students,
            "grants": grants,
            "devices": devices,
            "query": query,
            "status_filter": status_filter,
            "stats": {
                "total": total_students,
                "active": active_students,
                "new_today": new_students_today,
                "devices": total_devices,
            },
            "gov_stats": gov_stats,
        },
    )


@admin_required
def admin_institute_profile(request, institute_id):
    institute = get_object_or_404(Institute, id=institute_id)
    import_form = PartnerInstituteImportForm(prefix="import")
    edit_form = InstituteForm(instance=institute, prefix="edit")
    
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "import_institute_students":
            import_form = PartnerInstituteImportForm(request.POST, request.FILES, prefix="import")
            if import_form.is_valid():
                batch = AccessCodeBatch.objects.create(
                    name=import_form.cleaned_data["batch_name"],
                    course=import_form.cleaned_data["course"],
                    institute=institute,
                    code_prefix=import_form.cleaned_data["code_prefix"],
                    notes=f"طلاب مجانيون من {institute.name}",
                )
                created = create_codes_from_upload(batch, import_form.cleaned_data["file"], free_codes=True)
                messages.success(request, f"تم رفع {created} طالب وإنشاء أكواد مجانية مرتبطة بالمعهد.")
                return redirect("dashboard:admin_institute_profile", institute_id=institute.id)
        
        elif action == "edit_institute":
            edit_form = InstituteForm(request.POST, request.FILES, instance=institute, prefix="edit")
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "تم تحديث بيانات المعهد والشعار بنجاح.")
                return redirect("dashboard:admin_institute_profile", institute_id=institute.id)

    batches = (
        AccessCodeBatch.objects.filter(institute=institute)
        .select_related("course")
        .annotate(codes_total=Count("codes", distinct=True), activated_total=Count("codes__grants", distinct=True))
        .order_by("-created_at")
    )
    return render(
        request,
        "dashboard/admin_institute_profile.html",
        {"institute": institute, "import_form": import_form, "edit_form": edit_form, "batches": batches},
    )


@admin_required
def admin_sales_center_profile(request, center_id):
    center = get_object_or_404(SalesCenter.objects.select_related("institute"), id=center_id)
    batch_form = PartnerBatchForm(prefix="batch")
    edit_form = SalesCenterForm(instance=center, prefix="edit")
    
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_center_codes":
            batch_form = PartnerBatchForm(request.POST, prefix="batch")
            if batch_form.is_valid():
                batch = AccessCodeBatch.objects.create(
                    name=batch_form.cleaned_data["name"],
                    course=batch_form.cleaned_data["course"],
                    sales_center=center,
                    allocated_count=batch_form.cleaned_data["quantity"],
                    code_prefix=batch_form.cleaned_data["code_prefix"],
                    notes=batch_form.cleaned_data["notes"],
                )
                created = create_codes_for_batch(batch, batch_form.cleaned_data["quantity"], free_codes=False)
                messages.success(request, f"تم إنشاء {created} كود لمركز البيع.")
                return redirect("dashboard:admin_sales_center_profile", center_id=center.id)
        
        elif action == "edit_center":
            edit_form = SalesCenterForm(request.POST, instance=center, prefix="edit")
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "تم تحديث بيانات مركز البيع بنجاح.")
                return redirect("dashboard:admin_sales_center_profile", center_id=center.id)

    batches = (
        AccessCodeBatch.objects.filter(sales_center=center)
        .select_related("course")
        .annotate(codes_total=Count("codes", distinct=True), sold_total=Count("codes", filter=models.Q(codes__sale_status="sold"), distinct=True), activated_total=Count("codes__grants", distinct=True))
        .order_by("-created_at")
    )
    # Fetch Accounting Data
    wallet = Wallet.objects.filter(sales_center=center).first()
    recent_transactions = []
    if wallet:
        recent_transactions = WalletTransaction.objects.filter(wallet=wallet).order_by('-created_at')[:20]

    return render(
        request,
        "dashboard/admin_sales_center_profile.html",
        {
            "center": center, 
            "batch_form": batch_form, 
            "edit_form": edit_form, 
            "batches": batches,
            "wallet": wallet,
            "recent_transactions": recent_transactions
        },
    )



@admin_required
def admin_print_batch_cards(request, batch_id):
    """توليد صفحة قابلة للطباعة تحتوي على كروت الأكواد (وش وقفا)"""
    batch = get_object_or_404(
        AccessCodeBatch.objects.select_related("course", "package", "institute", "sales_center"), 
        id=batch_id
    )
    codes = batch.codes.all().order_by("created_at")
    target_title = batch.target_title
    card_course_title = (target_title or "").split(" - ", 1)[0].strip() or target_title
    
    # تحضير رابط المنصة لـ QR الخلفية
    platform_url = request.build_absolute_uri("/")
    platform_qr = _generate_qr_base64(platform_url)
    
    # تحضير الكروت مع الـ QR الخاص بكل كود
    cards = []
    for c in codes:
        cards.append({
            "obj": c,
            "qr": _generate_qr_base64(c.code)
        })
    AccessCodePrintLog.objects.create(
        batch=batch,
        printed_by=request.user if request.user.is_authenticated else None,
        cards_count=len(cards),
        notes="فتح صفحة طباعة الكروت",
    )
    
    # تقسيم الكروت لمجموعات (كل مجموعة 3 كروت لصفحة A4 واحدة)
    grouped_cards = [cards[i:i + 3] for i in range(0, len(cards), 3)]
        
    return render(request, "dashboard/admin_print_batch_cards.html", {
        "batch": batch,
        "card_course_title": card_course_title,
        "grouped_cards": grouped_cards,
        "platform_qr": platform_qr,
        "platform_name": "تركيز", # يمكن تغييرها حسب الرغبة
        "print_date": timezone.now(),
    })


def _generate_qr_base64(data):
    """دالة مساعدة لتوليد QR Code وتحويله لـ Base64 لدمجه في HTML"""
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffered = BytesIO()
    img.save(buffered, format="PNG")
    return base64.b64encode(buffered.getvalue()).decode()


@admin_required
def download_batch_codes(request, batch_id):
    batch = get_object_or_404(AccessCodeBatch.objects.select_related("course", "package", "institute", "sales_center"), id=batch_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "codes"
    sheet.append(["code", "target", "type", "partner", "status", "student_name", "student_phone", "activated"])
    partner = batch.institute.name if batch.institute else batch.sales_center.name if batch.sales_center else ""
    for code in batch.codes.order_by("created_at"):
        sheet.append([
            code.code,
            batch.target_title,
            code.get_access_type_display(),
            partner,
            code.sale_status,
            code.assigned_student_name,
            code.assigned_student_phone,
            "yes" if code.redeemed_count else "no",
        ])
    for column in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        sheet.column_dimensions[column].width = 24
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="codes-batch-{batch.id}.xlsx"'
    return response


@admin_required
def download_student_import_template(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "students"
    headers = ["student_name", "student_phone", "notes"]
    sheet.append(headers)
    sheet.append(["مثال: أحمد محمد", "09xxxxxxxx", "طلاب معهد اليمان"])
    for cell in sheet[1]:
        cell.style = "Headline 3"
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 18
    sheet.column_dimensions["C"].width = 34
    info = workbook.create_sheet("instructions")
    info.append(["استخدم الأعمدة كما هي بدون تغيير أسماء العناوين."])
    info.append(["student_name", "اسم الطالب كما سيظهر في السجل"])
    info.append(["student_phone", "رقم الطالب ويستخدم لربط الكود بحسابه"])
    info.append(["notes", "ملاحظات اختيارية"])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    filename = f"students-template-course-{course.id}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@admin_required
def admin_export_students(request):
    """تصدير كل بيانات الطلاب إلى ملف Excel للتواصل أو التسويق"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Students"
    sheet.append(["الاسم الكامل", "اسم المستخدم/الهاتف", "الفرع", "المحافظة", "تاريخ الانضمام", "آخر نشاط", "XP", "المستوى"])
    
    students = User.objects.filter(is_staff=False).select_related("student_profile").order_by("-date_joined")
    
    for s in students:
        profile = getattr(s, 'student_profile', None)
        sheet.append([
            s.get_full_name(),
            s.username,
            profile.track if profile else "",
            profile.governorate if profile else "",
            s.date_joined.strftime("%Y-%m-%d"),
            profile.last_activity_date.strftime("%Y-%m-%d") if profile and profile.last_activity_date else "",
            profile.xp if profile else 0,
            profile.level if profile else 1,
        ])
    
    for column in ["A", "B", "C", "D", "E", "F", "G", "H"]:
        sheet.column_dimensions[column].width = 20
        
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="all-students-{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    return response

@admin_required
def admin_instructor_add(request):
    form = InstructorAddForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            phone = form.cleaned_data["phone"]
            username = form.build_username()
            user = User.objects.create_user(
                username=username,
                password=phone,
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                is_staff=True,
            )
            InstructorProfile.objects.update_or_create(
                user=user,
                defaults={
                    "phone": phone,
                    "specialty": form.cleaned_data["specialty"],
                    "bio": form.cleaned_data["bio"],
                    "avatar": form.cleaned_data["photo"],
                    "force_password_change": True,
                },
            )
            messages.success(
                request,
                f"تم إضافة المدرس {user.get_full_name()} بنجاح. "
                f"يمكنه الدخول بالاسم الكامل أو رقم الهاتف ({phone}) · كلمة المرور: {phone}",
            )
            return redirect("dashboard:admin_instructors")
                
    return render(request, "dashboard/admin_instructor_add.html", {"form": form})


@admin_required
def admin_instructors(request):
    search = request.GET.get("q", "").strip()
    status = request.GET.get("status", "")
    instructors = (
        User.objects.filter(instructor_profile__isnull=False)
        .select_related("instructor_profile")
        .annotate(courses_total=Count("courses", distinct=True), students_total=Count("courses__access_grants__user", distinct=True))
        .order_by("first_name", "last_name", "username")
    )
    if search:
        instructors = instructors.filter(
            models.Q(first_name__icontains=search)
            | models.Q(last_name__icontains=search)
            | models.Q(username__icontains=search)
            | models.Q(instructor_profile__phone__icontains=search)
            | models.Q(instructor_profile__specialty__icontains=search)
        )
    if status:
        instructors = instructors.filter(instructor_profile__status=status)
    stats = {
        "total": User.objects.filter(instructor_profile__isnull=False).count(),
        "active": User.objects.filter(instructor_profile__status="active").count(),
        "pending": User.objects.filter(instructor_profile__status="pending").count(),
        "suspended": User.objects.filter(instructor_profile__status="suspended").count(),
    }
    return render(
        request,
        "dashboard/admin_instructors.html",
        {
            "instructors": instructors,
            "stats": stats,
            "search": search,
            "status": status,
            "status_choices": InstructorProfile.STATUS_CHOICES,
        },
    )


@admin_required
def admin_instructor_edit(request, instructor_id):
    instructor = get_object_or_404(
        User.objects.select_related("instructor_profile").annotate(
            courses_total=Count("courses", distinct=True),
            students_total=Count("courses__access_grants__user", distinct=True),
        ),
        id=instructor_id,
        instructor_profile__isnull=False,
    )
    form = InstructorEditForm(request.POST or None, request.FILES or None, instructor=instructor)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "تم تحديث بيانات المدرس بنجاح.")
        return redirect("dashboard:admin_instructors")
    courses = Course.objects.filter(instructor=instructor).select_related("subject").order_by("-created_at")[:12]
    return render(
        request,
        "dashboard/admin_instructor_edit.html",
        {
            "form": form,
            "instructor": instructor,
            "courses": courses,
        },
    )


@admin_required
def admin_system_backup(request):
    if not settings.ENABLE_ADMIN_BACKUP_EXPORT:
        raise Http404("Backup export is disabled.")

    """تصدير قاعدة البيانات كاملة كملف JSON للنسخ الاحتياطي"""
    buffer = BytesIO()
    # Use management command to dump data to the buffer
    from django.core.management import call_command
    
    # Create a temporary file to hold the dump
    temp_file = "backup_temp.json"
    try:
        # Note: We use -o to ensure proper encoding on Windows
        call_command('dumpdata', exclude=['contenttypes', 'auth.Permission'], indent=2, output=temp_file)
        with open(temp_file, 'rb') as f:
            data = f.read()
        
        response = HttpResponse(data, content_type="application/json")
        response["Content-Disposition"] = f'attachment; filename="backup-{timezone.now().strftime("%Y-%m-%d-%H%M")}.json"'
        return response
    finally:
        if os.path.exists(temp_file):
            os.remove(temp_file)


@admin_required
def admin_course_control(request, course_id):
    course = get_object_or_404(Course.objects.select_related("subject", "instructor", "instructor__instructor_profile").prefetch_related("units__lessons"), id=course_id)
    batch_form = CourseCodeBatchForm(prefix="batch")
    import_form = CourseStudentImportForm(prefix="import", course=course)
    lesson_form = CourseLessonUploadForm(prefix="lesson", course=course)
    sale_form = CourseCodeSaleForm(prefix="sale", course=course)
    unit_form = CourseUnitQuickForm(prefix="unit")
    media_form = CourseCardMediaForm(prefix="media", instance=course)
    edit_form = CourseEditForm(prefix="edit", instance=course)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "edit_course_details":
            edit_form = CourseEditForm(request.POST, prefix="edit", instance=course)
            if edit_form.is_valid():
                edit_form.save()
                messages.success(request, "تم تحديث بيانات الدورة بنجاح.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
            else:
                messages.error(request, "حدث خطأ أثناء حفظ بيانات الدورة، يرجى مراجعة الحقول.")
        elif action == "create_batch":
            batch_form = CourseCodeBatchForm(request.POST, prefix="batch")
            if batch_form.is_valid():
                batch = batch_form.save(commit=False)
                batch.course = course
                batch.save()
                messages.success(request, "تم إنشاء دفعة الأكواد. يمكنك الآن رفع ملف الطلاب عليها.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
        elif action == "import_students":
            import_form = CourseStudentImportForm(request.POST, request.FILES, prefix="import", course=course)
            if import_form.is_valid():
                created = create_codes_from_upload(
                    import_form.cleaned_data["batch"],
                    import_form.cleaned_data["file"],
                    import_form.cleaned_data["free_codes"],
                )
                messages.success(request, f"تم توليد {created} كود من ملف الطلاب.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
        elif action == "upload_lesson":
            lesson_form = CourseLessonUploadForm(request.POST, request.FILES, prefix="lesson", course=course)
            if lesson_form.is_valid():
                lesson = lesson_form.save(commit=False)
                if not lesson.lesson_type:
                    lesson.lesson_type = "video"
                lesson.save()
                messages.success(request, "تم تسجيل الدرس بنجاح.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
        elif action == "create_unit":
            unit_form = CourseUnitQuickForm(request.POST, request.FILES, prefix="unit")
            if unit_form.is_valid():
                unit = unit_form.save(commit=False)
                unit.course = course
                unit.save()
                messages.success(request, f"تم إنشاء الجلسة {unit.title}.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
        elif action == "update_course_media":
            media_form = CourseCardMediaForm(request.POST, request.FILES, prefix="media", instance=course)
            if media_form.is_valid():
                media_form.save()
                messages.success(request, "تم تحديث غلاف كرت الدورة.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
        elif action == "sell_code":
            sale_form = CourseCodeSaleForm(request.POST, prefix="sale", course=course)
            if sale_form.is_valid():
                with transaction.atomic():
                    access_code = AccessCode.objects.select_for_update().get(id=sale_form.cleaned_data["code"].id, course=course)
                    student_phone = sanitize_plain_text(sale_form.cleaned_data["student_phone"])
                    student_name = sanitize_plain_text(sale_form.cleaned_data["student_name"])
                    student, created = User.objects.get_or_create(
                        username=student_phone,
                        defaults={"first_name": student_name, "is_active": True},
                    )
                    if created:
                        student.set_unusable_password()
                        student.save(update_fields=["password"])
                    elif student_name and not student.get_full_name():
                        student.first_name = student_name
                        student.save(update_fields=["first_name"])
                    StudentProfile.objects.get_or_create(
                        user=student,
                        defaults={"phone": student_phone, "track": course.get_academic_track_display()},
                    )
                    access_code.assigned_student_name = student_name
                    access_code.assigned_student_phone = student_phone
                    access_code.sale_status = "sold"
                    access_code.sold_by = request.user
                    access_code.sold_at = timezone.now()
                    price_amount = sale_form.cleaned_data.get("price_amount")
                    if price_amount is not None:
                        access_code.sold_price_cents = int(price_amount * 100)
                        access_code.price_reason = "تحديد يدوي من الإدارة"
                    else:
                        base_price = course.price_cents or 0
                        best_rule, max_discount = _best_active_discount_for_price(base_price, course=course)
                        if best_rule and base_price > 0:
                            access_code.sold_price_cents = base_price - max_discount
                            if best_rule.discount_percent > 0:
                                access_code.price_reason = f"حسم {best_rule.discount_percent}% بمناسبة {best_rule.name}"
                            else:
                                access_code.price_reason = f"حسم بقيمة {best_rule.discount_amount_syp} ل.س بمناسبة {best_rule.name}"
                        else:
                            access_code.sold_price_cents = base_price
                            access_code.price_reason = "سعر كامل"
                    access_code.save(update_fields=[
                        "assigned_student_name",
                        "assigned_student_phone",
                        "sale_status",
                        "sold_by",
                        "sold_at",
                        "sold_price_cents",
                        "price_reason",
                        "updated_at",
                    ])
                    AccessGrant.objects.get_or_create(
                        user=student,
                        course=course,
                        lesson=None,
                        defaults={
                            "access_code": access_code,
                            "source": "admin",
                            "starts_at": timezone.now(),
                            "expires_at": access_code.valid_until,
                        },
                    )
                messages.success(request, f"تم بيع الكود {access_code.code} وإضافة الدورة لحساب الطالب.")
                return redirect("dashboard:admin_course_control", course_id=course.id)
        elif action == "clear_device":
            grant = get_object_or_404(AccessGrant, id=request.POST.get("grant_id"), course=course)
            grant.device_fingerprint = ""
            grant.save(update_fields=["device_fingerprint"])
            UserDevice.objects.filter(user=grant.user).update(is_active=False)
            messages.success(request, "تم نقل/فك جهاز الطالب. عند دخوله القادم سيتم ربط الجهاز الجديد.")
            return redirect("dashboard:admin_course_control", course_id=course.id)

    codes = AccessCode.objects.filter(course=course)
    batches = AccessCodeBatch.objects.filter(course=course).select_related("institute", "sales_center").order_by("-created_at")
    grants = AccessGrant.objects.filter(course=course).select_related("user", "access_code").order_by("-created_at")[:30]
    centers = (
        SalesCenter.objects.filter(access_codes__course=course)
        .select_related("institute")
        .annotate(total_codes=Count("access_codes", distinct=True), activated_codes=Count("access_codes__grants", distinct=True))
        .distinct()
        .order_by("name")
    )
    lessons = Lesson.objects.filter(unit__course=course).select_related("unit").order_by("unit__sort_order", "sort_order")
    units = course.units.prefetch_related("lessons").order_by("sort_order", "id")
    print_logs = AccessCodePrintLog.objects.filter(batch__course=course).select_related("batch", "printed_by").order_by("-created_at")[:12]
    stats = {
        "lessons": lessons.count(),
        "codes": codes.count(),
        "sold": codes.filter(sale_status="sold").count(),
        "available": codes.filter(sale_status__in=["available", "reserved"]).count(),
        "free": codes.filter(is_free_code=True).count(),
        "activated": codes.filter(redeemed_count__gt=0).count(),
        "inactive": codes.filter(redeemed_count=0).count(),
        "students": AccessGrant.objects.filter(course=course).values("user").distinct().count(),
        "gross_sales": codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0,
        "my_sales": codes.filter(sale_status="sold", sold_by=request.user).count(),
        "my_gross_sales": codes.filter(sale_status="sold", sold_by=request.user).aggregate(total=Sum("sold_price_cents"))["total"] or 0,
        "prints": AccessCodePrintLog.objects.filter(batch__course=course).count(),
        "printed_cards": AccessCodePrintLog.objects.filter(batch__course=course).aggregate(total=Sum("cards_count"))["total"] or 0,
    }
    stats["gross_sales_display"] = f"{stats['gross_sales'] // 100:,}"
    stats["my_gross_sales_display"] = f"{stats['my_gross_sales'] // 100:,}"
    return render(
        request,
        "dashboard/admin_course_control.html",
        {
            "course": course,
            "stats": stats,
            "batches": batches,
            "grants": grants,
            "centers": centers,
            "lessons": lessons[:20],
            "units": units,
            "print_logs": print_logs,
            "batch_form": batch_form,
            "import_form": import_form,
            "lesson_form": lesson_form,
            "sale_form": sale_form,
            "unit_form": unit_form,
            "media_form": media_form,
            "edit_form": edit_form,
        },
    )


def _get_instructor_context(request):
    # Safe profile avatar and specialty fallbacks
    avatar_url = ""
    specialty = "مدرس معتمد"
    try:
        profile = getattr(request.user, 'instructor_profile', None)
        if profile:
            specialty = profile.specialty or "مدرس معتمد"
            if profile.avatar:
                avatar_url = profile.avatar.url
    except Exception:
        pass

    # 1. Fetch courses safely
    try:
        courses = Course.objects.filter(instructor=request.user).select_related("subject")
        # Annotate lessons_total and grants_total safely in python memory to prevent MySQL/PostgreSQL Count(distinct) join bugs in production
        for course in courses:
            course.lessons_total = Lesson.objects.filter(unit__course=course).count()
            course.grants_total = AccessGrant.objects.filter(course=course).count()
            
            # Safely fetch cover URL
            try:
                course.cover_url = course.cover.url if course.cover else ""
            except Exception:
                course.cover_url = ""
                
            # Safely fetch teacher photo URL
            try:
                course.teacher_photo_url = course.teacher_photo.url if course.teacher_photo else ""
            except Exception:
                course.teacher_photo_url = ""
                
            # Safely fetch subject name
            try:
                course.subject_name = course.subject.name if course.subject else "غير محدد"
            except Exception:
                course.subject_name = "غير محدد"
                
            # Safely format created_at date
            try:
                course.formatted_date = course.created_at.strftime("%Y-%m-%d") if course.created_at else ""
            except Exception:
                course.formatted_date = ""
    except Exception:
        courses = Course.objects.none()

    # 2. Fetch sessions safely
    sessions = OnlineLessonSession.objects.none()
    try:
        if courses.exists():
            sessions = (
                OnlineLessonSession.objects.filter(lesson__unit__course__in=courses)
                .select_related("lesson", "lesson__unit", "lesson__unit__course")
                .annotate(attendance_total=Count("attendances", distinct=True))
                .order_by("starts_at")
            )
    except Exception:
        pass

    # 3. Fetch attendance rows safely
    attendance_rows = LessonAttendance.objects.none()
    try:
        if courses.exists():
            attendance_rows = LessonAttendance.objects.filter(session__lesson__unit__course__in=courses).select_related(
                "user", "session", "session__lesson", "session__lesson__unit", "session__lesson__unit__course"
            ).order_by("-created_at")
    except Exception:
        pass

    # 4. Fetch Wallet and transactions safely
    wallet = None
    transactions = []
    total_gross_sales = 0
    total_net_earnings = 0
    try:
        instructor_profile = getattr(request.user, 'instructor_profile', None)
        if instructor_profile:
            wallet = Wallet.objects.filter(instructor=instructor_profile).first()
            if wallet:
                transactions = list(wallet.transactions.all()[:20])
                
        # Calculate Total Gross Sales and Net Earnings dynamically
        from django.db.models import Sum
        from billing.models import AccessCode
        from apps.instructor_finance.models import RevenueShareAgreement
        
        # 1. Total Gross Sales (L.S)
        total_gross_sales = (AccessCode.objects.filter(course__instructor=request.user, sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) // 100
        
        # 2. Total Net Earnings based on agreements/defaults (L.S)
        courses_list = Course.objects.filter(instructor=request.user)
        for c in courses_list:
            c_gross_cents = AccessCode.objects.filter(course=c, sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0
            agreement = RevenueShareAgreement.objects.filter(instructor=request.user, course=c, is_active=True).first()
            if not agreement:
                agreement = RevenueShareAgreement.objects.filter(instructor=request.user, course=None, is_active=True).first()
            share_bps = agreement.commission_bps if agreement else 3000  # Default 30%
            total_net_earnings += int(c_gross_cents * (share_bps / 10000.0)) // 100
    except Exception:
        pass

    # 5. Fetch grants list safely
    grants_list = AccessGrant.objects.none()
    selected_course_id = request.GET.get("course_id", "")
    try:
        grants_list = AccessGrant.objects.filter(course__instructor=request.user).select_related(
            "user", "course"
        ).order_by("-created_at")
        if selected_course_id and selected_course_id.isdigit():
            grants_list = grants_list.filter(course_id=int(selected_course_id))
    except Exception:
        pass

    # 6. Gather counts safely
    try:
        courses_count = courses.count()
    except Exception:
        courses_count = 0

    try:
        lessons_count = Lesson.objects.filter(unit__course__in=courses).count()
    except Exception:
        lessons_count = 0

    try:
        sessions_count = sessions.count()
    except Exception:
        sessions_count = 0

    try:
        attendance_count = attendance_rows.count()
    except Exception:
        attendance_count = 0

    try:
        students_count = AccessGrant.objects.filter(course__in=courses).values("user").distinct().count()
    except Exception:
        students_count = 0

    try:
        codes_count = AccessCode.objects.filter(course__in=courses).count()
    except Exception:
        codes_count = 0

    return {
        "courses": courses,
        "courses_count": courses_count,
        "lessons_count": lessons_count,
        "sessions_count": sessions_count,
        "attendance_count": attendance_count,
        "students_count": students_count,
        "codes_count": codes_count,
        "sessions": sessions[:8] if hasattr(sessions, '__getitem__') else [],
        "attendance_rows": attendance_rows[:10] if hasattr(attendance_rows, '__getitem__') else [],
        "wallet": wallet,
        "transactions": transactions,
        "grants_list": grants_list,
        "selected_course_id": selected_course_id,
        "avatar_url": avatar_url,
        "specialty": specialty,
        "total_gross_sales": total_gross_sales,
        "total_net_earnings": total_net_earnings,
    }


@instructor_required
def instructor_dashboard(request):
    return redirect("dashboard:instructor_courses")


@instructor_required
def instructor_courses(request):
    context = _get_instructor_context(request)
    context["active_page"] = "courses"
    return render(request, "dashboard/instructor_courses.html", context)


@instructor_required
def instructor_students(request):
    context = _get_instructor_context(request)
    context["active_page"] = "students"
    return render(request, "dashboard/instructor_students.html", context)


@instructor_required
def instructor_finance(request):
    context = _get_instructor_context(request)
    context["active_page"] = "finance"
    return render(request, "dashboard/instructor_finance.html", context)


@instructor_required
def instructor_announcements(request):
    if request.method == "POST" and request.POST.get("action") == "send_message":
        course_id = request.POST.get("course_id", "").strip()
        message_body = request.POST.get("message_body", "").strip()
        send_whatsapp = request.POST.get("send_whatsapp") == "on"
        
        if not message_body:
            messages.warning(request, "تنبيه: نص الرسالة فارغ لا يمكن إرساله.")
        else:
            if course_id == "all":
                grants = AccessGrant.objects.filter(course__instructor=request.user)
            elif course_id.isdigit():
                grants = AccessGrant.objects.filter(course_id=int(course_id), course__instructor=request.user)
            else:
                grants = AccessGrant.objects.none()
                
            students = User.objects.filter(access_grants__in=grants).distinct()
            
            if not students.exists():
                messages.warning(request, "لم يتم العثور على طلاب مفعلين في المادة المحددة لإرسال الرسالة إليهم.")
            else:
                instructor_name = request.user.get_full_name() or request.user.first_name or request.user.username
                full_message = f"رسالة من الأستاذ {instructor_name}: {message_body}"
                
                notifications = [
                    StudentNotification(
                        user=student,
                        notification_type="system",
                        title=f"رسالة من الأستاذ {instructor_name}",
                        body=message_body,
                        url="/student/notifications/"
                    )
                    for student in students
                ]
                StudentNotification.objects.bulk_create(notifications)
                
                if send_whatsapp:
                    import threading
                    from .whatsapp_utils import send_whatsapp_message
                    
                    def send_bulk_whatsapp_bg(students_list, text):
                        for student in students_list:
                            phone = None
                            if hasattr(student, 'student_profile') and student.student_profile.phone:
                                phone = student.student_profile.phone
                            elif student.username.isdigit():
                                phone = student.username
                            if phone:
                                send_whatsapp_message(phone, text)
                                
                    threading.Thread(
                        target=send_bulk_whatsapp_bg,
                        args=(list(students), full_message),
                        daemon=True
                    ).start()
                    messages.success(request, f"✅ تم إرسال التنبيه الداخلي وجاري إرسال رسائل الواتساب بالخلفية لـ {students.count()} طالب.")
                else:
                    messages.success(request, f"✅ تم إرسال التنبيه الداخلي بنجاح لـ {students.count()} طالب.")
                    
        return redirect("dashboard:instructor_announcements")

    context = _get_instructor_context(request)
    context["active_page"] = "announcements"
    return render(request, "dashboard/instructor_announcements.html", context)


@instructor_required
def instructor_settings(request):
    if request.method == "POST" and request.POST.get("action") == "change_password":
        current_password = request.POST.get("current_password", "").strip()
        new_password = request.POST.get("new_password", "").strip()
        confirm_password = request.POST.get("confirm_password", "").strip()
        
        if not request.user.check_password(current_password):
            messages.error(request, "⚠️ كلمة المرور الحالية غير صحيحة.")
        elif len(new_password) < 8:
            messages.error(request, "⚠️ يجب أن تتكون كلمة المرور الجديدة من 8 خانات على الأقل.")
        elif new_password != confirm_password:
            messages.error(request, "⚠️ كلمة المرور الجديدة وتأكيدها غير متطابقتين.")
        else:
            request.user.set_password(new_password)
            request.user.save()
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            profile = request.user.instructor_profile
            profile.force_password_change = False
            profile.save(update_fields=["force_password_change"])
            request.session.pop("instructor_password_modal_dismissed", None)
            messages.success(request, "✅ تم تغيير كلمة المرور بنجاح.")

        return redirect("dashboard:instructor_settings")

    context = _get_instructor_context(request)
    context["active_page"] = "settings"
    return render(request, "dashboard/instructor_settings.html", context)



def _package_subject_groups(access_code, user, current_device):
    courses = (
        access_code.package.eligible_courses_queryset()
        .select_related("subject", "instructor", "instructor__instructor_profile")
        .order_by("subject__name", "instructor__first_name", "title", "id")
    )
    groups_map = defaultdict(list)
    for course in courses:
        existing_grant = AccessGrant.objects.filter(user=user, course=course, lesson=None).first()
        if existing_grant and existing_grant.device_fingerprint and existing_grant.device_fingerprint != current_device:
            continue
        groups_map[course.subject_id].append(course)
    return [courses for _subject_id, courses in sorted(groups_map.items(), key=lambda item: item[1][0].subject.name)]


def _build_package_selection(access_code, user, current_device):
    groups = _package_subject_groups(access_code, user, current_device)
    steps = []
    auto_course_ids = []
    for courses in groups:
        if len(courses) == 1:
            auto_course_ids.append(courses[0].id)
        else:
            steps.append(
                {
                    "subject": courses[0].subject,
                    "courses": courses,
                    "index": len(steps) + 1,
                }
            )
    return {
        "code": access_code.code,
        "package": access_code.package,
        "steps": steps,
        "auto_course_ids": auto_course_ids,
        "requires_choice": bool(steps),
        "total_steps": len(steps),
    }


def _redeem_package_choices(request, access_code, current_device, auto_select=False):
    groups = _package_subject_groups(access_code, request.user, current_device)
    if not groups:
        return {"ok": False, "message": "لا توجد دورات منشورة داخل هذه الباقة حالياً.", "count": 0}

    selected_course_ids = []
    posted_values = list(request.POST.getlist("course_choices"))
    posted_values.extend(value for key, value in request.POST.items() if key.startswith("course_choice_"))
    posted_ids = {int(value) for value in posted_values if str(value).isdigit()}
    for courses in groups:
        valid_ids = {course.id for course in courses}
        if len(courses) == 1:
            selected_course_ids.append(courses[0].id)
            continue
        if auto_select:
            return {"ok": False, "message": "هذه الباقة تحتاج اختيار مدرس لكل مادة.", "count": 0}
        matching_ids = valid_ids & posted_ids
        if len(matching_ids) != 1:
            return {"ok": False, "message": f"اختر دورة واحدة لمادة {courses[0].subject.name}.", "count": 0}
        selected_course_ids.append(next(iter(matching_ids)))

    grants = []
    now = timezone.now()
    for course in Course.objects.filter(id__in=selected_course_ids).select_related("subject"):
        grant, _created = AccessGrant.objects.get_or_create(
            user=request.user,
            course=course,
            lesson=None,
            defaults={
                "access_code": access_code,
                "source": "code",
                "device_fingerprint": current_device,
                "starts_at": now,
                "expires_at": access_code.valid_until,
            },
        )
        if grant.device_fingerprint and grant.device_fingerprint != current_device:
            return {"ok": False, "message": "إحدى الدورات مرتبطة بجهاز آخر. تواصل مع الإدارة لنقل الجهاز.", "count": 0}
        if not grant.device_fingerprint:
            grant.device_fingerprint = current_device
            grant.save(update_fields=["device_fingerprint"])
        grants.append(grant)

    access_code.redeemed_count += 1
    if not access_code.assigned_student_name:
        access_code.assigned_student_name = request.user.get_full_name() or request.user.first_name or request.user.username
    if not access_code.assigned_student_phone:
        profile = getattr(request.user, "student_profile", None)
        access_code.assigned_student_phone = profile.phone if profile else request.user.username
    if not access_code.sold_at:
        access_code.sold_at = timezone.now()

    if access_code.sale_status in {"available", "reserved"}:
        access_code.sale_status = "free" if access_code.is_free_code else "sold"
        if not access_code.is_free_code and access_code.package:
            base_price = access_code.package.price_cents or 0
            if base_price > 0:
                best_rule, max_discount = _best_active_discount_for_price(base_price, package=access_code.package)
                if best_rule:
                    access_code.sold_price_cents = base_price - max_discount
                    if best_rule.discount_percent > 0:
                        access_code.price_reason = f"حسم {best_rule.discount_percent}% بمناسبة {best_rule.name}"
                    else:
                        access_code.price_reason = f"حسم بقيمة {best_rule.discount_amount_syp} ل.س بمناسبة {best_rule.name}"
                else:
                    access_code.sold_price_cents = base_price
                    access_code.price_reason = "سعر كامل"
    access_code.save(update_fields=["redeemed_count", "sale_status", "sold_price_cents", "price_reason", "assigned_student_name", "assigned_student_phone", "sold_at", "updated_at"])
    StudentNotification.objects.create(
        user=request.user,
        notification_type="access",
        title="تم تفعيل باقة جديدة",
        body=f"تمت إضافة {len(grants)} دورة من باقة {access_code.package.name} إلى حسابك.",
        url="/student/my-courses/",
    )
    return {"ok": True, "message": "", "count": len(grants)}


@login_required
def student_dashboard(request):
    from analytics.services import TrackingService
    TrackingService.log_landing_visit(request)
    
    redeem_form = RedeemCodeForm(request.POST or None)
    current_device = _current_device_fingerprint(request)
    package_selection = None

    if request.method == "POST" and request.POST.get("action") == "transfer_devices":
        AccessGrant.objects.filter(user=request.user).update(device_fingerprint=current_device)
        messages.success(request, "تم نقل تفعيل كافة موادك ودوراتك إلى هذا الجهاز الحالي بنجاح!")
        return redirect("dashboard:student_dashboard")

    if request.method == "POST" and request.POST.get("action") == "complete_package_redeem":
        code_value = request.POST.get("package_code", "").strip()
        arabic_digits = "٠١٢٣٤٥٦٧٨٩"
        english_digits = "0123456789"
        for i in range(10):
            code_value = code_value.replace(arabic_digits[i], english_digits[i])
        code_value = "".join(code_value.split()).upper()
        
        with transaction.atomic():
            access_code = AccessCode.objects.select_for_update().filter(code__iexact=code_value).first()
            if access_code is None or access_code.access_type != "package" or not access_code.package:
                messages.error(request, "كود الباقة غير صحيح.")
            else:
                allowed, reason = access_code.is_redeemable(timezone.now())
                if not allowed:
                    messages.error(request, reason)
                elif not _access_code_matches_student(access_code, request.user):
                    messages.error(request, "هذا الكود مخصص لطالب آخر ولا يمكن تفعيله على هذا الحساب.")
                else:
                    result = _redeem_package_choices(request, access_code, current_device)
                    if result["ok"]:
                        messages.success(request, f"تم تفعيل الباقة وإضافة {result['count']} دورة إلى مكتبتك.")
                        return redirect("dashboard:student_dashboard")
                    messages.error(request, result["message"])
                    
    elif request.method == "POST" and redeem_form.is_valid():
        code_value = redeem_form.cleaned_data["code"].strip()
        arabic_digits = "٠١٢٣٤٥٦٧٨٩"
        english_digits = "0123456789"
        for i in range(10):
            code_value = code_value.replace(arabic_digits[i], english_digits[i])
        code_value = "".join(code_value.split()).upper()
        
        with transaction.atomic():
            access_code = AccessCode.objects.select_for_update().filter(code__iexact=code_value).first()
            if access_code is None:
                messages.error(request, "الكود غير صحيح.")
            else:
                allowed, reason = access_code.is_redeemable(timezone.now())
                if not allowed:
                    messages.error(request, reason)
                elif not _access_code_matches_student(access_code, request.user):
                    messages.error(request, "هذا الكود مخصص لطالب آخر ولا يمكن تفعيله على هذا الحساب.")
                elif access_code.access_type == "package" and access_code.package:
                    package_selection = _build_package_selection(access_code, request.user, current_device)
                    if package_selection["requires_choice"]:
                        messages.info(request, "اختر مدرساً واحداً لكل مادة لإكمال تفعيل الباقة.")
                    else:
                        result = _redeem_package_choices(request, access_code, current_device, auto_select=True)
                        if result["ok"]:
                            messages.success(request, f"تم تفعيل الباقة وإضافة {result['count']} دورة إلى مكتبتك.")
                            return redirect("dashboard:student_dashboard")
                        messages.error(request, result["message"])
                else:
                    grant, created = AccessGrant.objects.get_or_create(
                        user=request.user,
                        course=access_code.course,
                        lesson=access_code.lesson,
                        defaults={
                            "access_code": access_code,
                            "source": "code",
                            "device_fingerprint": current_device,
                            "starts_at": timezone.now(),
                            "expires_at": access_code.valid_until,
                        },
                    )
                    if created:
                        access_code.redeemed_count += 1
                        if not access_code.assigned_student_name:
                            access_code.assigned_student_name = request.user.get_full_name() or request.user.first_name or request.user.username
                        if not access_code.assigned_student_phone:
                            profile = getattr(request.user, "student_profile", None)
                            access_code.assigned_student_phone = profile.phone if profile else request.user.username
                        if not access_code.sold_at:
                            access_code.sold_at = timezone.now()

                        if access_code.sale_status in {"available", "reserved"}:
                            access_code.sale_status = "free" if access_code.is_free_code else "sold"
                            if not access_code.is_free_code:
                                base_price = 0
                                if access_code.course:
                                    base_price = access_code.course.price_cents or 0
                                elif access_code.lesson:
                                    if not base_price and access_code.course:
                                        base_price = access_code.course.price_cents or 0
                                if base_price > 0:
                                    best_rule, max_discount = _best_active_discount_for_price(base_price, course=access_code.course)
                                    if best_rule:
                                        access_code.sold_price_cents = base_price - max_discount
                                        if best_rule.discount_percent > 0:
                                            access_code.price_reason = f"حسم {best_rule.discount_percent}% بمناسبة {best_rule.name}"
                                        else:
                                            access_code.price_reason = f"حسم بقيمة {best_rule.discount_amount_syp} ل.س بمناسبة {best_rule.name}"
                                    else:
                                        access_code.sold_price_cents = base_price
                                        access_code.price_reason = "سعر كامل"
                        access_code.save(update_fields=["redeemed_count", "sale_status", "sold_price_cents", "price_reason", "assigned_student_name", "assigned_student_phone", "sold_at", "updated_at"])
                        StudentNotification.objects.create(
                            user=request.user,
                            notification_type="access",
                            title="تم تفعيل كود جديد",
                            body=f"تمت إضافة {access_code.course or access_code.lesson or access_code.plan} إلى حسابك.",
                            url="/student/",
                        )
                        messages.success(request, "تم تفعيل الكود وإضافة المادة إلى حسابك.")
                    else:
                        if grant.device_fingerprint and grant.device_fingerprint != current_device:
                            messages.error(request, "هذه المكثفة مرتبطة بجهاز آخر. تواصل مع الإدارة لنقل الجهاز.")
                        else:
                            if not grant.device_fingerprint:
                                grant.device_fingerprint = current_device
                                grant.save(update_fields=["device_fingerprint"])
                            messages.info(request, "هذه المادة أو الدرس مفعّل لديك مسبقًا.")
        if not package_selection:
            return redirect("dashboard:student_dashboard")

    grants = _device_grants(request.user, current_device).select_related("course", "lesson", "access_code").order_by("-created_at")
    
    # Courses that are activated but on a DIFFERENT device (student should see them with lock icon)
    device_grant_ids = set(grants.values_list("id", flat=True))
    other_device_grants = _active_access_grants(request.user).select_related(
        "course", "course__instructor", "lesson", "access_code"
    ).exclude(id__in=device_grant_ids).filter(course__isnull=False).order_by("-created_at")
    
    sessions = _available_sessions_for_user(request.user, current_device)
    unlocked_lessons = (
        Lesson.objects.filter(id__in=_student_lesson_ids_for_device(request.user, current_device))
        .select_related("unit", "unit__course")
        .order_by("unit__course__title", "unit__sort_order", "sort_order")[:16]
    )

    # Gamification and Continue Learning
    student_profile = None
    if hasattr(request.user, 'student_profile'):
        student_profile = request.user.student_profile
        
    continue_learning = LessonProgress.objects.filter(user=request.user).select_related("lesson", "lesson__unit", "lesson__unit__course").order_by("-updated_at").first()

    context = {
        "redeem_form": redeem_form,
        "grants": grants,
        "other_device_grants": other_device_grants,
        "unlocked_lessons": unlocked_lessons,
        "sessions": sessions[:12],
        "attendances": LessonAttendance.objects.filter(user=request.user).select_related("session").order_by("-created_at")[:12],
        "attempts": Attempt.objects.filter(user=request.user).select_related("exam").order_by("-created_at")[:8],
        "notifications": StudentNotification.objects.filter(user=request.user).order_by("-created_at")[:8],
        "student_profile": student_profile,
        "continue_learning": continue_learning,
        "package_selection": package_selection,
    }
    return render(request, "dashboard/student_dashboard.html", context)


@login_required
def student_device_ping(request):
    response = JsonResponse({"status": "active"})
    response["Cache-Control"] = "no-store"
    return response


@login_required
def my_courses_page(request):
    current_device = _current_device_fingerprint(request)
    # Load ALL active access grants for this student (regardless of device)
    all_grants = _active_access_grants(request.user).select_related(
        "course", "course__subject", "course__instructor", "course__instructor__instructor_profile", "lesson", "access_code"
    ).order_by("-created_at")
    
    # Build set of device-specific grant IDs for fast lookup
    device_grant_ids = set(
        _device_grants(request.user, current_device).values_list("id", flat=True)
    )
    
    progress_map = {}
    for cp in CourseProgress.objects.filter(user=request.user):
        progress_map[cp.course_id] = int(cp.completion_percent)
    for grant in all_grants:
        if grant.course:
            grant.progress_percent = progress_map.get(grant.course.id, 0)
        else:
            grant.progress_percent = 0
        grant.is_on_current_device = (grant.id in device_grant_ids)
    context = {
        "grants": all_grants,
    }
    return render(request, "dashboard/my_courses.html", context)



@login_required
def favorites_page(request):
    context = {
        "favorites": [],
    }
    return render(request, "dashboard/favorites.html", context)


@login_required
def notifications_page(request):
    notifications = StudentNotification.objects.filter(user=request.user).order_by("-created_at")
    unread_count = notifications.filter(read_at__isnull=True).count()
    if request.method == "POST" and request.POST.get("action") == "mark_all_read":
        notifications.filter(read_at__isnull=True).update(read_at=timezone.now())
        messages.success(request, "تم تعليم جميع الإشعارات كمقروءة.")
        return redirect(reverse("dashboard:notifications"))
    context = {
        "notifications": notifications,
        "unread_count": unread_count,
    }
    return render(request, "dashboard/notifications.html", context)


@login_required
def profile_page(request):
    try:
        student_profile = None
        if hasattr(request.user, "student_profile"):
            student_profile = request.user.student_profile
        
        current_device = _current_device_fingerprint(request)
        grants = _device_grants(request.user, current_device).select_related("course", "lesson").order_by("-created_at")
        
        total_xp = student_profile.xp if student_profile else 0
        level = student_profile.level if student_profile else 1
        streak = student_profile.streak_days if student_profile else 0
        
        completed_lessons = LessonProgress.objects.filter(user=request.user, completed_at__isnull=False).count()
        
        # Optimize query for total lessons
        course_ids = grants.filter(course__isnull=False).values_list("course_id", flat=True)
        total_lessons = Lesson.objects.filter(unit__course_id__in=course_ids).distinct().count()
        
        context = {
            "student_profile": student_profile,
            "grants": grants,
            "total_xp": total_xp,
            "level": level,
            "streak": streak,
            "completed_lessons": completed_lessons,
            "total_lessons": total_lessons,
        }

        return render(request, "dashboard/profile.html", context)
    except Exception as e:
        import traceback
        with open("profile_error_log.txt", "a", encoding="utf-8") as f:
            f.write(f"Error at {timezone.now()}: {str(e)}\n")
            f.write(traceback.format_exc() + "\n")
        return render(request, "500.html", status=500)

    
@login_required
def profile_edit(request):
    student_profile, _created = StudentProfile.objects.get_or_create(user=request.user)
        
    if request.method == "POST":
        try:
            first_name = sanitize_plain_text(request.POST.get("first_name"), max_length=120)
            last_name = sanitize_plain_text(request.POST.get("last_name"), max_length=120)
            phone = validate_syrian_mobile(request.POST.get("phone"), user=request.user)
            bio = sanitize_plain_text(request.POST.get("bio"), max_length=1000)
        except ValidationError as exc:
            messages.error(request, " ".join(exc.messages))
            return redirect("dashboard:profile_edit")
        
        request.user.first_name = first_name
        request.user.last_name = last_name
        request.user.username = phone
        request.user.save()
        
        student_profile.phone = phone
        student_profile.bio = bio
        if request.FILES.get("avatar"):
            student_profile.avatar = request.FILES.get("avatar")
        student_profile.save()
            
        messages.success(request, "تم تحديث بياناتك الشخصية بنجاح.")
        return redirect("dashboard:profile")
        
    context = {
        "student_profile": student_profile,
    }
    return render(request, "dashboard/profile_edit.html", context)

@login_required
def view_certificate(request, course_id):
    from learning.models import Certificate
    course = get_object_or_404(Course, id=course_id)
    certificate = Certificate.objects.filter(user=request.user, course=course).first()
    
    if not certificate:
        # Check if student completed the course to issue a certificate
        progress = CourseProgress.objects.filter(user=request.user, course=course).first()
        if progress and progress.completion_percent >= 100:
            certificate = Certificate.objects.create(user=request.user, course=course)
        else:
            messages.error(request, "لم تقم بإكمال هذه الدورة بعد لتتمكن من الحصول على الشهادة.")
            return redirect("dashboard:student_course_detail", course_id=course.id)
            
    context = {
        "certificate": certificate,
        "course": course,
        "student": request.user,
    }
    return render(request, "dashboard/certificate.html", context)

def departments_list(request):
    from learning.models import Subject
    # Grouping logic or just listing subjects
    subjects = Subject.objects.all().prefetch_related('courses')
    context = {
        "subjects": subjects,
    }
    return render(request, "dashboard/departments.html", context)

def search_results(request):
    query = request.GET.get("q", "")
    courses = []
    if query:
        courses = Course.objects.filter(
            models.Q(title__icontains=query) | 
            models.Q(description__icontains=query) |
            models.Q(subject__name__icontains=query)
        ).filter(status="published").distinct()
    
    context = {
        "query": query,
        "courses": courses,
    }
    return render(request, "dashboard/search_results.html", context)

def about_page(request):
    return render(request, "dashboard/about.html")

def contact_page(request):
    if request.method == "POST":
        # Logic for contact form could go here
        messages.success(request, "تم استلام رسالتك بنجاح. سنقوم بالرد عليك في أقرب وقت.")
        return redirect("dashboard:contact")
    return render(request, "dashboard/contact.html")

def instructors_list(request):
    # Fetching users who are staff and have at least one published course
    from django.contrib.auth.models import User
    from learning.models import Course
    
    # Show all instructors who have a profile, regardless of published courses
    instructors = User.objects.filter(is_staff=True, instructor_profile__isnull=False).select_related('instructor_profile').distinct()
    
    context = {
        "instructors": instructors,
    }
    return render(request, "dashboard/instructors.html", context)

def faq_page(request):
    return render(request, "dashboard/faq.html")

def privacy_page(request):
    return render(request, "dashboard/privacy.html")

def terms_page(request):
    return render(request, "dashboard/terms.html")


def _student_course_ids(user):
    return list(_active_access_grants(user).filter(course__isnull=False).values_list("course_id", flat=True))


def _student_course_ids_for_device(user, device_fingerprint):
    return list(_device_grants(user, device_fingerprint).filter(course__isnull=False).values_list("course_id", flat=True))


def _student_lesson_ids(user):
    direct_lessons = list(_active_access_grants(user).filter(lesson__isnull=False).values_list("lesson_id", flat=True))
    course_lessons = list(Lesson.objects.filter(unit__course_id__in=_student_course_ids(user)).values_list("id", flat=True))
    return list(set(direct_lessons + course_lessons))


def _student_lesson_ids_for_device(user, device_fingerprint):
    direct_lessons = list(_device_grants(user, device_fingerprint).filter(lesson__isnull=False).values_list("lesson_id", flat=True))
    course_lessons = list(Lesson.objects.filter(unit__course_id__in=_student_course_ids_for_device(user, device_fingerprint)).values_list("id", flat=True))
    return list(set(direct_lessons + course_lessons))


def _available_sessions_for_user(user, device_fingerprint=None):
    lesson_ids = _student_lesson_ids_for_device(user, device_fingerprint) if device_fingerprint else _student_lesson_ids(user)
    return (
        OnlineLessonSession.objects.filter(lesson_id__in=lesson_ids)
        .select_related("lesson", "lesson__unit", "lesson__unit__course")
        .order_by("starts_at")
    )


@login_required
def student_course_detail(request, course_id):
    current_device = _current_device_fingerprint(request)
    course = get_object_or_404(Course.objects.select_related("subject", "instructor", "instructor__instructor_profile").prefetch_related("units__lessons"), id=course_id)
    
    is_staff = request.user.is_staff
    is_impersonation = (current_device == "impersonation_bypass")
    has_impersonation_bypass = bool(request.session.get("impersonator_admin_id"))
    
    # Check if student has active access to this course globally
    has_global_access = _active_access_grants(request.user).filter(course=course).exists()
    if not has_global_access:
        raise Http404("Course not found")
    
    # Auto-bind device fingerprint for grants that don't have one yet (e.g. admin manual grants)
    # This ensures the course locks to the FIRST device the student opens it from.
    if current_device and current_device != "impersonation_bypass" and not is_staff and not has_impersonation_bypass:
        unbound_grant = _active_access_grants(request.user).filter(course=course, device_fingerprint="").first()
        if unbound_grant:
            unbound_grant.device_fingerprint = current_device
            unbound_grant.save(update_fields=["device_fingerprint"])
        
    # Check if the active grant allows accessing on this device
    has_device_access = is_staff or has_impersonation_bypass or is_impersonation or _device_grants(request.user, current_device).filter(course=course).exists()
    device_locked = not has_device_access
    
    active_grant = _active_access_grants(request.user).filter(course=course).first()
    active_device_fingerprint = active_grant.device_fingerprint if active_grant else ""
    
    whatsapp_url = ""
    if device_locked:
        import urllib.parse
        student_name = request.user.get_full_name() or request.user.username
        student_username = request.user.username
        profile = getattr(request.user, "student_profile", None)
        student_phone = profile.phone if (profile and getattr(profile, "phone", None)) else request.user.username
        
        whatsapp_text = (
            f"مرحباً منصة محترفو التعليم، أود طلب نقل تفعيل دورة ({course.title}) إلى جهازي الحالي.\n"
            f"اسم الطالب: {student_name}\n"
            f"اسم المستخدم: {student_username}\n"
            f"رقم الطالب: {student_phone}\n"
            f"رمز الجهاز الجديد: {current_device}"
        )
        whatsapp_url = f"https://wa.me/963984011372?text={urllib.parse.quote(whatsapp_text)}"
        
    progress = CourseProgress.objects.filter(user=request.user, course=course).first()
    attempts = Attempt.objects.filter(user=request.user, exam__course=course).select_related("exam").order_by("-created_at")
    completed_lesson_ids = set(
        LessonProgress.objects.filter(
            user=request.user,
            lesson__unit__course=course,
            completed_at__isnull=False,
        ).values_list("lesson_id", flat=True)
    )
    return render(
        request,
        "dashboard/student_course_detail.html",
        {
            "course": course,
            "attempts": attempts,
            "completed_lesson_ids": completed_lesson_ids,
            "progress": progress,
            "device_locked": device_locked,
            "current_device": current_device,
            "active_device_fingerprint": active_device_fingerprint,
            "whatsapp_url": whatsapp_url,
            "active_grant": active_grant,
        },
    )




@login_required
def student_lesson_detail(request, lesson_id):
    current_device = _current_device_fingerprint(request)
    lesson = get_object_or_404(Lesson.objects.select_related("unit", "unit__course"), id=lesson_id)
    if not request.user.is_staff and lesson.id not in _student_lesson_ids_for_device(request.user, current_device):
        raise Http404("Lesson not found")

    progress, _created = LessonProgress.objects.get_or_create(
        user=request.user,
        lesson=lesson,
    )
    _refresh_course_progress(request.user, lesson.unit.course)
    signed_video_url = ""
    if lesson.video_file:
        signed_video_url = _signed_lesson_video_url(request, lesson)
    return render(
        request,
        "dashboard/student_lesson_detail.html",
        {"lesson": lesson, "progress": progress, "video_embed_url": _video_embed_url(lesson.video_url), "signed_video_url": signed_video_url},
    )


@login_required
def signed_lesson_video(request, lesson_id, token):
    lesson = get_object_or_404(Lesson.objects.select_related("unit", "unit__course"), id=lesson_id)
    current_device = _current_device_fingerprint(request)
    if not request.user.is_staff and lesson.id not in _student_lesson_ids_for_device(request.user, current_device):
        raise Http404("Video not found")
    try:
        payload = signing.TimestampSigner(salt="lesson-video").unsign_object(token, max_age=60 * 10)
    except signing.BadSignature as exc:
        raise Http404("Video not found") from exc
    if payload.get("lesson_id") != lesson.id or payload.get("user_id") != request.user.id or payload.get("device") != current_device:
        raise Http404("Video not found")
    if not lesson.video_file:
        raise Http404("Video not found")
    response = FileResponse(lesson.video_file.open("rb"), content_type="application/octet-stream")
    response["Cache-Control"] = "no-store"
    response["Content-Disposition"] = 'inline; filename="lesson-video.dat"'
    return response


def _video_embed_url(video_url):
    if not video_url:
        return ""
    if "youtube.com/watch?v=" in video_url:
        return video_url.replace("watch?v=", "embed/")
    if "youtu.be/" in video_url:
        return video_url.replace("youtu.be/", "www.youtube.com/embed/")
    if "vimeo.com/" in video_url and "player.vimeo.com" not in video_url:
        video_id = video_url.rstrip("/").split("/")[-1]
        return f"https://player.vimeo.com/video/{video_id}"
    
    # Bunny.net Support with Secure Token Authentication
    if "mediadelivery.net" in video_url:
        from urllib.parse import urlparse
        import hashlib
        import time
        import os
        parsed = urlparse(video_url)
        path_parts = [p for p in parsed.path.split("/") if p]
        library_id = ""
        video_id = ""
        if len(path_parts) >= 3:
            library_id = path_parts[1]
            video_id = path_parts[2]
        elif len(path_parts) == 2:
            library_id = path_parts[0]
            video_id = path_parts[1]
        if library_id and video_id:
            token_key = os.environ.get("BUNNY_STREAM_TOKEN_KEY", "").strip()
            if token_key:
                expires = int(time.time()) + 7200
                token_input = f"{token_key}{video_id}{expires}"
                token = hashlib.sha256(token_input.encode("utf-8")).hexdigest()
                return f"https://iframe.mediadelivery.net/embed/{library_id}/{video_id}?token={token}&expires={expires}"
            else:
                return f"https://iframe.mediadelivery.net/embed/{library_id}/{video_id}"
    # Legacy Fallback if parsing library_id and video_id failed
    if "mediadelivery.net" in video_url:
        if "/play/" in video_url:
            return video_url.replace("/play/", "/embed/")
        return video_url
        
    return video_url


def _signed_lesson_video_url(request, lesson):
    from django.urls import reverse

    payload = {
        "lesson_id": lesson.id,
        "user_id": request.user.id,
        "device": _current_device_fingerprint(request),
    }
    token = signing.TimestampSigner(salt="lesson-video").sign_object(payload)
    return reverse("dashboard:signed_lesson_video", args=[lesson.id, token])


@login_required
def complete_lesson(request, lesson_id):
    current_device = _current_device_fingerprint(request)
    lesson = get_object_or_404(Lesson.objects.select_related("unit", "unit__course"), id=lesson_id)
    if lesson.id not in _student_lesson_ids_for_device(request.user, current_device):
        messages.error(request, "هذا الدرس غير مفعّل على هذا الجهاز.")
        return redirect("dashboard:student_dashboard")

    from datetime import date
    from datetime import timedelta

    LessonProgress.objects.update_or_create(
        user=request.user,
        lesson=lesson,
        defaults={
            "watched_seconds": lesson.duration_seconds or 0,
            "last_position_seconds": lesson.duration_seconds or 0,
            "completed_at": timezone.now(),
        },
    )
    _refresh_course_progress(request.user, lesson.unit.course)
    
    # Gamification Engine
    if hasattr(request.user, 'student_profile'):
        profile = request.user.student_profile
        today = date.today()
        
        # Streak Logic
        if profile.last_activity_date != today:
            if profile.last_activity_date == today - timedelta(days=1):
                profile.streak_days += 1
            else:
                profile.streak_days = 1
            profile.last_activity_date = today
            
        # XP & Level Logic
        profile.xp += 50
        new_level = (profile.xp // 500) + 1
        level_up = False
        if new_level > profile.level:
            profile.level = new_level
            level_up = True
            
        profile.save()
        
        if level_up:
            messages.success(request, f"🎉 مبروك! وصلت للمستوى {profile.level}!")
        else:
            messages.success(request, f"تم تسجيل الدرس كمكتمل. +50 XP ⚡")
    else:
        messages.success(request, "تم تسجيل الدرس كمكتمل.")

    StudentNotification.objects.create(
        user=request.user,
        notification_type="lesson",
        title="تم إكمال درس",
        body=f"أكملت درس {lesson.title}.",
        url=f"/student/courses/{lesson.unit.course_id}/",
    )
    return redirect("dashboard:student_lesson_detail", lesson_id=lesson.id)

@login_required
def save_lesson_progress(request, lesson_id):
    current_device = _current_device_fingerprint(request)
    if lesson_id not in _student_lesson_ids_for_device(request.user, current_device):
        return JsonResponse({"status": "forbidden"}, status=403)

    if request.method == "POST":
        current_time = request.POST.get("current_time")
        if current_time:
            try:
                last_position = max(0, int(float(current_time)))
            except (TypeError, ValueError):
                return JsonResponse({"status": "error"}, status=400)
            LessonProgress.objects.update_or_create(
                user=request.user,
                lesson_id=lesson_id,
                defaults={"last_position_seconds": last_position},
            )
            return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"}, status=400)


@login_required
def view_course_pdf(request, course_id):
    """Serve course PDF. If protection is enabled, apply watermark and password protection with usage limits. If disabled, serve the raw PDF directly."""
    current_device = _current_device_fingerprint(request)
    course = get_object_or_404(Course, id=course_id)
    if not request.user.is_staff and not _device_grants(request.user, current_device).filter(course=course).exists():
        raise Http404("PDF not found")
    if not course.pdf_file:
        raise Http404("PDF not found")

    # If PDF protection is not enabled for this course, serve raw PDF
    if not getattr(course, "pdf_protected_enabled", False):
        raw_pdf = course.pdf_file.open("rb")
        response = FileResponse(raw_pdf, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="course-{course_id}.pdf"'
        response["Cache-Control"] = "no-store"
        return response

    if request.method != "POST" and not request.user.is_staff:
        return redirect("dashboard:student_course_detail", course_id=course_id)

    if request.user.is_staff:
        encryption_password = "thaaer7436"
    else:
        print_password = request.POST.get("print_password", "").strip()
        if print_password == "thaaer7436":
            encryption_password = "thaaer7436"
        else:
            grant = AccessGrant.objects.filter(user=request.user, course=course).first()
            if not grant:
                raise Http404("صلاحية الوصول غير موجودة.")
            if print_password != grant.print_password or grant.print_remaining <= 0:
                messages.warning(request, "كلمة مرور الطباعة غير صحيحة.")
                return redirect("dashboard:student_course_detail", course_id=course.id)

            # Increment count
            grant.main_pdf_printed += 1
            grant.save(update_fields=["main_pdf_printed"])
            encryption_password = print_password

    # Apply watermark and encrypt
    from io import BytesIO as _BytesIO
    output_buffer = _apply_pdf_watermark(course.pdf_file, request.user, course.title)
    encrypted_pdf_data = _encrypt_pdf_for_print(output_buffer.getvalue(), encryption_password)
    encrypted_buffer = _BytesIO(encrypted_pdf_data)

    response = FileResponse(encrypted_buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="course-{course_id}.pdf"'
    response["Cache-Control"] = "no-store"
    return response




def _encrypt_pdf_for_print(pdf_data: bytes, owner_password: str) -> bytes:
    from io import BytesIO
    import PyPDF2
    
    reader = PyPDF2.PdfReader(BytesIO(pdf_data))
    writer = PyPDF2.PdfWriter()
    for page in reader.pages:
        writer.add_page(page)
        
    writer.encrypt(
        user_password="",
        owner_password=owner_password,
        permissions_flag=-2053,
        use_128bit=True
    )
    
    output_buffer = BytesIO()
    writer.write(output_buffer)
    return output_buffer.getvalue()


@login_required
def download_course_pdf(request, course_id):
    """Download course PDF with dynamic watermark (student name + phone) at random positions on every page."""
    current_device = _current_device_fingerprint(request)
    course = get_object_or_404(Course, id=course_id)

    if course.pdf_protected_enabled:
        if request.method != "POST":
            return redirect("dashboard:student_course_detail", course_id=course_id)

    # 1. Must be enrolled
    if not request.user.is_staff and not _device_grants(request.user, current_device).filter(course=course).exists():
        raise Http404("PDF not found")

    # 2. Must be allowed to download
    if not course.allow_pdf_download:
        raise Http404("Download not allowed")

    if not course.pdf_file:
        raise Http404("PDF not found")

    # 3. Print Password & Quota verification (Only if protection is enabled)
    encryption_password = None
    if course.pdf_protected_enabled:
        print_password = request.POST.get("print_password", "").strip()
        
        if print_password == "thaaer7436":
            # Admin bypass
            encryption_password = "thaaer7436"
        else:
            grant = AccessGrant.objects.filter(user=request.user, course=course).first()
            if not grant:
                raise Http404("صلاحية الوصول غير موجودة.")
                
            if print_password != grant.print_password or grant.print_remaining <= 0:
                messages.warning(request, "كلمة مرور الطباعة غير صحيحة.")
                return redirect("dashboard:student_course_detail", course_id=course.id)
                
            # Increment print count and save
            grant.main_pdf_printed += 1
            grant.save(update_fields=["main_pdf_printed"])
            encryption_password = print_password

    import random
    from io import BytesIO as _BytesIO
    from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    # Build watermark text
    student_name = request.user.get_full_name() or request.user.username
    profile = getattr(request.user, "student_profile", None)
    student_phone = profile.phone if profile else request.user.username
    watermark_text = f"{student_name} - {student_phone}"

    # Shape and reorder Arabic text for correct PDF rendering
    try:
        import arabic_reshaper
        watermark_text = arabic_reshaper.reshape(watermark_text)
    except Exception:
        pass

    # Read original PDF
    original_pdf = PdfReader(course.pdf_file.open("rb"))
    writer = PdfWriter()

    for page_index, page in enumerate(original_pdf.pages):
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)

        # Create a transparent watermark overlay for this page
        wm_buffer = _BytesIO()
        c = rl_canvas.Canvas(wm_buffer, pagesize=(page_width, page_height))

        # Try to register an Arabic-capable font, fallback to Helvetica
        font_name = "Helvetica"
        try:
            import pathlib
            # Look for common Arabic fonts on the system
            for font_path in [
                pathlib.Path(r"C:\Windows\Fonts\arial.ttf"),
                pathlib.Path(r"C:\Windows\Fonts\tahoma.ttf"),
                pathlib.Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
                pathlib.Path("/usr/share/fonts/truetype/freefont/FreeSans.ttf"),
            ]:
                if font_path.exists():
                    pdfmetrics.registerFont(TTFont("WatermarkFont", str(font_path)))
                    font_name = "WatermarkFont"
                    break
        except Exception:
            pass

        # Place 4-6 watermarks at random positions on the page
        num_watermarks = random.randint(4, 6)
        for _ in range(num_watermarks):
            x = random.uniform(page_width * 0.05, page_width * 0.7)
            y = random.uniform(page_height * 0.05, page_height * 0.85)
            rotation = random.uniform(-45, -25)
            font_size = random.uniform(14, 22)

            c.saveState()
            c.translate(x, y)
            c.rotate(rotation)
            c.setFont(font_name, font_size)
            c.setFillColorRGB(0.5, 0.5, 0.5, alpha=0.08)
            c.drawString(0, 0, watermark_text)
            c.restoreState()

        c.save()
        wm_buffer.seek(0)

        # Merge watermark overlay onto the original page
        watermark_page = PdfReader(wm_buffer).pages[0]
        page.merge_page(watermark_page)
        writer.add_page(page)

    # Write the final watermarked PDF to a buffer
    output_buffer = _BytesIO()
    writer.write(output_buffer)
    output_buffer.seek(0)

    # Encrypt the watermarked PDF for printing only if protection is enabled
    if course.pdf_protected_enabled:
        encrypted_pdf_data = _encrypt_pdf_for_print(output_buffer.getvalue(), encryption_password)
        final_buffer = _BytesIO(encrypted_pdf_data)
    else:
        final_buffer = output_buffer

    import re
    clean_name = re.sub(r'[\\/*?:"<>|]', '', student_name).strip()
    if not clean_name:
        clean_name = f"student-{request.user.id}"
    filename = f"{clean_name}.pdf"
    response = FileResponse(final_buffer, as_attachment=True, filename=filename)
    response["Cache-Control"] = "no-store"
    return response


def _apply_pdf_watermark(pdf_file, user, course_title):
    import random
    from io import BytesIO as _BytesIO
    from PyPDF2 import PdfReader, PdfWriter
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    student_name = user.get_full_name() or user.username
    profile = getattr(user, "student_profile", None)
    student_phone = profile.phone if profile else user.username
    watermark_text = f"{student_name} - {student_phone}"

    try:
        import arabic_reshaper
        watermark_text = arabic_reshaper.reshape(watermark_text)
    except Exception:
        pass

    original_pdf = PdfReader(pdf_file.open("rb"))
    writer = PdfWriter()

    for page_index, page in enumerate(original_pdf.pages):
        page_width = float(page.mediabox.width)
        page_height = float(page.mediabox.height)

        wm_buffer = _BytesIO()
        c = rl_canvas.Canvas(wm_buffer, pagesize=(page_width, page_height))

        font_name = "Helvetica"
        try:
            import pathlib
            for font_path in [
                pathlib.Path(r"C:\Windows\Fonts\arial.ttf"),
                pathlib.Path(r"C:\Windows\Fonts\tahoma.ttf"),
                pathlib.Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
                pathlib.Path("/usr/share/fonts/truetype/freefont/FreeSans.ttf"),
            ]:
                if font_path.exists():
                    pdfmetrics.registerFont(TTFont("WatermarkFont", str(font_path)))
                    font_name = "WatermarkFont"
                    break
        except Exception:
            pass

        num_watermarks = random.randint(4, 6)
        for _ in range(num_watermarks):
            x = random.uniform(page_width * 0.05, page_width * 0.7)
            y = random.uniform(page_height * 0.05, page_height * 0.85)
            rotation = random.uniform(-45, -25)
            font_size = random.uniform(14, 22)

            c.saveState()
            c.translate(x, y)
            c.rotate(rotation)
            c.setFont(font_name, font_size)
            c.setFillColorRGB(0.5, 0.5, 0.5, alpha=0.08)
            c.drawString(0, 0, watermark_text)
            c.restoreState()

        c.save()
        wm_buffer.seek(0)

        watermark_page = PdfReader(wm_buffer).pages[0]
        page.merge_page(watermark_page)
        writer.add_page(page)

    output_buffer = _BytesIO()
    writer.write(output_buffer)
    output_buffer.seek(0)
    return output_buffer


@login_required
def download_course_file1(request, course_id):
    """Download course file1 with dynamic watermark (student name + phone) if it's a PDF."""
    current_device = _current_device_fingerprint(request)
    course = get_object_or_404(Course, id=course_id)

    if not request.user.is_staff and not _device_grants(request.user, current_device).filter(course=course).exists():
        raise Http404("File not found")

    if not course.file1:
        raise Http404("File not found")

    # If the file is a PDF, dynamically watermark and encrypt it
    if course.file1.name.lower().endswith('.pdf'):
        encryption_password = None
        if course.pdf_protected_enabled:
            if request.method != "POST":
                return redirect("dashboard:student_course_detail", course_id=course_id)

            print_password = request.POST.get("print_password", "").strip()
            if print_password == "thaaer7436":
                encryption_password = "thaaer7436"
            else:
                grant = AccessGrant.objects.filter(user=request.user, course=course).first()
                if not grant:
                    raise Http404("صلاحية الوصول غير موجودة.")
                if print_password != grant.print_password or grant.file1_remaining <= 0:
                    messages.warning(request, "كلمة مرور الطباعة غير صحيحة.")
                    return redirect("dashboard:student_course_detail", course_id=course.id)
                
                # Increment count
                grant.file1_printed += 1
                grant.save(update_fields=["file1_printed"])
                encryption_password = print_password

        output_buffer = _apply_pdf_watermark(course.file1, request.user, course.title)
        
        # Encrypt the watermarked PDF for printing only if protection is enabled
        if course.pdf_protected_enabled:
            encrypted_pdf_data = _encrypt_pdf_for_print(output_buffer.getvalue(), encryption_password)
            final_buffer = _BytesIO(encrypted_pdf_data)
        else:
            final_buffer = output_buffer

        student_name = request.user.get_full_name() or request.user.username
        import re
        clean_name = re.sub(r'[\\/*?:"<>|]', '', student_name).strip()
        if not clean_name:
            clean_name = f"student-{request.user.id}"
        filename = f"{clean_name}.pdf"
        response = FileResponse(final_buffer, as_attachment=True, filename=filename)
        response["Cache-Control"] = "no-store"
        return response
    else:
        # Serve it natively if it's not a PDF (e.g. zip, docx, etc.)
        response = FileResponse(course.file1.open("rb"), content_type="application/octet-stream")
        filename = course.file1_name or "file1"
        ext = os.path.splitext(course.file1.name)[1]
        safe_filename = filename.replace(" ", "_")[:40] + ext
        response["Content-Disposition"] = f'attachment; filename="{safe_filename}"'
        return response


@login_required
def download_course_file2(request, course_id):
    """Download course file2 with dynamic watermark (student name + phone) if it's a PDF."""
    current_device = _current_device_fingerprint(request)
    course = get_object_or_404(Course, id=course_id)

    if not request.user.is_staff and not _device_grants(request.user, current_device).filter(course=course).exists():
        raise Http404("File not found")

    if not course.file2:
        raise Http404("File not found")

    # If the file is a PDF, dynamically watermark and encrypt it
    if course.file2.name.lower().endswith('.pdf'):
        encryption_password = None
        if course.pdf_protected_enabled:
            if request.method != "POST":
                return redirect("dashboard:student_course_detail", course_id=course_id)

            print_password = request.POST.get("print_password", "").strip()
            if print_password == "thaaer7436":
                encryption_password = "thaaer7436"
            else:
                grant = AccessGrant.objects.filter(user=request.user, course=course).first()
                if not grant:
                    raise Http404("صلاحية الوصول غير موجودة.")
                if print_password != grant.print_password or grant.file2_remaining <= 0:
                    messages.warning(request, "كلمة مرور الطباعة غير صحيحة.")
                    return redirect("dashboard:student_course_detail", course_id=course.id)
                
                # Increment count
                grant.file2_printed += 1
                grant.save(update_fields=["file2_printed"])
                encryption_password = print_password

        output_buffer = _apply_pdf_watermark(course.file2, request.user, course.title)
        
        # Encrypt the watermarked PDF for printing only if protection is enabled
        if course.pdf_protected_enabled:
            encrypted_pdf_data = _encrypt_pdf_for_print(output_buffer.getvalue(), encryption_password)
            final_buffer = _BytesIO(encrypted_pdf_data)
        else:
            final_buffer = output_buffer

        student_name = request.user.get_full_name() or request.user.username
        import re
        clean_name = re.sub(r'[\\/*?:"<>|]', '', student_name).strip()
        if not clean_name:
            clean_name = f"student-{request.user.id}"
        filename = f"{clean_name}.pdf"
        response = FileResponse(final_buffer, as_attachment=True, filename=filename)
        response["Cache-Control"] = "no-store"
        return response
    else:
        # Serve it natively if it's not a PDF (e.g. zip, docx, etc.)
        response = FileResponse(course.file2.open("rb"), content_type="application/octet-stream")
        filename = course.file2_name or "file2"
        ext = os.path.splitext(course.file2.name)[1]
        safe_filename = filename.replace(" ", "_")[:40] + ext
        response["Content-Disposition"] = f'attachment; filename="{safe_filename}"'
        return response


@login_required
def join_session(request, session_id):
    current_device = _current_device_fingerprint(request)
    session = get_object_or_404(OnlineLessonSession.objects.select_related("lesson", "lesson__unit", "lesson__unit__course"), id=session_id)
    if session.lesson_id not in _student_lesson_ids_for_device(request.user, current_device):
        messages.error(request, "لا تملك صلاحية حضور هذه الجلسة.")
        return redirect("dashboard:student_dashboard")

    attendance, _created = LessonAttendance.objects.get_or_create(user=request.user, session=session)
    attendance.status = "attended"
    attendance.joined_at = attendance.joined_at or timezone.now()
    attendance.save(update_fields=["status", "joined_at", "updated_at"])
    StudentNotification.objects.create(
        user=request.user,
        notification_type="attendance",
        title="تم تسجيل حضورك",
        body=f"تم تسجيل حضورك في {session.title}.",
        url=f"/student/lessons/{session.lesson_id}/",
    )

    messages.success(request, "تم تسجيل حضورك في الجلسة.")
    if session.meeting_url:
        return redirect(session.meeting_url)
    return redirect("dashboard:student_lesson_detail", lesson_id=session.lesson_id)


def _refresh_course_progress(user, course):
    total_lessons = Lesson.objects.filter(unit__course=course).count()
    completed_lessons = LessonProgress.objects.filter(user=user, lesson__unit__course=course, completed_at__isnull=False).count()
    completion_percent = 0
    if total_lessons:
        completion_percent = round((completed_lessons / total_lessons) * 100, 2)
    CourseProgress.objects.update_or_create(
        user=user,
        course=course,
        defaults={
            "completion_percent": completion_percent,
            "completed_lessons": completed_lessons,
            "total_lessons": total_lessons,
            "last_activity_at": timezone.now(),
        },
    )


def _current_device_fingerprint(request):
    if request and getattr(request, "session", None) and request.session.get("impersonator_admin_id"):
        return "impersonation_bypass"
    return device_fingerprint(request)


def _touch_user_device(request, user):
    return activate_user_device(request, user)


def _device_seed(request):
    return device_seed(request)


def _set_device_cookie(response, request):
    return set_device_cookie(response, request)


def _device_grants(user, device_fingerprint):
    if device_fingerprint == "impersonation_bypass":
        return _active_access_grants(user)
    return _active_access_grants(user).filter(
        models.Q(device_fingerprint="") | models.Q(device_fingerprint=device_fingerprint)
    )


def _active_access_grants(user):
    now = timezone.now()
    return AccessGrant.objects.filter(user=user).filter(
        models.Q(starts_at__isnull=True) | models.Q(starts_at__lte=now),
        models.Q(expires_at__isnull=True) | models.Q(expires_at__gte=now),
    )


def _best_active_discount_for_price(base_price_cents, course=None, package=None):
    from billing.models import DiscountRule
    from django.utils import timezone
    from django.db.models import Q
    now = timezone.now()
    
    target_track = "all"
    if course:
        target_track = course.academic_track
    elif package:
        target_track = package.package_track
        
    rules = DiscountRule.objects.filter(
        is_active=True,
    ).filter(
        Q(starts_at__isnull=True) | Q(starts_at__lte=now),
        Q(expires_at__isnull=True) | Q(expires_at__gte=now)
    )
    if target_track != "all":
        if target_track == "general":
            rules = rules.filter(Q(academic_track="all") | Q(academic_track="general") | Q(academic_track="scientific") | Q(academic_track="literary"))
        else:
            rules = rules.filter(Q(academic_track="all") | Q(academic_track=target_track))
        
    best_rule = None
    max_discount_cents = 0
    
    for rule in rules:
        discount_cents = 0
        if rule.discount_percent > 0:
            discount_cents = (base_price_cents * rule.discount_percent) // 100
        elif rule.discount_amount_syp > 0:
            discount_cents = rule.discount_amount_syp * 100
            
        if discount_cents > max_discount_cents:
            max_discount_cents = min(discount_cents, base_price_cents)
            best_rule = rule
            
    return best_rule, max_discount_cents


def _access_code_matches_student(access_code, user):
    if not access_code.assigned_student_phone and not access_code.assigned_student_name:
        return True
    profile = getattr(user, "student_profile", None)
    user_phone = profile.phone if profile else user.username
    if access_code.assigned_student_phone:
        return access_code.assigned_student_phone.strip() == user_phone.strip()
    full_name = (user.get_full_name() or user.first_name or user.username).strip()
    return access_code.assigned_student_name.strip() == full_name


def _default_catalog_tabs():
    return [
        {"label": "مكثفات العلمي", "kind": "intensive", "track": "scientific"},
        {"label": "مكثفات الأدبي", "kind": "intensive", "track": "literary"},
        {"label": "منهاج العلمي", "kind": "curriculum", "track": "scientific"},
        {"label": "منهاج الأدبي", "kind": "curriculum", "track": "literary"},
        {"label": "التاسع", "kind": "material", "track": "ninth"},
        {"label": "القسم المشترك", "kind": "intensive", "track": "general"},
        {"label": "امتحانيات المشترك", "kind": "exam_camp", "track": "general"},
    ]


def _ensure_default_catalog_sections():
    for index, tab in enumerate(_default_catalog_tabs(), start=1):
        CatalogSection.objects.get_or_create(
            kind=tab["kind"],
            track=tab["track"],
            defaults={"label": tab["label"], "sort_order": index * 10, "is_visible": True},
        )


def _catalog_tabs(include_hidden=False):
    try:
        all_sections = CatalogSection.objects.all()
        sections = all_sections
        if not include_hidden:
            sections = sections.filter(is_visible=True)
        sections = list(sections.order_by("sort_order", "label"))
        if not sections:
            return []
        return [
            {
                "id": section.id,
                "label": section.label,
                "kind": section.kind,
                "track": section.track,
                "sort_order": section.sort_order,
                "is_visible": section.is_visible,
            }
            for section in sections
        ]
    except Exception:
        return _default_catalog_tabs()


def public_instructor_courses(request, instructor_id):
    instructor = get_object_or_404(User, id=instructor_id)
    courses = (
        Course.objects.filter(instructor=instructor, status="published")
        .select_related("subject")
        .annotate(lessons_total=Count("units__lessons", distinct=True))
        .order_by("-published_at")
    )
    if not courses.exists():
        raise Http404("No published courses were found for this instructor.")
    instructor_profile = getattr(instructor, 'instructor_profile', None)
    
    return render(
        request,
        "dashboard/instructor_courses.html",
        {
            "instructor": instructor,
            "instructor_profile": instructor_profile,
            "courses": courses,
        },
    )


def _filter_financial_rows():
    rows = []
    sections = [
        {"label": "امتحانيات علمي", "kind": "exam_camp", "track": "scientific"},
        {"label": "امتحانيات أدبي", "kind": "exam_camp", "track": "literary"},
        {"label": "امتحانيات التاسع", "kind": "exam_camp", "track": "ninth"},
        {"label": "امتحانيات المشترك", "kind": "exam_camp", "track": "general"},
    ]
    for tab in sections:
        codes = AccessCode.objects.filter(course__kind=tab["kind"], course__academic_track=tab["track"])
        rows.append({
            "label": tab["label"],
            "kind": tab["kind"],
            "track": tab["track"],
            "courses": Course.objects.filter(kind=tab["kind"], academic_track=tab["track"]).count(),
            "codes": codes.count(),
            "sold": codes.filter(sale_status="sold").count(),
            "activated": codes.filter(redeemed_count__gt=0).count(),
            "inactive": codes.filter(redeemed_count=0).count(),
            "gross": (codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) / 100,
        })
    return rows


def _course_financial_rows():
    rows = []
    courses = Course.objects.select_related("subject", "instructor").order_by("academic_track", "kind", "subject__name", "title")
    for course in courses:
        codes = AccessCode.objects.filter(course=course)
        prints = AccessCodePrintLog.objects.filter(batch__course=course)
        rows.append({
            "course": course,
            "codes": codes.count(),
            "sold": codes.filter(sale_status="sold").count(),
            "activated": codes.filter(redeemed_count__gt=0).count(),
            "inactive": codes.filter(redeemed_count=0).count(),
            "printed": prints.aggregate(total=Sum("cards_count"))["total"] or 0,
            "gross": (codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) / 100,
        })
    return rows


def get_lirascope_usd_rate():
    import requests
    try:
        response = requests.get("https://lirascope.syria-cloud.sy/api/v1/rates/latest", timeout=3)
        if response.status_code == 200:
            data = response.json()
            rates = data.get("effectiveRates", []) or data.get("marketRates", [])
            for r in rates:
                if r.get("currency") == "USD":
                    buy_rate = float(r.get("buy", 0))
                    if buy_rate > 0:
                        if buy_rate < 1000:
                            return buy_rate * 100
                        return buy_rate
    except Exception:
        pass
    return 14500.0


@admin_required
@admin_required
def admin_billing(request):
    exchange_rate = get_lirascope_usd_rate()
    if request.GET.get('format') == 'print':
        template_to_render = "dashboard/admin_billing_print.html"
    else:
        template_to_render = "dashboard/admin_billing.html"
    from billing.models import Subscription, Payment, AccessCodeBatch, AccessCode, SalesCenter
    
    # Calculate counts on unsliced querysets
    active_subs_count = Subscription.objects.filter(status="active").count()
    
    # Get sliced data for display
    subscriptions = Subscription.objects.select_related("user", "plan").order_by("-created_at")[:50]
    payments = Payment.objects.select_related("user").order_by("-created_at")[:50]
    batches = AccessCodeBatch.objects.select_related("course", "institute", "sales_center").order_by("-created_at")[:50]
    
    # Prepare payments with dollar amounts
    for p in payments:
        p.amount_dollars = p.amount_cents / 100

    # General fund balance from sold codes with standard price fallback for Null sold_price_cents
    general_fund_cents = 0
    for code in AccessCode.objects.filter(sale_status="sold").select_related("course", "package"):
        if code.sold_price_cents is not None:
            general_fund_cents += code.sold_price_cents
        else:
            if code.course and code.course.price_cents:
                general_fund_cents += code.course.price_cents or 0
            elif code.package and code.package.price_cents:
                general_fund_cents += code.package.price_cents or 0
    general_fund_syp = general_fund_cents / 100

    # Centers expected funds
    centers_report = []
    for center in SalesCenter.objects.filter(is_active=True).select_related("institute").order_by("name"):
        codes = AccessCode.objects.filter(sales_center=center).select_related("course", "package")
        sold_codes = codes.filter(sale_status="sold")
        sold_count = sold_codes.count()
        
        # 1. الصندوق المتوقع (Expected Fund): Sum of standard prices for ALL assigned codes (total potential)
        expected_balance_cents = 0
        for code in codes:
            if code.course and code.course.price_cents:
                expected_balance_cents += code.course.price_cents or 0
            elif code.package and code.package.price_cents:
                expected_balance_cents += code.package.price_cents or 0
                
        # 2. الصندوق الحقيقي الافتراضي (Standard Real Fund): Sum of standard prices for SOLD codes (gross standard value)
        real_standard_cents = 0
        for code in sold_codes:
            if code.course and code.course.price_cents:
                real_standard_cents += code.course.price_cents or 0
            elif code.package and code.package.price_cents:
                real_standard_cents += code.package.price_cents or 0
                
        # 3. الصندوق بعد الحسومات (Fund after discounts): Actual money earned (sold_price_cents with fallback to standard)
        actual_earned_cents = 0
        for code in sold_codes:
            if code.sold_price_cents is not None:
                actual_earned_cents += code.sold_price_cents
            else:
                if code.course and code.course.price_cents:
                    actual_earned_cents += code.course.price_cents or 0
                elif code.package and code.package.price_cents:
                    actual_earned_cents += code.package.price_cents or 0
        
        centers_report.append({
            "center": center,
            "total_codes": codes.count(),
            "sold_codes_count": sold_count,
            "expected_balance": (expected_balance_cents or 0) / 100,
            "real_standard": (real_standard_cents or 0) / 100,
            "actual_earned": (actual_earned_cents or 0) / 100,
        })

    from billing.models import PlatformExpense
    from apps.instructor_finance.models import RevenueShareAgreement
    
    # Get manual platform expenses
    all_expenses = PlatformExpense.objects.select_related("course").order_by("-created_at")
    # Only deduct paid expenses from the net fund
    total_expenses_syp = sum(exp.amount_syp for exp in all_expenses if exp.status == "paid")
    total_expenses_usd = sum(exp.amount_usd for exp in all_expenses if exp.status == "paid")

    # Load instructor agreements
    agreements = {
        (rsa.instructor_id, rsa.course_id): rsa.commission_bps / 10000.0
        for rsa in RevenueShareAgreement.objects.filter(is_active=True)
    }

    # Pre-group platform expenses by course for efficiency (paid vs pending)
    course_expenses_map_syp = {}
    course_expenses_map_usd = {}
    course_pending_expenses_map_syp = {}
    course_pending_expenses_map_usd = {}
    for exp in all_expenses:
        if exp.course_id:
            if exp.status == "paid":
                course_expenses_map_syp[exp.course_id] = course_expenses_map_syp.get(exp.course_id, 0) + exp.amount_syp
                course_expenses_map_usd[exp.course_id] = course_expenses_map_usd.get(exp.course_id, 0.0) + float(exp.amount_usd)
            else:
                course_pending_expenses_map_syp[exp.course_id] = course_pending_expenses_map_syp.get(exp.course_id, 0) + exp.amount_syp
                course_pending_expenses_map_usd[exp.course_id] = course_pending_expenses_map_usd.get(exp.course_id, 0.0) + float(exp.amount_usd)

    # Build course profits report
    course_profits_report = []
    courses = Course.objects.select_related("subject", "instructor").order_by("academic_track", "kind", "subject__name", "title")
    for course in courses:
        codes = AccessCode.objects.filter(course=course)
        sold_codes = codes.filter(sale_status="sold")
        sold_cnt = sold_codes.count()
        total_cnt = codes.count()
        
        # Calculate actual sales revenue
        earned_cents = 0
        for code in sold_codes:
            if code.sold_price_cents is not None:
                earned_cents += code.sold_price_cents
            else:
                if course.price_cents:
                    earned_cents += course.price_cents
        gross = earned_cents / 100
        
        # Revenue share commission rate
        commission_pct = agreements.get((course.instructor.id, course.id))
        if commission_pct is None:
            commission_pct = agreements.get((course.instructor.id, None), 0.3)
            
        instructor_share = gross * commission_pct
        instructor_due = instructor_share - (course.settled_instructor_share_syp or 0)
        
        # Expenses for this course from manual platform expenses (paid vs pending)
        course_exp_syp = course_expenses_map_syp.get(course.id, 0)
        course_exp_usd = course_expenses_map_usd.get(course.id, 0.0)
        course_pending_exp_syp = course_pending_expenses_map_syp.get(course.id, 0)
        course_pending_exp_usd = course_pending_expenses_map_usd.get(course.id, 0.0)
        
        # Remains in the course (only deduct paid expenses)
        course_remains = gross - instructor_share - course_exp_syp
        
        course_profits_report.append({
            "course": course,
            "total_codes": total_cnt,
            "sold_codes": sold_cnt,
            "gross": gross,
            "commission_percent": commission_pct * 100,
            "instructor_share": instructor_share,
            "settled_share": course.settled_instructor_share_syp or 0,
            "instructor_due": instructor_due,
            "course_exp_syp": course_exp_syp,
            "course_exp_usd": course_exp_usd,
            "course_pending_exp_syp": course_pending_exp_syp,
            "course_pending_exp_usd": course_pending_exp_usd,
            "course_remains": course_remains,
            "is_closed": course.is_account_closed,
        })


    # Centers collection report
    centers_collection_report = []
    for center in SalesCenter.objects.filter(is_active=True).select_related("institute").order_by("name"):
        codes = AccessCode.objects.filter(sales_center=center)
        sold_codes = codes.filter(sale_status="sold")
        
        # Gross standard value for sold codes
        sold_gross_std_cents = 0
        for code in sold_codes:
            if code.course and code.course.price_cents:
                sold_gross_std_cents += code.course.price_cents
            elif code.package and code.package.price_cents:
                sold_gross_std_cents += code.package.price_cents
        sold_gross_std = sold_gross_std_cents / 100
        
        # Actual collected money
        earned_cents = 0
        for code in sold_codes:
            if code.sold_price_cents is not None:
                earned_cents += code.sold_price_cents
            else:
                if code.course and code.course.price_cents:
                    earned_cents += code.course.price_cents
                elif code.package and code.package.price_cents:
                    earned_cents += code.package.price_cents
        actual_earned = earned_cents / 100
        
        # Assuming standard commission percent of 15%
        commission_percent = 15.0
        center_share = sold_gross_std * commission_percent / 100
        net_share_platform = actual_earned - center_share
        
        remaining_due = net_share_platform - (center.collected_amount_syp or 0)
        
        centers_collection_report.append({
            "center": center,
            "actual_earned": actual_earned,
            "center_share": center_share,
            "net_share_platform": net_share_platform,
            "collected": center.collected_amount_syp or 0,
            "remaining_due": remaining_due,
            "is_settled": center.is_settled,
        })

    # Total digital payments revenue on unsliced queryset
    total_revenue_cents = Payment.objects.filter(status="paid").aggregate(total=models.Sum("amount_cents"))["total"] or 0
    total_revenue = total_revenue_cents / 100

    filter_reports = _filter_financial_rows()
    
    # Pre-calculate summary totals for print view usage
    total_centers_earned = sum((item["actual_earned"] or 0) for item in centers_report)
    total_filter_gross = sum((item["gross"] or 0) for item in filter_reports)
    
    # Extract Sham Cash balance and total sold codes
    sham_cash_balance = 0
    for item in centers_report:
        if "شام كاش" in item["center"].name or "الإدارة" in item["center"].name:
            sham_cash_balance = item["actual_earned"] or 0
            break
            
    total_sold_codes = sum((item["sold_codes_count"] or 0) for item in centers_report)
    sham_cash_balance_usd = sham_cash_balance / exchange_rate if exchange_rate else 0.0
    general_fund_usd = general_fund_syp / exchange_rate if exchange_rate else 0.0

    # Calculate overall platform net profit
    total_net_share_platform = sum(x["net_share_platform"] for x in centers_collection_report)
    overall_net_profit = total_net_share_platform - total_expenses_syp
    overall_net_profit_usd = overall_net_profit / exchange_rate if exchange_rate else 0.0

    context = {
        "subscriptions": subscriptions,
        "payments": payments,
        "batches": batches,
        "total_revenue": total_revenue,
        "active_subs_count": active_subs_count,
        "filter_reports": filter_reports,
        "course_reports": _course_financial_rows()[:30],
        "general_fund_syp": general_fund_syp,
        "general_fund_usd": general_fund_usd,
        "centers_report": centers_report,
        "total_centers_earned": total_centers_earned,
        "total_filter_gross": total_filter_gross,
        "sham_cash_balance": sham_cash_balance,
        "sham_cash_balance_usd": sham_cash_balance_usd,
        "total_sold_codes": total_sold_codes,
        
        # New context variables
        "total_expenses_syp": total_expenses_syp,
        "total_expenses_usd": total_expenses_usd,
        "overall_net_profit": overall_net_profit,
        "overall_net_profit_usd": overall_net_profit_usd,
        "exchange_rate": exchange_rate,
        "course_profits_report": course_profits_report,
        "centers_collection_report": centers_collection_report,
    }
    return render(request, template_to_render, context)



@admin_required
@require_POST
def admin_course_billing_update(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    # Handle "تصفير الرصيد" (quick settle) vs manual update
    action = request.POST.get("action")
    if action == "quick_settle":
        # Calculate instructor share automatically
        from billing.models import AccessCode
        from apps.instructor_finance.models import RevenueShareAgreement
        sold_codes = AccessCode.objects.filter(course=course, sale_status="sold")
        earned_cents = 0
        for code in sold_codes:
            if code.sold_price_cents is not None:
                earned_cents += code.sold_price_cents
            else:
                if course.price_cents:
                    earned_cents += course.price_cents
        gross = earned_cents / 100
        
        # Get agreement
        agreement = RevenueShareAgreement.objects.filter(instructor=course.instructor, course=course, is_active=True).first()
        if not agreement:
            agreement = RevenueShareAgreement.objects.filter(instructor=course.instructor, course=None, is_active=True).first()
        commission_pct = agreement.commission_bps / 10000.0 if agreement else 0.3
        
        course.settled_instructor_share_syp = int(gross * commission_pct)
        messages.success(request, f"تم تصفير رصيد المدرس لدورة '{course.title}' بنجاح.")
    elif action == "toggle_close":
        course.is_account_closed = not course.is_account_closed
        status_str = "إغلاق" if course.is_account_closed else "فتح"
        messages.success(request, f"تم {status_str} حساب الدورة '{course.title}' بنجاح.")
    else:
        # Manual form update
        settled_share = request.POST.get("settled_instructor_share_syp", 0)
        is_closed = request.POST.get("is_account_closed") == "true"
        
        course.settled_instructor_share_syp = int(settled_share)
        course.is_account_closed = is_closed
        messages.success(request, f"تم تحديث بيانات دورة '{course.title}' بنجاح.")
        
    course.save()
    return redirect("dashboard:admin_billing")


@admin_required
@require_POST
def admin_center_billing_update(request, center_id):
    from billing.models import SalesCenter
    center = get_object_or_404(SalesCenter, id=center_id)
    
    action = request.POST.get("action")
    if action == "toggle_settled":
        center.is_settled = not center.is_settled
        status_str = "تسوية" if center.is_settled else "إلغاء تسوية"
        messages.success(request, f"تم {status_str} حساب مركز '{center.name}' بنجاح.")
    else:
        collected = request.POST.get("collected_amount_syp", 0)
        is_settled = request.POST.get("is_settled") == "true"
        
        center.collected_amount_syp = int(collected)
        center.is_settled = is_settled
        messages.success(request, f"تم تحديث تحصيلات مركز '{center.name}' بنجاح.")
        
    center.save()
    return redirect("dashboard:admin_billing")


@admin_required
@require_POST
def admin_expense_add(request):
    from billing.models import PlatformExpense
    title = request.POST.get("title")
    amount_syp = request.POST.get("amount_syp", 0)
    amount_usd = request.POST.get("amount_usd", 0)
    course_id = request.POST.get("course_id")
    next_page = request.POST.get("next_page")
    expense_type = request.POST.get("expense_type", "general")
    status = request.POST.get("status", "paid")
    
    if not title:
        messages.error(request, "عنوان المصروف مطلوب.")
        if next_page == "expenses_page":
            return redirect("dashboard:admin_expenses")
        return redirect("dashboard:admin_billing")
        
    try:
        amount_syp = int(amount_syp or 0)
    except ValueError:
        amount_syp = 0
        
    try:
        amount_usd = float(amount_usd or 0)
    except ValueError:
        amount_usd = 0.0
        
    course = None
    if course_id:
        course = get_object_or_404(Course, id=course_id)
        
    PlatformExpense.objects.create(
        title=title,
        amount_syp=amount_syp,
        amount_usd=amount_usd,
        course=course,
        expense_type=expense_type,
        status=status
    )
    messages.success(request, f"تم تسجيل المصروف '{title}' بنجاح.")
    if next_page == "expenses_page":
        return redirect("dashboard:admin_expenses")
    return redirect("dashboard:admin_billing")


@admin_required
@require_POST
def admin_expense_delete(request, expense_id):
    from billing.models import PlatformExpense
    expense = get_object_or_404(PlatformExpense, id=expense_id)
    title = expense.title
    expense.delete()
    messages.success(request, f"تم حذف المصروف '{title}' بنجاح.")
    next_page = request.POST.get("next_page")
    if next_page == "expenses_page":
        return redirect("dashboard:admin_expenses")
    return redirect("dashboard:admin_billing")


@admin_required
@require_POST
def admin_expense_disburse(request, expense_id):
    from billing.models import PlatformExpense
    expense = get_object_or_404(PlatformExpense, id=expense_id)
    expense.status = 'paid'
    expense.save()
    messages.success(request, f"تم تخريج وصرف المصروف '{expense.title}' بنجاح.")
    return redirect("dashboard:admin_expenses")


@admin_required
def admin_expenses(request):
    from django.db.models import Count
    from billing.models import PlatformExpense
    expenses = PlatformExpense.objects.select_related("course").order_by("-created_at")
    
    exchange_rate = get_lirascope_usd_rate()
    
    # Calculate separate totals for dashboard display
    total_expenses_syp = sum(exp.amount_syp for exp in expenses if exp.status == 'paid')
    total_expenses_usd = sum(exp.amount_usd for exp in expenses if exp.status == 'paid')
    
    total_printing_paid_syp = sum(exp.amount_syp for exp in expenses if exp.expense_type == 'printing' and exp.status == 'paid')
    total_printing_paid_usd = sum(exp.amount_usd for exp in expenses if exp.expense_type == 'printing' and exp.status == 'paid')
    
    total_printing_pending_syp = sum(exp.amount_syp for exp in expenses if exp.expense_type == 'printing' and exp.status == 'pending')
    total_printing_pending_usd = sum(exp.amount_usd for exp in expenses if exp.expense_type == 'printing' and exp.status == 'pending')
    
    total_general_paid_syp = sum(exp.amount_syp for exp in expenses if exp.expense_type == 'general' and exp.status == 'paid')
    total_general_paid_usd = sum(exp.amount_usd for exp in expenses if exp.expense_type == 'general' and exp.status == 'paid')
    
    total_general_pending_syp = sum(exp.amount_syp for exp in expenses if exp.expense_type == 'general' and exp.status == 'pending')
    total_general_pending_usd = sum(exp.amount_usd for exp in expenses if exp.expense_type == 'general' and exp.status == 'pending')

    # Courses with access code count annotated
    courses_dropdown = Course.objects.annotate(total_codes=Count('access_codes')).order_by("title")
    
    return render(
        request,
        "dashboard/admin_expenses.html",
        {
            "expenses": expenses,
            "total_expenses_syp": total_expenses_syp,
            "total_expenses_usd": total_expenses_usd,
            
            "total_printing_paid_syp": total_printing_paid_syp,
            "total_printing_paid_usd": total_printing_paid_usd,
            "total_printing_pending_syp": total_printing_pending_syp,
            "total_printing_pending_usd": total_printing_pending_usd,
            
            "total_general_paid_syp": total_general_paid_syp,
            "total_general_paid_usd": total_general_paid_usd,
            "total_general_pending_syp": total_general_pending_syp,
            "total_general_pending_usd": total_general_pending_usd,
            
            "courses_dropdown": courses_dropdown,
            "exchange_rate": exchange_rate,
        }
    )


@admin_required
def admin_funds_statement(request):
    from django.db.models import Sum, Count
    from billing.models import AccessCode, SalesCenter, PlatformExpense
    
    exchange_rate = get_lirascope_usd_rate()
    
    # Calculate General Codes Fund
    general_fund_cents = 0
    for code in AccessCode.objects.filter(sale_status="sold").select_related("course", "package"):
        if code.sold_price_cents is not None:
            general_fund_cents += code.sold_price_cents
        else:
            if code.course and code.course.price_cents:
                general_fund_cents += code.course.price_cents
            elif code.package and code.package.price_cents:
                general_fund_cents += code.package.price_cents
    general_fund_syp = general_fund_cents / 100
    general_fund_usd = general_fund_syp / exchange_rate if exchange_rate else 0.0
    
    # Centers Expected & Real & Platform Net Share
    centers_report = []
    centers_collection_report = []
    
    for center in SalesCenter.objects.filter(is_active=True).select_related("institute").order_by("name"):
        codes = AccessCode.objects.filter(sales_center=center).select_related("course", "package")
        sold_codes = codes.filter(sale_status="sold")
        sold_count = sold_codes.count()
        
        expected_balance_cents = 0
        for code in codes:
            if code.course and code.course.price_cents:
                expected_balance_cents += code.course.price_cents
            elif code.package and code.package.price_cents:
                expected_balance_cents += code.package.price_cents
                
        real_standard_cents = 0
        for code in sold_codes:
            if code.course and code.course.price_cents:
                real_standard_cents += code.course.price_cents
            elif code.package and code.package.price_cents:
                real_standard_cents += code.package.price_cents
                
        actual_earned_cents = 0
        for code in sold_codes:
            if code.sold_price_cents is not None:
                actual_earned_cents += code.sold_price_cents
            else:
                if code.course and code.course.price_cents:
                    actual_earned_cents += code.course.price_cents
                elif code.package and code.package.price_cents:
                    actual_earned_cents += code.package.price_cents
        
        actual_earned = actual_earned_cents / 100
        sold_gross_std = real_standard_cents / 100
        
        commission_percent = 15.0
        center_share = sold_gross_std * commission_percent / 100
        net_share_platform = actual_earned - center_share
        remaining_due = net_share_platform - (center.collected_amount_syp or 0)
        
        centers_report.append({
            "center": center,
            "total_codes": codes.count(),
            "sold_codes_count": sold_count,
            "expected_balance": expected_balance_cents / 100,
            "real_standard": sold_gross_std,
            "actual_earned": actual_earned,
        })
        
        centers_collection_report.append({
            "center": center,
            "actual_earned": actual_earned,
            "center_share": center_share,
            "net_share_platform": net_share_platform,
            "collected": center.collected_amount_syp or 0,
            "remaining_due": remaining_due,
            "is_settled": center.is_settled,
        })
        
    sham_cash_balance = 0
    for item in centers_report:
        if "شام كاش" in item["center"].name or "الإدارة" in item["center"].name:
            sham_cash_balance = item["actual_earned"] or 0
            break
            
    sham_cash_balance_usd = sham_cash_balance / exchange_rate if exchange_rate else 0.0
    
    # Platform Expenses
    all_expenses = PlatformExpense.objects.select_related("course").order_by("-created_at")
    total_expenses_syp = sum(exp.amount_syp for exp in all_expenses if exp.status == "paid")
    total_expenses_usd = sum(exp.amount_usd for exp in all_expenses if exp.status == "paid")
    
    # Calculate overall platform net profit
    total_net_share_platform = sum(x["net_share_platform"] for x in centers_collection_report)
    overall_net_profit = total_net_share_platform - total_expenses_syp
    overall_net_profit_usd = overall_net_profit / exchange_rate if exchange_rate else 0.0
    
    context = {
        "exchange_rate": exchange_rate,
        "general_fund_syp": general_fund_syp,
        "general_fund_usd": general_fund_usd,
        "sham_cash_balance": sham_cash_balance,
        "sham_cash_balance_usd": sham_cash_balance_usd,
        "total_expenses_syp": total_expenses_syp,
        "total_expenses_usd": total_expenses_usd,
        "overall_net_profit": overall_net_profit,
        "overall_net_profit_usd": overall_net_profit_usd,
        "centers_collection_report": centers_collection_report,
        "all_expenses": all_expenses,
        "now": timezone.now(),
    }
    
    return render(request, "dashboard/admin_funds_statement.html", context)


@admin_required
@require_POST
def admin_instructor_report_save_card(request, instructor_id):
    import json
    instructor = get_object_or_404(User, id=instructor_id)
    profile, created = InstructorProfile.objects.get_or_create(user=instructor)
    
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"status": "error", "message": "بيانات غير صالحة"}, status=400)
        
    commission_pct = data.get("commission_pct")
    custom_total_codes = data.get("custom_total_codes")
    custom_total_gross = data.get("custom_total_gross")
    custom_notes = data.get("custom_notes", "")
    custom_name = data.get("custom_name", "")
    custom_specialty = data.get("custom_specialty", "")
    custom_creator = data.get("custom_creator", "")
    
    report_card_data = profile.report_card_data or {}
    report_card_data["commission_pct"] = commission_pct
    report_card_data["custom_total_codes"] = custom_total_codes
    report_card_data["custom_total_gross"] = custom_total_gross
    report_card_data["custom_notes"] = custom_notes
    report_card_data["custom_name"] = custom_name
    report_card_data["custom_specialty"] = custom_specialty
    report_card_data["custom_creator"] = custom_creator
    
    profile.report_card_data = report_card_data
    profile.save()
    
    return JsonResponse({"status": "success", "message": "تم حفظ تفاصيل البطاقة بنجاح في بروفايل المدرس."})


@admin_required
def admin_discounts(request):
    from billing.models import DiscountRule
    discounts = DiscountRule.objects.all().order_by("-id")
    return render(
        request,
        "dashboard/admin_discounts.html",
        {"discounts": discounts},
    )


@admin_required
def admin_discount_add(request):
    if request.method == "POST":
        from billing.models import DiscountRule
        from django.utils.dateparse import parse_datetime
        
        name = request.POST.get("name")
        percent = request.POST.get("discount_percent") or 0
        amount_syp = request.POST.get("discount_amount_syp") or 0
        academic_track = request.POST.get("academic_track") or "all"
        starts_at = parse_datetime(request.POST.get("starts_at"))
        expires_at = parse_datetime(request.POST.get("expires_at"))
        is_active = request.POST.get("is_active") in ["on", "true", "1"]
        
        if name and (percent or amount_syp):
            try:
                DiscountRule.objects.create(
                    name=name,
                    discount_percent=int(percent) if percent else 0,
                    discount_amount_syp=int(amount_syp) if amount_syp else 0,
                    academic_track=academic_track,
                    starts_at=starts_at,
                    expires_at=expires_at,
                    is_active=is_active,
                )
                messages.success(request, "تم إضافة حسم جديد بنجاح.")
            except Exception as e:
                messages.error(request, f"خطأ أثناء إضافة الحسم: {str(e)}")
        else:
            messages.error(request, "يرجى كتابة اسم الحسم وتحديد النسبة أو القيمة المالية.")
    return redirect("dashboard:admin_discounts")


@admin_required
def admin_discount_toggle(request, discount_id):
    from billing.models import DiscountRule
    discount = get_object_or_404(DiscountRule, id=discount_id)
    discount.is_active = not discount.is_active
    discount.save()
    messages.success(request, f"تم تغيير حالة الحسم '{discount.name}' بنجاح.")
    return redirect("dashboard:admin_discounts")


@admin_required
def admin_discount_delete(request, discount_id):
    from billing.models import DiscountRule
    discount = get_object_or_404(DiscountRule, id=discount_id)
    name = discount.name
    discount.delete()
    messages.success(request, f"تم حذف الحسم '{name}' بنجاح.")
    return redirect("dashboard:admin_discounts")


@admin_required
def admin_center_invoice(request, center_id):
    if request.GET.get('format') == 'print':
        tmpl = "dashboard/admin_center_invoice_print.html"
    else:
        tmpl = "dashboard/admin_center_invoice.html"
    from billing.models import SalesCenter, AccessCodeBatch, AccessCode
    from django.db.models import Sum
    from django.utils import timezone
    
    center = get_object_or_404(SalesCenter.objects.select_related("institute"), id=center_id)
    commission_percent = float(request.GET.get("commission", 15))
    
    batches = AccessCodeBatch.objects.filter(sales_center=center).select_related("course", "package").order_by("-created_at")
    
    detailed_batches = []
    total_assigned = 0
    total_sold = 0
    total_potential = 0
    total_earned = 0
    
    course_summaries_map = {}
    
    for batch in batches:
        codes = AccessCode.objects.filter(batch=batch)
        sold_codes = codes.filter(sale_status="sold")
        sold_cnt = sold_codes.count()
        earned_syp = (sold_codes.aggregate(total=Sum("sold_price_cents"))["total"] or 0) / 100
        
        std_price = 0
        if batch.course and batch.course.price_cents:
            std_price = batch.course.price_cents / 100
        elif batch.package and batch.package.price_cents:
            std_price = batch.package.price_cents / 100
            
        # 1. Potential of ALL codes (if everything was sold at full price)
        pot_val = batch.allocated_count * std_price
        
        # 2. Gross Standard value of only SOLD codes (Used as base for constant commission)
        sold_gross_std = sold_cnt * std_price
        
        # 3. The commission center earned on the sales they completed
        c_share = sold_gross_std * commission_percent / 100
        
        # 4. Net remainder for platform = (Actual Money Collected - Center's Share)
        n_share = earned_syp - c_share
        
        detailed_batches.append({
            "batch": batch,
            "total_codes": batch.allocated_count,
            "sold_count": sold_cnt,
            "standard_price": std_price,
            "potential_value": pot_val,
            "actual_earned": earned_syp,
            "center_share": c_share,
            "net_share": n_share,
        })
        
        total_assigned += batch.allocated_count
        total_sold += sold_cnt
        total_potential += pot_val
        total_earned += earned_syp

        # Aggregate by course only
        if batch.course:
            key = f"course_{batch.course.id}"
            item_name = batch.course.title
            item_type = "مادة"
            instructor_name = batch.course.instructor.get_full_name() or batch.course.instructor.username if batch.course.instructor else ""
        else:
            continue
            
        if key not in course_summaries_map:
            course_summaries_map[key] = {
                "name": item_name,
                "type": item_type,
                "instructor": instructor_name,
                "standard_price": std_price,
                "sold_count": 0,
                "actual_earned": 0.0,
            }
        
        course_summaries_map[key]["sold_count"] += sold_cnt
        course_summaries_map[key]["actual_earned"] += earned_syp
        
    # Calculate standard total and discount value for each grouped item
    course_summaries = []
    for data in course_summaries_map.values():
        std_total = data["sold_count"] * data["standard_price"]
        discount_value = std_total - data["actual_earned"]
        data["standard_total"] = std_total
        data["discount_value"] = max(0.0, discount_value)
        course_summaries.append(data)
        
    # Aggregating overall summaries correctly from detail rows
    total_center_commission = sum(x["center_share"] for x in detailed_batches)
    total_net_share = total_earned - total_center_commission
    
    context = {
        "center": center,
        "commission_percent": commission_percent,
        "detailed_batches": detailed_batches,
        "course_summaries": course_summaries,
        "total_assigned": total_assigned,
        "total_sold": total_sold,
        "total_potential": total_potential,
        "total_earned": total_earned,
        "total_center_commission": total_center_commission,
        "total_net_share": total_net_share,
        "invoice_number": f"INV-{center.id}-{timezone.now().strftime('%Y%m%d')}",
    }
    return render(request, tmpl, context)


def _apply_premium_excel_styling(ws, column_width=22):
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    header_fill = PatternFill(start_color="8A6D3B", end_color="8A6D3B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    # 1. Basic Alignment and Borders for all populated cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    # 2. Elegant Styling for the Header Row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # 3. UI Optimizations: Freezing panes & Filter
    ws.freeze_panes = "A2"
    if ws.max_column > 0 and ws.max_row > 0:
        ws.auto_filter.ref = ws.dimensions

    # 4. Proper Sizing for Columns
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width


def _get_enhanced_filter_financials():
    from django.db.models import Sum
    from billing.models import AccessCode, AccessCodePrintLog
    from learning.models import Course
    
    rows = []
    sections = [
        {"label": "امتحانيات - فرع علمي", "kind": "exam_camp", "track": "scientific"},
        {"label": "امتحانيات - فرع أدبي", "kind": "exam_camp", "track": "literary"},
        {"label": "امتحانيات - تاسع", "kind": "exam_camp", "track": "ninth"},
        {"label": "امتحانيات - مواد مشتركة", "kind": "exam_camp", "track": "general"},
        {"label": "مكثفات - علمي", "kind": "intensive", "track": "scientific"},
        {"label": "مكثفات - أدبي", "kind": "intensive", "track": "literary"},
    ]
    
    for tab in sections:
        codes_qs = AccessCode.objects.filter(course__kind=tab["kind"], course__academic_track=tab["track"])
        courses_qs = Course.objects.filter(kind=tab["kind"], academic_track=tab["track"])
        
        total_codes = codes_qs.count()
        sold_codes_qs = codes_qs.filter(sale_status="sold")
        sold_count = sold_codes_qs.count()
        
        gross_earned = (sold_codes_qs.aggregate(total=Sum("sold_price_cents"))["total"] or 0) / 100
        
        # Calculate Forecast using standard price * total codes generated
        potential_forecast = 0
        for c in courses_qs:
            c_price = (c.price_cents or 0) / 100
            c_codes_cnt = AccessCode.objects.filter(course=c).count()
            potential_forecast += (c_price * c_codes_cnt)

        prints_cnt = AccessCodePrintLog.objects.filter(
            batch__course__kind=tab["kind"], 
            batch__course__academic_track=tab["track"]
        ).aggregate(total=Sum("cards_count"))["total"] or 0
        
        rows.append([
            tab["label"],
            courses_qs.count(),
            total_codes,
            sold_count,
            codes_qs.filter(redeemed_count__gt=0).count(),
            prints_cnt,
            int(gross_earned),
            int(potential_forecast)
        ])
    return rows


def _apply_god_tier_excel_styling(ws, column_width=25, force_zebra=True):
    """Super premium styling: RTL native, Heavy borders, Modern Font, High-End Header, Zebra Stripes."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 1. CRITICAL FOR ARABIC: Native RTL Interface
    ws.sheet_view.rightToLeft = True
    
    # 2. Color Palette for Executive Look
    brand_navy = PatternFill(start_color="233646", end_color="233646", fill_type="solid")
    zebra_light = PatternFill(start_color="F9FBFC", end_color="F9FBFC", fill_type="solid")
    
    header_font = Font(color="FFFFFF", bold=True, size=12, name="Segoe UI")
    data_font = Font(color="2C3E50", size=10, name="Segoe UI")
    
    heavy_edge = Side(style="thin", color="C0C9D0")
    uniform_border = Border(left=heavy_edge, right=heavy_edge, top=heavy_edge, bottom=heavy_edge)
    
    # 3. Loop to inject styles
    for r_idx, row in enumerate(ws.iter_rows(), start=1):
        # Set row height for readability
        ws.row_dimensions[r_idx].height = 28 if r_idx == 1 else 22
        
        for cell in row:
            cell.border = uniform_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            
            if r_idx == 1:
                cell.fill = brand_navy
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = data_font
                if force_zebra and r_idx % 2 == 0:
                    cell.fill = zebra_light
                    
    # 4. Formatting: Freeze & Filter
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
        
    # 5. Responsive Columns (Manual Override)
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = column_width


def _draw_erp_card(ws, row, col, title, val_formula, sub_label, hex_theme="233646"):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    head_f = PatternFill(start_color=hex_theme, end_color=hex_theme, fill_type="solid")
    body_f = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    foot_f = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin = Side(style="thin", color="D1D5DB")
    heavy_b = Side(style="medium", color=hex_theme)
    
    c1 = ws.cell(row=row, column=col)
    c1.value = title
    c1.font = Font(bold=True, color="FFFFFF", size=11, name="Segoe UI")
    c1.fill = head_f
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = Border(top=thin, left=thin, right=thin)

    c2 = ws.cell(row=row+1, column=col)
    c2.value = val_formula
    c2.font = Font(bold=True, color="1F2937", size=24, name="Segoe UI")
    c2.fill = body_f
    c2.alignment = Alignment(horizontal="center", vertical="center")
    c2.number_format = '#,##0'
    c2.border = Border(left=thin, right=thin)

    c3 = ws.cell(row=row+2, column=col)
    c3.value = sub_label
    c3.font = Font(italic=True, color="6B7280", size=9, name="Segoe UI")
    c3.fill = foot_f
    c3.alignment = Alignment(horizontal="center", vertical="center")
    c3.border = Border(left=thin, right=thin, bottom=heavy_b)

def _draw_erp_table(ws, sr, er, sc, ec, zebra=True):
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    head_f = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    head_font = Font(bold=True, color="FFFFFF", size=10, name="Segoe UI")
    row_even = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_b = Side(style="thin", color="E2E8F0")
    uni_border = Border(left=thin_b, right=thin_b, top=thin_b, bottom=thin_b)
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = uni_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r == sr:
                cell.fill = head_f
                cell.font = head_font
            else:
                cell.font = Font(size=9, name="Segoe UI", color="334155")
                if zebra and (r % 2 == 0):
                    cell.fill = row_even

@admin_required
def admin_financial_report_export(request):
    from openpyxl import Workbook
    from openpyxl.chart import PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
    from django.utils import timezone
    from django.db.models import Sum, Count, Avg, Q
    from openpyxl.utils import get_column_letter

    from billing.models import AccessCode, AccessCodePrintLog, CoursePackage, SalesCenter
    from learning.models import Course, Lesson, LessonProgress
    from accounts.models import StudentProfile, InstructorProfile

    wb = Workbook()
    act = wb.active
    wb.remove(act)

    # 1. ERP SHEET CONFIGURATION (ALL REQUIRED SHEETS)
    sheets = {}
    sheet_config = [
        ("01 - لوحة القيادة التنفيذية", True),
        ("02 - نظرة عامة مالية", True),
        ("03 - الإعدادات المالية", True),
        ("04 - المحرك المالي", True),
        ("05 - تقرير أرباح المدرسين", True),
        ("06 - تقرير أرباح المراكز", True),
        ("07 - التقرير المالي للطلاب", True),
        ("08 - ربحية الدورات", True),
        ("09 - تتبع الأقساط", True),
        ("10 - الديون المتأخرة", True),
        ("11 - تقرير الاستحقاقات", True),
        ("12 - الاعتراف بالإيرادات", True),
        ("13 - قائمة الدخل", True),
        ("14 - توقعات التدفق النقدي", True),
        ("15 - لوحة مؤشرات الأداء", True),
        ("16 - المخاطر والتنبيهات المالية", True),
        ("99 - قاعدة البيانات الخام", True),
    ]
    for name, rtl in sheet_config:
        ws = wb.create_sheet(name)
        if rtl: ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        sheets[name] = ws

    # ---------------------------------------------------------
    # 03 - FINANCIAL SETTINGS
    # ---------------------------------------------------------
    stg = sheets["03 - الإعدادات المالية"]
    stg.column_dimensions['B'].width = 35
    stg.column_dimensions['C'].width = 20
    
    stg.append(["", "⚙️ الإعدادات العامة (الافتراضية)", "القيمة الافتراضية"])
    stg.append(["", "نسبة عمولة المدرس الافتراضية", 0.40])      # C2
    stg.append(["", "نسبة حصة المركز الافتراضية", 0.15])      # C3
    stg.append(["", "نسبة حصة المنصة الافتراضية", 0.45])      # C4
    stg.append(["", "نسبة مصاريف التشغيل/التسويق", 0.05])     # C5
    stg.append(["", "نسبة الضريبة المقتطعة", 0.00])            # C6
    
    for r in range(2, 7):
        stg.cell(row=r, column=3).number_format = '0%'
    _draw_erp_table(stg, 1, 6, 2, 3)

    stg.append([])
    stg.append(["", "👨‍🏫 استثناءات نسب المدرسين", "النسبة المخصصة"])
    stg.append(["", "اسم المدرس", "نسبة الحصة"]) 
    t_row = 10
    for t in InstructorProfile.objects.select_related("user"):
        nm = t.user.get_full_name() or t.user.username
        stg.append(["", nm, ""])
        stg.cell(row=t_row, column=3).number_format = '0%'
        t_row += 1
    # Adding a couple blank placeholders for user addition
    stg.append(["", "مدرس إضافي (تعديل يدوي)", ""])
    stg.cell(row=t_row, column=3).number_format = '0%'
    t_row += 1
    _draw_erp_table(stg, 8, max(t_row-1, 10), 2, 3)
    
    c_start = t_row + 1
    stg.cell(row=c_start, column=2, value="🏢 استثناءات نسب المراكز")
    stg.cell(row=c_start+1, column=2, value="المركز")
    stg.cell(row=c_start+1, column=3, value="نسبة الحصة")
    c_row = c_start + 2
    for c in SalesCenter.objects.all():
        stg.cell(row=c_row, column=2, value=c.name)
        stg.cell(row=c_row, column=3, value="")
        stg.cell(row=c_row, column=3).number_format = '0%'
        c_row += 1
    _draw_erp_table(stg, c_start, max(c_row-1, c_start+1), 2, 3)

    cr_start = c_row + 1
    stg.cell(row=cr_start, column=2, value="📚 استثناءات المواد (أقوى أولوية)")
    stg.cell(row=cr_start+1, column=2, value="المادة")
    stg.cell(row=cr_start+1, column=3, value="حصة المدرس للمادة")
    cr_row = cr_start + 2
    for c in Course.objects.all():
        stg.cell(row=cr_row, column=2, value=c.title)
        stg.cell(row=cr_row, column=3, value="")
        stg.cell(row=cr_row, column=3).number_format = '0%'
        cr_row += 1
    _draw_erp_table(stg, cr_start, max(cr_row-1, cr_start+1), 2, 3)

    # ---------------------------------------------------------
    # 99 - RAW DATA
    # ---------------------------------------------------------
    raw = sheets["99 - قاعدة البيانات الخام"]
    for i in range(1, 9): raw.column_dimensions[get_column_letter(i)].width = 20
    raw.append(["رقم المعاملة", "التاريخ", "الطالب", "المادة", "المدرس", "المركز", "المبلغ المدفوع", "الحالة"])
    
    sales = AccessCode.objects.filter(sale_status="sold").select_related("course", "course__instructor", "sales_center").order_by("-sold_at")
    raw_r = 2
    for s in sales.iterator(chunk_size=1000):
        dt = s.sold_at.strftime("%Y-%m-%d") if s.sold_at else "2026-01-01"
        st_nm = s.assigned_student_name or "مجهول"
        crs_nm = s.course.title if s.course else "عام"
        tch_nm = s.course.instructor.get_full_name() if (s.course and s.course.instructor) else "-"
        cnt_nm = s.sales_center.name if s.sales_center else "منصة مباشر"
        amt = int((s.sold_price_cents or 0)/100)
        
        raw.append([s.id, dt, st_nm, crs_nm, tch_nm, cnt_nm, amt, "مكتمل"])
        raw_r += 1
    _draw_erp_table(raw, 1, max(2, raw_r-1), 1, 8)
    raw.freeze_panes = "A2"

    # ---------------------------------------------------------
    # 04 - FINANCIAL ENGINE
    # ---------------------------------------------------------
    eng = sheets["04 - المحرك المالي"]
    eng_cols = ["ID", "التاريخ", "الطالب", "المادة", "المدرس", "المركز", "إجمالي الإيراد", 
                "نسبة المدرس", "قيمة المدرس", "نسبة المركز", "قيمة المركز", 
                "نسبة التشغيل", "قيمة التشغيل", "صافي ربح المنصة"]
    eng.append(eng_cols)
    for i in range(1, 15): eng.column_dimensions[get_column_letter(i)].width = 16
    
    # Ranges in Settings Sheet
    c_rng = f"'03 - الإعدادات المالية'!B${cr_start+2}:C${max(cr_row-1, cr_start+2)}"
    t_rng = f"'03 - الإعدادات المالية'!B$10:C${max(t_row-1, 10)}"
    cnt_rng = f"'03 - الإعدادات المالية'!B${c_start+2}:C${max(c_row-1, c_start+2)}"
    
    def_t_pct = "'03 - الإعدادات المالية'!C$2"
    def_c_pct = "'03 - الإعدادات المالية'!C$3"
    def_m_pct = "'03 - الإعدادات المالية'!C$5" 
    
    for r in range(2, max(3, raw_r)):
        if r < raw_r:
            eng.append([
                f"='99 - قاعدة البيانات الخام'!A{r}", f"='99 - قاعدة البيانات الخام'!B{r}", f"='99 - قاعدة البيانات الخام'!C{r}",
                f"='99 - قاعدة البيانات الخام'!D{r}", f"='99 - قاعدة البيانات الخام'!E{r}", f"='99 - قاعدة البيانات الخام'!F{r}",
                f"='99 - قاعدة البيانات الخام'!G{r}"
            ])
        else:
            eng.append(["", "", "", "", "", "", 0])

        # Teacher %: IF(ISNUMBER(VLOOKUP(Course...)), VLOOKUP, IF(ISNUMBER(VLOOKUP(Teacher...)), VLOOKUP, Default))
        # Note: IFERROR(1/(1/VLOOKUP)) handles both #N/A and blank/0 as error, falling back beautifully.
        t_pct_f = f'=IFERROR(1/(1/VLOOKUP(D{r},{c_rng},2,FALSE)), IFERROR(1/(1/VLOOKUP(E{r},{t_rng},2,FALSE)), {def_t_pct}))'
        c_pct_f = f'=IF(F{r}="منصة مباشر", 0, IFERROR(1/(1/VLOOKUP(F{r},{cnt_rng},2,FALSE)), {def_c_pct}))'
        o_pct_f = f'={def_m_pct}'
        
        eng.cell(row=r, column=8, value=t_pct_f).number_format = '0%'
        eng.cell(row=r, column=9, value=f'=G{r}*H{r}') 
        eng.cell(row=r, column=10, value=c_pct_f).number_format = '0%'
        eng.cell(row=r, column=11, value=f'=G{r}*J{r}') 
        eng.cell(row=r, column=12, value=o_pct_f).number_format = '0%'
        eng.cell(row=r, column=13, value=f'=G{r}*L{r}') 
        eng.cell(row=r, column=14, value=f'=G{r}-I{r}-K{r}-M{r}') 
        
    _draw_erp_table(eng, 1, max(2, raw_r-1), 1, 14)
    eng.freeze_panes = "A2"

    # ---------------------------------------------------------
    # 05 - TEACHER EARNINGS REPORT
    # ---------------------------------------------------------
    tch = sheets["05 - تقرير أرباح المدرسين"]
    tch.append(["اسم المدرس", "عدد العمليات", "إجمالي الإيرادات", "مستحقات المدرس", "تم سداده (يدوي)", "الرصيد المتبقي (دين)"])
    for i in range(1, 7): tch.column_dimensions[get_column_letter(i)].width = 22
    t_out = 2
    for t in InstructorProfile.objects.select_related("user"):
        nm = t.user.get_full_name() or t.user.username
        tch.append([
            nm,
            f'=COUNTIF(\'04 - المحرك المالي\'!E:E, A{t_out})',
            f'=SUMIF(\'04 - المحرك المالي\'!E:E, A{t_out}, \'04 - المحرك المالي\'!G:G)',
            f'=SUMIF(\'04 - المحرك المالي\'!E:E, A{t_out}, \'04 - المحرك المالي\'!I:I)',
            0,
            f'=D{t_out}-E{t_out}'
        ])
        t_out += 1
    _draw_erp_table(tch, 1, max(2, t_out-1), 1, 6)

    # ---------------------------------------------------------
    # 06 - CENTER EARNINGS REPORT
    # ---------------------------------------------------------
    cnt = sheets["06 - تقرير أرباح المراكز"]
    cnt.append(["المركز", "إجمالي الإيرادات المولدة", "مستحقات المركز", "تم سداده (يدوي)", "الرصيد المتبقي"])
    for i in range(1, 6): cnt.column_dimensions[get_column_letter(i)].width = 22
    c_out = 2
    for c in SalesCenter.objects.all():
        cnt.append([
            c.name,
            f'=SUMIF(\'04 - المحرك المالي\'!F:F, A{c_out}, \'04 - المحرك المالي\'!G:G)',
            f'=SUMIF(\'04 - المحرك المالي\'!F:F, A{c_out}, \'04 - المحرك المالي\'!K:K)',
            0,
            f'=C{c_out}-D{c_out}'
        ])
        c_out += 1
    _draw_erp_table(cnt, 1, max(2, c_out-1), 1, 5)

    # ---------------------------------------------------------
    # 08 - COURSE PROFITABILITY
    # ---------------------------------------------------------
    cpf = sheets["08 - ربحية الدورات"]
    cpf.append(["المادة", "عدد المبيعات", "إجمالي الإيراد", "تكلفة المدرس", "تكلفة المراكز", "تكلفة تشغيلية", "صافي الربح", "هامش الربح %"])
    for i in range(1, 9): cpf.column_dimensions[get_column_letter(i)].width = 18
    cf_out = 2
    for c in Course.objects.all():
        cpf.append([
            c.title,
            f'=COUNTIF(\'04 - المحرك المالي\'!D:D, A{cf_out})',
            f'=SUMIF(\'04 - المحرك المالي\'!D:D, A{cf_out}, \'04 - المحرك المالي\'!G:G)',
            f'=SUMIF(\'04 - المحرك المالي\'!D:D, A{cf_out}, \'04 - المحرك المالي\'!I:I)',
            f'=SUMIF(\'04 - المحرك المالي\'!D:D, A{cf_out}, \'04 - المحرك المالي\'!K:K)',
            f'=SUMIF(\'04 - المحرك المالي\'!D:D, A{cf_out}, \'04 - المحرك المالي\'!M:M)',
            f'=SUMIF(\'04 - المحرك المالي\'!D:D, A{cf_out}, \'04 - المحرك المالي\'!N:N)',
            f'=IFERROR(G{cf_out}/C{cf_out}, 0)'
        ])
        cpf.cell(row=cf_out, column=8).number_format = '0%'
        cf_out += 1
    _draw_erp_table(cpf, 1, max(2, cf_out-1), 1, 8)
    cpf.conditional_formatting.add(f'H2:H{cf_out}', ColorScaleRule(start_type='min', start_color='FCA5A5', end_type='max', end_color='86EFAC'))

    # ---------------------------------------------------------
    # 13 - PROFIT & LOSS (P&L)
    # ---------------------------------------------------------
    pnl = sheets["13 - قائمة الدخل"]
    pnl.column_dimensions['B'].width = 35
    pnl.column_dimensions['C'].width = 25
    pnl.append(["", "قائمة الدخل الشاملة (P&L)", "القيمة"])
    pnl.append(["", "إجمالي الإيرادات (Gross Revenue)", "=SUM('04 - المحرك المالي'!G:G)"])
    pnl.append(["", "(-) الخصومات والمنح", 0])
    pnl.append(["", "صافي الإيرادات (Net Revenue)", "=C2-C3"])
    pnl.append(["", "(-) مستحقات المدرسين (COGS)", "=SUM('04 - المحرك المالي'!I:I)"])
    pnl.append(["", "(-) عمولات المراكز (COGS)", "=SUM('04 - المحرك المالي'!K:K)"])
    pnl.append(["", "إجمالي الربح (Gross Profit)", "=C4-C5-C6"])
    pnl.append(["", "(-) المصاريف التشغيلية والتسويق", "=SUM('04 - المحرك المالي'!M:M)"])
    pnl.append(["", "صافي الربح قبل الضريبة (EBT)", "=C7-C8"])
    pnl.append(["", "(-) الضريبة المقتطعة", "='03 - الإعدادات المالية'!C6 * C9"])
    pnl.append(["", "صافي الدخل النهائي (Net Income)", "=C9-C10"])
    _draw_erp_table(pnl, 1, 11, 2, 3)
    pnl['C11'].font = Font(bold=True, size=12, color="16A34A")

    # ---------------------------------------------------------
    # 01 - EXECUTIVE DASHBOARD
    # ---------------------------------------------------------
    dsh = sheets["01 - لوحة القيادة التنفيذية"]
    for char in ['B','C','D','E', 'F']: dsh.column_dimensions[char].width = 25
    dsh.merge_cells('B2:F2')
    dsh['B2'] = "محترفو التعليم | 🚀 نظام تخطيط الموارد المالية (ERP)"
    dsh['B2'].font = Font(bold=True, size=18, color="FFFFFF", name="Segoe UI")
    dsh['B2'].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    dsh['B2'].alignment = Alignment(horizontal="center", vertical="center")
    
    _draw_erp_card(dsh, 4, 2, "إجمالي الإيرادات", "=SUM('04 - المحرك المالي'!G:G)", "مبيعات الكلية", "059669")
    _draw_erp_card(dsh, 4, 3, "مستحقات المدرسين", "=SUM('04 - المحرك المالي'!I:I)", "التزامات مستحقة", "DC2626")
    _draw_erp_card(dsh, 4, 4, "عمولات المراكز", "=SUM('04 - المحرك المالي'!K:K)", "أرباح الفروع", "EA580C")
    _draw_erp_card(dsh, 4, 5, "مصاريف تشغيلية", "=SUM('04 - المحرك المالي'!M:M)", "استضافة وتسويق", "64748B")
    _draw_erp_card(dsh, 4, 6, "صافي ربح المنصة", "=SUM('04 - المحرك المالي'!N:N)", "الربح الفعلي الدقيق", "2563EB")
    
    dsh.cell(row=8, column=2, value="📈 مؤشرات الأداء الحيوية (KPIs)")
    dsh.cell(row=8, column=2).font = Font(bold=True, size=14, color="1E293B")
    
    dsh.cell(row=10, column=2, value="متوسط نسبة ربح المنصة")
    dsh.cell(row=10, column=3, value="=IFERROR(F5/B5, 0)")
    dsh.cell(row=10, column=3).number_format = '0%'
    
    dsh.cell(row=11, column=2, value="أفضل مادة مبيعاً")
    dsh.cell(row=11, column=3, value="=INDEX('08 - ربحية الدورات'!A2:A100, MATCH(MAX('08 - ربحية الدورات'!B2:B100), '08 - ربحية الدورات'!B2:B100, 0))")

    dsh.cell(row=12, column=2, value="أفضل مدرس مبيعاً")
    dsh.cell(row=12, column=3, value="=INDEX('05 - تقرير أرباح المدرسين'!A2:A100, MATCH(MAX('05 - تقرير أرباح المدرسين'!B2:B100), '05 - تقرير أرباح المدرسين'!B2:B100, 0))")
    _draw_erp_table(dsh, 10, 12, 2, 3)

    ts = timezone.now().strftime("%Y-%m-%d_%H%M")
    fn = f"PLATFORM_ERP_FINANCIAL_ENGINE_{ts}.xlsx"
    return _workbook_response(wb, fn)



@admin_required
def admin_packages_report_export(request):
    from openpyxl import Workbook
    from django.utils import timezone
    from django.db.models import Sum
    from billing.models import AccessCode, CoursePackage
    
    workbook = Workbook()
    summary = workbook.active
    summary.title = "packages"
    summary.append(["package", "track", "subjects", "courses", "codes", "sold", "activated", "inactive", "gross_syp"])
    
    packages = CoursePackage.objects.prefetch_related("courses").order_by("name")
    for package in packages:
        codes_qs = AccessCode.objects.filter(access_type="package", package=package)
        eligible_courses = package.eligible_courses_queryset()
        
        summary.append([
            package.name,
            package.get_package_track_display(),
            eligible_courses.values("subject").distinct().count(),
            eligible_courses.count(),
            codes_qs.count(),
            codes_qs.filter(sale_status="sold").count(),
            codes_qs.filter(redeemed_count__gt=0).count(),
            codes_qs.filter(redeemed_count=0).count(),
            int((codes_qs.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) / 100),
        ])

    codes_sheet = workbook.create_sheet("codes")
    codes_sheet.append(["code", "package", "sale_status", "student_name", "student_phone", "sold_by", "sold_at", "sold_price_syp", "activated", "batch", "sales_center"])
    codes = (
        AccessCode.objects.filter(access_type="package")
        .select_related("package", "sold_by", "batch", "sales_center")
        .order_by("package__name", "-created_at")
    )
    for code in codes:
        codes_sheet.append([
            code.code,
            code.package.name if code.package else "",
            code.sale_status,
            code.assigned_student_name,
            code.assigned_student_phone,
            code.sold_by.username if code.sold_by else "",
            code.sold_at.strftime("%Y-%m-%d %H:%M") if code.sold_at else "",
            (code.sold_price_cents or 0) // 100,
            "yes" if code.redeemed_count else "no",
            code.batch.name if code.batch else "",
            code.sales_center.name if code.sales_center else "",
        ])
    for sheet in workbook.worksheets:
        for column in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            sheet.column_dimensions[column].width = 22
    return _workbook_response(workbook, f"packages-report-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx")


@admin_required
def admin_course_report_export(request, course_id):
    course = get_object_or_404(Course.objects.select_related("subject", "instructor", "instructor__instructor_profile"), id=course_id)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "summary"
    codes = AccessCode.objects.filter(course=course)
    prints = AccessCodePrintLog.objects.filter(batch__course=course)
    summary.append(["course", course.title])
    summary.append(["codes", codes.count()])
    summary.append(["sold", codes.filter(sale_status="sold").count()])
    summary.append(["activated", codes.filter(redeemed_count__gt=0).count()])
    summary.append(["inactive", codes.filter(redeemed_count=0).count()])
    summary.append(["printed_cards", prints.aggregate(total=Sum("cards_count"))["total"] or 0])
    summary.append(["gross_syp", (codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) // 100])

    codes_sheet = workbook.create_sheet("codes")
    codes_sheet.append(["code", "sale_status", "student_name", "student_phone", "sold_by", "sold_at", "activated", "sales_center", "batch"])
    for code in codes.select_related("sold_by", "sales_center", "batch").order_by("-created_at"):
        codes_sheet.append([
            code.code,
            code.sale_status,
            code.assigned_student_name,
            code.assigned_student_phone,
            code.sold_by.username if code.sold_by else "",
            code.sold_at.strftime("%Y-%m-%d %H:%M") if code.sold_at else "",
            "yes" if code.redeemed_count else "no",
            code.sales_center.name if code.sales_center else "",
            code.batch.name if code.batch else "",
        ])

    activations = workbook.create_sheet("activations")
    activations.append(["student", "username", "code", "device", "created_at"])
    for grant in AccessGrant.objects.filter(course=course).select_related("user", "access_code").order_by("-created_at"):
        activations.append([
            grant.user.get_full_name() or grant.user.username,
            grant.user.username,
            grant.access_code.code if grant.access_code else "",
            grant.device_fingerprint,
            grant.created_at.strftime("%Y-%m-%d %H:%M"),
        ])
    return _workbook_response(workbook, f"course-report-{course.id}-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx")


@admin_required
def admin_instructor_report(request, instructor_id):
    if request.GET.get('format') == 'print':
        render_tmpl = "dashboard/admin_instructor_report_print.html"
    else:
        render_tmpl = "dashboard/admin_instructor_report.html"
    from django.db.models import Sum, Count
    from billing.models import AccessCode
    from accounts.models import InstructorProfile
    from apps.instructor_finance.models import RevenueShareAgreement
    
    instructor = get_object_or_404(User, id=instructor_id)
    profile, created = InstructorProfile.objects.get_or_create(user=instructor)
    saved_card_data = profile.report_card_data or {}
    
    # Get active agreement or default (None course) commission percentage
    agreement = RevenueShareAgreement.objects.filter(instructor=instructor, course=None, is_active=True).first()
    if not agreement:
        agreement = RevenueShareAgreement.objects.filter(instructor=instructor, is_active=True).first()
    default_commission_pct = (agreement.commission_bps / 100.0) if agreement else 50.0

    courses = Course.objects.filter(instructor=instructor).select_related("subject")
    
    total_codes = 0
    total_sold = 0
    total_activated = 0
    total_gross = 0
    
    course_data = []
    for course in courses:
        codes = AccessCode.objects.filter(course=course)
        cnt_all = codes.count()
        cnt_sold = codes.filter(sale_status="sold").count()
        cnt_active = codes.filter(redeemed_count__gt=0).count()
        sum_gross = (codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) // 100
        
        course_centers = (
            AccessCode.objects.filter(course=course, sale_status="sold")
            .select_related("sales_center")
            .values("sales_center__name")
            .annotate(sold_count=Count("id"), total_sales=Sum("sold_price_cents"))
            .order_by("-sold_count")
        )
        
        detailed_codes = codes.select_related("sold_by", "sales_center").prefetch_related("grants", "grants__user", "grants__user__student_profile").order_by("-sold_at")
        
        course_data.append({
            "course": course,
            "cnt_all": cnt_all,
            "cnt_sold": cnt_sold,
            "cnt_active": cnt_active,
            "cnt_inactive": cnt_all - cnt_active,
            "sum_gross": sum_gross,
            "centers": course_centers,
            "codes": detailed_codes,
        })
        
        total_codes += cnt_all
        total_sold += cnt_sold
        total_activated += cnt_active
        total_gross += sum_gross

    centers_sales = (
        AccessCode.objects.filter(course__instructor=instructor, sale_status="sold")
        .select_related("sales_center", "sales_center__institute")
        .values("sales_center__name", "sales_center__institute__name")
        .annotate(sold_count=Count("id"), total_sales=Sum("sold_price_cents"))
        .order_by("-sold_count")
    )
    
    from billing.models import AccessGrant
    governorate_sales = (
        AccessGrant.objects.filter(course__instructor=instructor, source="code")
        .select_related("user__student_profile")
        .values("user__student_profile__governorate")
        .annotate(activated_count=Count("id"))
        .order_by("-activated_count")
    )

    discounts_sales = (
        AccessCode.objects.filter(course__instructor=instructor, sale_status="sold")
        .exclude(price_reason="")
        .exclude(price_reason="سعر كامل")
        .values("price_reason")
        .annotate(sold_count=Count("id"), total_sales=Sum("sold_price_cents"))
        .order_by("-sold_count")
    )

    if request.GET.get('format') == 'print-card':
        card_name = request.GET.get('name') or saved_card_data.get('custom_name') or instructor.get_full_name() or instructor.username
        card_specialty = request.GET.get('specialty') or saved_card_data.get('custom_specialty') or profile.specialty or "مدرس مادة الرياضيات"
        
        try:
            comm_pct = float(request.GET.get('comm_pct'))
        except (ValueError, TypeError):
            comm_pct = saved_card_data.get('commission_pct') or default_commission_pct
            
        try:
            sold_cnt = int(request.GET.get('sold_cnt'))
        except (ValueError, TypeError):
            sold_cnt = saved_card_data.get('custom_total_codes') or total_sold
            
        try:
            gross_val = float(request.GET.get('gross_val'))
        except (ValueError, TypeError):
            gross_val = saved_card_data.get('custom_total_gross') or total_gross
            
        notes = request.GET.get('notes') or saved_card_data.get('custom_notes') or "حساب مستحقات مبيعات الأكواد"
        creator = request.GET.get('creator') or saved_card_data.get('custom_creator') or "Manager thaaer almasre"
        
        net_share = round(gross_val * (comm_pct / 100))
        
        context = {
            "instructor": instructor,
            "profile": profile,
            "card_name": card_name,
            "card_specialty": card_specialty,
            "comm_pct": comm_pct,
            "sold_cnt": sold_cnt,
            "gross_val": gross_val,
            "notes": notes,
            "creator": creator,
            "net_share": net_share,
        }
        return render(request, "dashboard/admin_instructor_card_print.html", context)

    return render(
        request,
        render_tmpl,
        {
            "instructor": instructor,
            "profile": profile,
            "course_data": course_data,
            "total_codes": total_codes,
            "total_sold": total_sold,
            "total_activated": total_activated,
            "total_inactive": total_codes - total_activated,
            "total_gross": total_gross,
            "centers_sales": centers_sales,
            "governorate_sales": governorate_sales,
            "discounts_sales": discounts_sales,
            "saved_card_data": saved_card_data,
            "default_commission_pct": default_commission_pct,
        }
    )


@admin_required
def admin_instructor_report_export(request, instructor_id):
    from django.db.models import Sum, Count
    from billing.models import AccessCode
    instructor = get_object_or_404(User, id=instructor_id)
    courses = Course.objects.filter(instructor=instructor).select_related("subject")
    
    workbook = Workbook()
    
    # 1. Summary Sheet ("الملخص")
    summary_sheet = workbook.active
    summary_sheet.title = "الملخص الشامل"
    
    summary_sheet.append([f"التقرير المالي وتفصيل مبيعات المدرس: {instructor.get_full_name() or instructor.username}"])
    summary_sheet.append([]) # Empty row
    summary_sheet.append(["الدورة", "المادة", "المسار الأكاديمي", "إجمالي الأكواد", "المباعة", "المفعلة", "غير المفعلة", "إجمالي المبيعات (ل.س)"])
    
    total_all_codes = 0
    total_all_sold = 0
    total_all_activated = 0
    total_all_inactive = 0
    total_all_gross = 0
    
    for course in courses:
        codes = AccessCode.objects.filter(course=course)
        cnt_all = codes.count()
        cnt_sold = codes.filter(sale_status="sold").count()
        cnt_active = codes.filter(redeemed_count__gt=0).count()
        cnt_inactive = cnt_all - cnt_active
        sum_gross = (codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) // 100
        
        summary_sheet.append([
            course.title,
            course.subject.name if course.subject else "",
            f"{course.get_kind_display()} - {course.get_academic_track_display()}",
            cnt_all,
            cnt_sold,
            cnt_active,
            cnt_inactive,
            sum_gross
        ])
        
        total_all_codes += cnt_all
        total_all_sold += cnt_sold
        total_all_activated += cnt_active
        total_all_inactive += cnt_inactive
        total_all_gross += sum_gross
        
    summary_sheet.append([
        "إجمالي المدرس",
        "",
        "",
        total_all_codes,
        total_all_sold,
        total_all_activated,
        total_all_inactive,
        total_all_gross
    ])
    
    # 2. Sales Centers Sheet ("مراكز البيع")
    centers_sheet = workbook.create_sheet("مراكز البيع")
    centers_sheet.append(["مركز البيع", "المعهد", "عدد الأكواد المباعة", "إجمالي المبيعات (ل.س)"])
    
    centers_sales = (
        AccessCode.objects.filter(course__instructor=instructor, sale_status="sold")
        .select_related("sales_center", "sales_center__institute")
        .values("sales_center__name", "sales_center__institute__name")
        .annotate(sold_count=Count("id"), total_sales=Sum("sold_price_cents"))
        .order_by("-sold_count")
    )
    
    total_centers_sold = 0
    total_centers_gross = 0
    
    for row in centers_sales:
        center_name = row["sales_center__name"] or "مبيعات المنصة / مباشرة"
        institute_name = row["sales_center__institute__name"] or "-"
        sold_count = row["sold_count"]
        sales_gross = (row["total_sales"] or 0) // 100
        
        centers_sheet.append([
            center_name,
            institute_name,
            sold_count,
            sales_gross
        ])
        total_centers_sold += sold_count
        total_centers_gross += sales_gross
        
    centers_sheet.append([
        "إجمالي مبيعات المراكز",
        "",
        total_centers_sold,
        total_centers_gross
    ])
    
    # 3. Course sheets ("تفصيل الدورات")
    for course in courses:
        sheet_title = course.title
        for char in "[]*?:/\\":
            sheet_title = sheet_title.replace(char, "")
        sheet_title = sheet_title[:30]
        
        cs = workbook.create_sheet(sheet_title)
        
        # Course Header
        cs.append([f"تفاصيل دورة: {course.title}"])
        cs.append([])
        
        codes = AccessCode.objects.filter(course=course).select_related("sold_by", "sales_center", "batch")
        cnt_all = codes.count()
        cnt_sold = codes.filter(sale_status="sold").count()
        cnt_active = codes.filter(redeemed_count__gt=0).count()
        cnt_inactive = cnt_all - cnt_active
        sum_gross = (codes.filter(sale_status="sold").aggregate(total=Sum("sold_price_cents"))["total"] or 0) // 100
        
        cs.append(["مؤشر الدورة", "القيمة"])
        cs.append(["المادة", course.subject.name if course.subject else ""])
        cs.append(["النوع / الفرع", f"{course.get_kind_display()} - {course.get_academic_track_display()}"])
        cs.append(["حالة الدورة", course.get_status_display()])
        cs.append(["إجمالي الأكواد", cnt_all])
        cs.append(["المباعة", cnt_sold])
        cs.append(["المفعلة", cnt_active])
        cs.append(["إجمالي المبيعات (ل.س)", sum_gross])
        
        cs.append([])
        cs.append([])
        
        cs.append(["مبيعات مراكز البيع لهذه الدورة"])
        cs.append(["مركز البيع", "عدد الأكواد المباعة", "إجمالي المبيعات (ل.س)"])
        
        course_centers = (
            AccessCode.objects.filter(course=course, sale_status="sold")
            .select_related("sales_center")
            .values("sales_center__name")
            .annotate(sold_count=Count("id"), total_sales=Sum("sold_price_cents"))
            .order_by("-sold_count")
        )
        
        for cc in course_centers:
            cs.append([
                cc["sales_center__name"] or "مبيعات المنصة / مباشرة",
                cc["sold_count"],
                (cc["total_sales"] or 0) // 100
            ])
            
        cs.append([])
        cs.append([])
        
        cs.append(["تفاصيل الأكواد والطلاب"])
        cs.append(["الكود", "حالة البيع", "اسم الطالب", "هاتف الطالب", "المباع بواسطة", "تاريخ البيع", "السعر (ل.س)", "مفعل", "مركز البيع"])
        
        for code in codes.order_by("-sold_at"):
            cs.append([
                code.code,
                code.get_sale_status_display(),
                code.assigned_student_name,
                code.assigned_student_phone,
                code.sold_by.username if code.sold_by else "",
                code.sold_at.strftime("%Y-%m-%d %H:%M") if code.sold_at else "",
                (code.sold_price_cents or 0) // 100,
                "نعم" if code.redeemed_count > 0 else "لا",
                code.sales_center.name if code.sales_center else ""
            ])
            
    return _workbook_response(workbook, f"instructor-report-{instructor.id}-{timezone.now().strftime('%Y%m%d-%H%M')}.xlsx")


from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def _style_workbook(workbook):
    # Premium corporate colors matching the platform
    header_font = Font(name="Cairo", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid") # Dark Slate Gray
    
    data_font = Font(name="Cairo", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    align_right = Alignment(horizontal='right', vertical='center')
    align_center = Alignment(horizontal='center', vertical='center')

    for sheet in workbook.worksheets:
        # Set RTL layout for Arabic
        sheet.sheet_view.rightToLeft = True
        
        # Ensure gridlines are visible when printing
        if sheet.views.sheetView:
            sheet.views.sheetView[0].showGridLines = True
            
        # Setup page layout optimized for A4 print
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.orientation = sheet.ORIENTATION_PORTRAIT
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        
        # Format rows and cells
        for r_idx, row in enumerate(sheet.iter_rows(), start=1):
            is_header = r_idx == 1
            sheet.row_dimensions[r_idx].height = 28 if is_header else 22
            
            for cell in row:
                cell.border = thin_border
                if is_header:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                else:
                    cell.font = data_font
                    val = str(cell.value or '')
                    # Align numbers and short codes in the center, and text/names to the right
                    if val.isdigit() or len(val) <= 12 or cell.value is None:
                        cell.alignment = align_center
                    else:
                        cell.alignment = align_right
                        
        # Auto-adjust column widths to prevent truncation
        for col in sheet.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or '')
                val_len = len(val.encode('utf-8')) // 2 if any(ord(c) > 127 for c in val) else len(val)
                if val_len > max_len:
                    max_len = val_len
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _workbook_response(workbook, filename):
    # Style workbook with Cairo, RTL and A4 printing support before saving
    try:
        _style_workbook(workbook)
    except Exception:
        pass
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@admin_required
def admin_packages(request):
    package_form = CoursePackageForm(prefix="package")
    code_form = PackageCodeGenerateForm(prefix="codes")
    batch_form = PackageCodeBatchForm(prefix="batch")
    sale_form = PackageCodeSaleForm(prefix="sale")

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "create_package":
            package_form = CoursePackageForm(request.POST, prefix="package")
            if package_form.is_valid():
                package = package_form.save()
                messages.success(request, f"تم إنشاء الباقة {package.name}.")
                return redirect("dashboard:admin_packages")
        elif action == "generate_package_codes":
            code_form = PackageCodeGenerateForm(request.POST, prefix="codes")
            if code_form.is_valid():
                package = code_form.cleaned_data["package"]
                quantity = code_form.cleaned_data["quantity"]
                prefix = code_form.cleaned_data["code_prefix"] or package.code or "PKG"
                created = 0
                for _index in range(quantity):
                    shim = type("CodePrefix", (), {"code_prefix": prefix})()
                    AccessCode.objects.create(
                        code=unique_code(shim),
                        access_type="package",
                        package=package,
                        sale_status="available",
                        max_redemptions=1,
                        valid_until=timezone.now() + timezone.timedelta(days=180),
                        notes=f"كود باقة: {package.name}",
                    )
                    created += 1
                messages.success(request, f"تم توليد {created} كود للباقة {package.name}.")
                return redirect("dashboard:admin_packages")
        elif action == "create_package_batch":
            batch_form = PackageCodeBatchForm(request.POST, prefix="batch")
            if batch_form.is_valid():
                package = request.POST.get("package_id")
                package_obj = get_object_or_404(CoursePackage, id=package)
                batch = batch_form.save(commit=False)
                batch.package = package_obj
                batch.allocated_count = batch_form.cleaned_data["quantity"]
                batch.free_count = batch_form.cleaned_data["quantity"] if batch_form.cleaned_data["free_codes"] else 0
                batch.save()
                created = create_codes_for_batch(
                    batch,
                    batch_form.cleaned_data["quantity"],
                    free_codes=batch_form.cleaned_data["free_codes"],
                )
                messages.success(request, f"تم إنشاء دفعة {batch.name} وتوليد {created} كود للباقة {package_obj.name}.")
                return redirect("dashboard:admin_packages")
        elif action == "sell_package_code":
            sale_form = PackageCodeSaleForm(request.POST, prefix="sale")
            if sale_form.is_valid():
                with transaction.atomic():
                    access_code = AccessCode.objects.select_for_update().get(
                        id=sale_form.cleaned_data["code"].id,
                        access_type="package",
                    )
                    student_phone = sanitize_plain_text(sale_form.cleaned_data["student_phone"])
                    student_name = sanitize_plain_text(sale_form.cleaned_data["student_name"])
                    student, created = User.objects.get_or_create(
                        username=student_phone,
                        defaults={"first_name": student_name, "is_active": True},
                    )
                    if created:
                        student.set_unusable_password()
                        student.save(update_fields=["password"])
                    elif student_name and not student.get_full_name():
                        student.first_name = student_name
                        student.save(update_fields=["first_name"])
                    StudentProfile.objects.get_or_create(
                        user=student,
                        defaults={"phone": student_phone, "track": access_code.package.get_package_track_display() if access_code.package else ""},
                    )
                    access_code.assigned_student_name = student_name
                    access_code.assigned_student_phone = student_phone
                    access_code.sale_status = "sold"
                    access_code.sold_by = request.user
                    access_code.sold_at = timezone.now()
                    price_amount = sale_form.cleaned_data.get("price_amount")
                    if price_amount is not None:
                        access_code.sold_price_cents = int(price_amount * 100)
                        access_code.price_reason = "تحديد يدوي من الإدارة"
                    else:
                        base_price = access_code.package.price_cents if access_code.package else 0
                        best_rule, max_discount = _best_active_discount_for_price(base_price, package=access_code.package)
                        if best_rule and base_price > 0:
                            access_code.sold_price_cents = base_price - max_discount
                            if best_rule.discount_percent > 0:
                                access_code.price_reason = f"حسم {best_rule.discount_percent}% بمناسبة {best_rule.name}"
                            else:
                                access_code.price_reason = f"حسم بقيمة {best_rule.discount_amount_syp} ل.س بمناسبة {best_rule.name}"
                        else:
                            access_code.sold_price_cents = base_price
                            access_code.price_reason = "سعر كامل"
                    access_code.save(update_fields=[
                        "assigned_student_name",
                        "assigned_student_phone",
                        "sale_status",
                        "sold_by",
                        "sold_at",
                        "sold_price_cents",
                        "price_reason",
                        "updated_at",
                    ])
                messages.success(request, f"تم بيع كود الباقة {access_code.code} للطالب {student_name}.")
                return redirect("dashboard:admin_packages")

    packages = (
        CoursePackage.objects.prefetch_related("courses")
        .annotate(
            courses_total=Count("courses", distinct=True),
            codes_total=Count("access_codes", distinct=True),
            sold_total=Count("access_codes", filter=models.Q(access_codes__sale_status="sold"), distinct=True),
            activated_total=Count("access_codes", filter=models.Q(access_codes__redeemed_count__gt=0), distinct=True),
            gross_sales=Sum("access_codes__sold_price_cents", filter=models.Q(access_codes__sale_status="sold")),
        )
        .order_by("name")
    )
    for package in packages:
        eligible_courses = package.eligible_courses_queryset()
        package.eligible_courses_count = eligible_courses.count()
        package.eligible_subjects_count = eligible_courses.values("subject").distinct().count()
    package_codes = AccessCode.objects.filter(access_type="package").select_related("package", "sold_by").order_by("-created_at")[:80]
    package_batches = AccessCodeBatch.objects.filter(package__isnull=False).select_related("package", "institute", "sales_center").order_by("-created_at")[:80]
    return render(
        request,
        "dashboard/admin_packages.html",
        {
            "package_form": package_form,
            "code_form": code_form,
            "batch_form": batch_form,
            "sale_form": sale_form,
            "packages": packages,
            "package_codes": package_codes,
            "package_batches": package_batches,
        },
    )


@admin_required
def admin_content_hub(request):
    from learning.models import Subject, Course
    from .forms import SubjectForm
    from django.utils.text import slugify

    subject_form = SubjectForm()
    
    if request.method == "POST" and request.POST.get("action") == "create_subject":
        subject_form = SubjectForm(request.POST)
        if subject_form.is_valid():
            subject = subject_form.save(commit=False)
            if not subject.slug:
                base_slug = slugify(subject.name, allow_unicode=True) or "subject"
                slug = base_slug
                counter = 2
                while Subject.objects.filter(slug=slug).exists():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                subject.slug = slug
            subject.save()
            messages.success(request, f"تم إضافة المادة {subject.name} بنجاح.")
            return redirect("dashboard:admin_content_hub")

    subjects = Subject.objects.annotate(courses_count=Count("courses")).order_by("name")
    recent_courses = Course.objects.select_related("subject", "instructor", "instructor__instructor_profile").order_by("-created_at")
    
    return render(
        request,
        "dashboard/admin_content_hub.html",
        {
            "subjects": subjects,
            "recent_courses": recent_courses,
            "subject_form": subject_form,
        },
    )


@admin_required
def admin_exams_hub(request):
    from exams.models import Exam, Question, Attempt
    exams = Exam.objects.select_related("course").annotate(questions_count=Count("questions")).order_by("-created_at")
    recent_attempts = Attempt.objects.select_related("user", "exam").order_by("-created_at")[:50]
    
    return render(
        request,
        "dashboard/admin_exams_hub.html",
        {
            "exams": exams,
            "recent_attempts": recent_attempts,
            "total_questions": Question.objects.count(),
        },
    )


@admin_required
def admin_departments(request):
    from accounts.models import AcademicBranch
    from .forms import AcademicBranchForm
    
    if request.method == "POST":
        form = AcademicBranchForm(request.POST)
        if form.is_valid():
            branch = form.save()
            messages.success(request, f"تمت إضافة قسم {branch.name} بنجاح.")
            return redirect("dashboard:admin_departments")
            
    departments = AcademicBranch.objects.all().order_by("sort_order", "name")
    return render(request, "dashboard/admin_departments.html", {
        "departments": departments,
        "department_form": AcademicBranchForm()
    })


@admin_required
def admin_department_edit(request, department_id):
    from accounts.models import AcademicBranch
    from .forms import AcademicBranchForm
    branch = get_object_or_404(AcademicBranch, id=department_id)
    
    if request.method == "POST":
        form = AcademicBranchForm(request.POST, instance=branch)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تحديث قسم {branch.name} بنجاح.")
        else:
            messages.error(request, "حدث خطأ أثناء التحديث.")
            
    return redirect("dashboard:admin_departments")


@admin_required
def admin_subjects(request):
    from learning.models import Subject
    from .forms import SubjectForm
    from django.utils.text import slugify

    if request.method == "POST":
        form = SubjectForm(request.POST)
        if form.is_valid():
            subject = form.save(commit=False)
            if not subject.slug:
                subject.slug = slugify(subject.name, allow_unicode=True) or f"sub-{Subject.objects.count() + 1}"
            subject.save()
            messages.success(request, f"تمت إضافة مادة {subject.name} بنجاح.")
            return redirect("dashboard:admin_subjects")

    subjects = Subject.objects.annotate(
        courses_count=Count("courses", distinct=True),
        units_count=Count("courses__units", distinct=True),
        lessons_count=Count("courses__units__lessons", distinct=True)
    ).order_by("name")
    
    return render(request, "dashboard/admin_subjects.html", {
        "subjects": subjects,
        "subject_form": SubjectForm()
    })


@admin_required
def admin_subject_edit(request, subject_id):
    from learning.models import Subject
    from .forms import SubjectForm
    subject = get_object_or_404(Subject, id=subject_id)
    
    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تحديث مادة {subject.name} بنجاح.")
        else:
            messages.error(request, "حدث خطأ أثناء التحديث.")
            
    return redirect("dashboard:admin_subjects")


@admin_required
def admin_units(request):
    from learning.models import Unit, Course
    from .forms import UnitForm
    course_id = request.GET.get("course_id")
    
    if request.method == "POST":
        form = UnitForm(request.POST)
        if form.is_valid():
            unit = form.save()
            messages.success(request, f"تمت إضافة وحدة {unit.title} بنجاح.")
            next_url = request.POST.get("next")
            if next_url and next_url.startswith("/admin-dashboard/content/units/"):
                return redirect(next_url)
            return redirect("dashboard:admin_units")

    units = Unit.objects.select_related("course").annotate(
        lessons_count=Count("lessons")
    ).order_by("course__title", "sort_order")
    
    if course_id:
        units = units.filter(course_id=course_id)
        
    courses = Course.objects.only("id", "title").order_by("title")
    
    return render(request, "dashboard/admin_units.html", {
        "units": units,
        "courses": courses,
        "selected_course_id": course_id,
        "unit_form": UnitForm()
    })


@admin_required
def admin_unit_edit(request, unit_id):
    from learning.models import Unit
    from .forms import UnitForm
    unit = get_object_or_404(Unit, id=unit_id)
    
    if request.method == "POST":
        form = UnitForm(request.POST, request.FILES, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, f"تم تحديث وحدة {unit.title} بنجاح.")
        else:
            messages.error(request, "حدث خطأ أثناء تحديث الوحدة.")
    next_url = request.POST.get("next")
    if next_url and next_url.startswith("/admin-dashboard/content/units/"):
        return redirect(next_url)
    return redirect("dashboard:admin_units")


@admin_required
def admin_lessons(request):
    from learning.models import Lesson, Course, Unit
    course_id = request.GET.get("course_id")
    unit_id = request.GET.get("unit_id")
    
    if request.method == "POST":
        unit = get_object_or_404(Unit.objects.select_related("course"), id=request.POST.get("unit"))
        lesson = Lesson.objects.create(
            unit=unit,
            title=request.POST.get("title", "").strip(),
            description=request.POST.get("description", "").strip(),
            lesson_type=request.POST.get("lesson_type", "video"),
            video_url=request.POST.get("video_url", ""),
            duration_seconds=request.POST.get("duration_seconds") or None,
            sort_order=request.POST.get("sort_order") or 0,
            is_free_preview=request.POST.get("is_free_preview") == "on",
        )
        if request.FILES.get("video_file"):
            lesson.video_file = request.FILES["video_file"]
        if request.FILES.get("pdf_file"):
            lesson.pdf_file = request.FILES["pdf_file"]
        if lesson.video_file or lesson.pdf_file:
            lesson.save(update_fields=["video_file", "pdf_file"])
        messages.success(request, f"تمت إضافة درس {lesson.title} بنجاح.")
        next_url = request.POST.get("next")
        if next_url and next_url.startswith("/admin-dashboard/content/lessons/"):
            return redirect(next_url)
        return redirect("dashboard:admin_lessons")

    lessons = Lesson.objects.select_related("unit", "unit__course").order_by("unit__course__title", "unit__sort_order", "sort_order")
    
    if unit_id:
        lessons = lessons.filter(unit_id=unit_id)
    elif course_id:
        lessons = lessons.filter(unit__course_id=course_id)
        
    courses = Course.objects.only("id", "title").order_by("title")
    units = Unit.objects.filter(course_id=course_id) if course_id else Unit.objects.none()
    
    return render(request, "dashboard/admin_lessons.html", {
        "lessons": lessons,
        "courses": courses,
        "units": units,
        "all_units": Unit.objects.select_related("course").all().order_by("course__title", "sort_order"),
        "selected_course_id": course_id,
        "selected_unit_id": unit_id
    })


@admin_required
def admin_lesson_edit(request, lesson_id):
    from learning.models import Lesson, Unit
    lesson = get_object_or_404(Lesson, id=lesson_id)
    
    if request.method == "POST":
        if request.POST.get("unit"):
            lesson.unit = get_object_or_404(Unit, id=request.POST.get("unit"))
        lesson.title = request.POST.get("title", "").strip()
        lesson.description = request.POST.get("description", "").strip()
        lesson.lesson_type = request.POST.get("lesson_type", lesson.lesson_type)
        lesson.video_url = request.POST.get("video_url", "")
        lesson.duration_seconds = request.POST.get("duration_seconds") or None
        lesson.sort_order = request.POST.get("sort_order") or 0
        lesson.is_free_preview = request.POST.get("is_free_preview") == "on"
        if request.FILES.get("video_file"):
            lesson.video_file = request.FILES["video_file"]
        if request.FILES.get("pdf_file"):
            lesson.pdf_file = request.FILES["pdf_file"]
        lesson.save()
        messages.success(request, f"تم تحديث درس {lesson.title} بنجاح.")
    next_url = request.POST.get("next")
    if next_url and next_url.startswith("/admin-dashboard/content/lessons/"):
        return redirect(next_url)
    return redirect("dashboard:admin_lessons")


@admin_required
def admin_student_detail(request, user_id):
    student = get_object_or_404(User.objects.select_related("student_profile"), id=user_id)
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "toggle_status":
            student.is_active = not student.is_active
            student.save()
            status = "نشط" if student.is_active else "موقوف"
            messages.success(request, f"تم تغيير حالة الطالب إلى {status} بنجاح.")
            
        elif action == "reset_devices":
            student.devices.all().delete()
            messages.success(request, "تم تصفير جميع أجهزة الطالب بنجاح.")
            
        elif action == "delete_device":
            device_id = request.POST.get("device_id")
            UserDevice.objects.filter(id=device_id, user=student).delete()
            messages.success(request, "تم إزالة الجهاز المحدد بنجاح.")
            
        elif action == "toggle_device_active":
            device_id = request.POST.get("device_id")
            device = get_object_or_404(UserDevice, id=device_id, user=student)
            device.is_active = not device.is_active
            device.save()
            status = "نشط" if device.is_active else "موقوف (محظور)"
            messages.success(request, f"تم تغيير حالة الجهاز إلى {status} بنجاح.")
            
        elif action == "grant_course":
            course_id = request.POST.get("course_id")
            course = get_object_or_404(Course, id=course_id)
            AccessGrant.objects.get_or_create(
                user=student,
                course=course,
                defaults={"source": "admin_manual"}
            )
            messages.success(request, f"تم تفعيل مادة {course.title} للطالب بنجاح.")
            
        elif action == "revoke_course":
            grant_id = request.POST.get("grant_id")
            AccessGrant.objects.filter(id=grant_id, user=student).delete()
            messages.success(request, "تم سحب صلاحية المادة من الطالب.")
            
        elif action == "transfer_grant_device":
            grant_id = request.POST.get("grant_id")
            target_device_fingerprint = request.POST.get("target_device_fingerprint")
            grant = get_object_or_404(AccessGrant, id=grant_id, user=student)
            grant.device_fingerprint = target_device_fingerprint
            grant.save(update_fields=["device_fingerprint"])
            messages.success(request, f"تم نقل تفعيل مادة {grant.course.title if grant.course else 'الخاصة'} إلى الجهاز الجديد بنجاح.")
            
        elif action == "reset_print_quota":
            grant_id = request.POST.get("grant_id")
            grant = get_object_or_404(AccessGrant, id=grant_id, user=student)
            grant.main_pdf_printed = 0
            grant.file1_printed = 0
            grant.file2_printed = 0
            grant.save(update_fields=["main_pdf_printed", "file1_printed", "file2_printed"])
            messages.success(request, f"تم إعادة تعيين عدادات الطباعة لمادة {grant.course.title if grant.course else ''} بنجاح.")

        elif action == "regenerate_print_password":
            grant_id = request.POST.get("grant_id")
            grant = get_object_or_404(AccessGrant, id=grant_id, user=student)
            import random
            import string
            grant.print_password = "".join(random.choices(string.digits, k=6))
            grant.save(update_fields=["print_password"])
            messages.success(request, f"تم توليد كلمة مرور طباعة جديدة لمادة {grant.course.title if grant.course else ''} وهي: {grant.print_password}")
            
        elif action == "update_profile_admin":
            governorate = request.POST.get("governorate")
            track = request.POST.get("track")
            profile = student.student_profile
            profile.governorate = governorate
            profile.track = track
            profile.save()
            messages.success(request, "تم تحديث الفرع والمحافظة للطالب بنجاح.")
            
        elif action == "whatsapp_message":
            message_text = request.POST.get("message", "").strip()
            profile = getattr(student, "student_profile", None)
            if not profile or not profile.phone:
                messages.error(request, "الطالب لا يمتلك رقم هاتف مسجل للإرسال.")
            elif not message_text:
                messages.warning(request, "نص الرسالة فارغ.")
            else:
                sent = send_whatsapp_message(profile.phone, message_text)
                if sent:
                    messages.success(request, f"✅ تم إرسال رسالة الواتساب بنجاح إلى {student.get_full_name() or student.username}.")
                else:
                    messages.error(request, "❌ فشل الإرسال، تأكد أن البوابة الداخلية للواتساب متصلة حالياً.")

        elif action == "reset_otp_limit":
            from accounts.models import StudentProfile
            profile, _ = StudentProfile.objects.get_or_create(user=student)
            phone = profile.phone or student.username
            
            from dashboard.otp_service import _otp_daily_key
            daily_key = _otp_daily_key(phone)
            cache.delete(daily_key)
            
            for purpose in ["login", "register", "reset_password"]:
                cache.delete(f"otp_attempts:{purpose}:{phone}")
                cache.delete(f"otp_resend:{purpose}:{phone}")
            
            # Reset login attempts blocks for their recent IP addresses
            from dashboard.models import OTPVerificationLog
            ips = OTPVerificationLog.objects.filter(phone=phone).values_list('ip_address', flat=True).distinct()
            for ip in ips:
                if ip:
                    cache.delete(f"login_attempts:{ip}:{student.username}")
                    cache.delete(f"login_attempts:{ip}:{phone}")
            
            messages.success(request, "تم تصفير وفتح إمكانية طلب كود التحقق والدخول للطالب بنجاح.")
            
        elif action == "update_otp_ceiling":
            from accounts.models import StudentProfile
            profile, _ = StudentProfile.objects.get_or_create(user=student)
            try:
                max_daily_otp = int(request.POST.get("max_daily_otp", 10))
                profile.max_daily_otp = max_daily_otp
                profile.save()
                messages.success(request, f"تم تحديث سقف طلب الرموز اليومي إلى {max_daily_otp} بنجاح.")
            except Exception as e:
                messages.error(request, f"فشل تحديث سقف الرموز: {str(e)}")

        return redirect("dashboard:admin_student_detail", user_id=student.id)

    # Context data
    all_courses = Course.objects.filter(status="published").select_related("instructor").order_by("title")
    student_grants = AccessGrant.objects.filter(user=student).select_related("course", "course__instructor", "access_code")
    enrolled_course_ids = student_grants.values_list("course_id", flat=True)
    available_courses = all_courses.exclude(id__in=enrolled_course_ids)
    
    attempts = Attempt.objects.filter(user=student).select_related("exam", "exam__course").order_by("-created_at")[:10]
    devices = UserDevice.objects.filter(user=student).order_by("-last_seen_at")
    for device in devices:
        device.associated_grants = AccessGrant.objects.filter(
            user=student,
            device_fingerprint=device.fingerprint
        ).select_related("course", "course__instructor", "access_code")
    
    from accounts.models import Governorate, AcademicBranch
    all_governorates = Governorate.objects.filter(is_active=True).order_by("sort_order", "name")
    all_tracks = AcademicBranch.objects.filter(is_active=True).order_by("sort_order", "name")

    # OTP verification details
    from dashboard.otp_service import _otp_daily_key
    from dashboard.models import OTPVerificationLog
    
    phone_number = student.student_profile.phone or student.username if hasattr(student, 'student_profile') else student.username
    otp_count_today = OTPVerificationLog.objects.filter(
        phone=phone_number,
        created_at__date=timezone.now().date()
    ).count()
    
    daily_key = _otp_daily_key(phone_number)
    otp_cache_count = int(cache.get(daily_key, 0))
    max_daily_otp = student.student_profile.max_daily_otp if hasattr(student, 'student_profile') else 10

    context = {
        "student": student,
        "available_courses": available_courses,
        "student_grants": student_grants,
        "attempts": attempts,
        "devices": devices,
        "all_governorates": all_governorates,
        "all_tracks": all_tracks,
        "otp_count_today": otp_count_today,
        "otp_cache_count": otp_cache_count,
        "max_daily_otp": max_daily_otp,
    }
    return render(request, "dashboard/admin_student_detail.html", context)


def _get_sham_cash_center():
    """Get or Create the permanent administration accounting center for Sham Cash tracking."""
    from billing.models import SalesCenter
    center, _ = SalesCenter.objects.get_or_create(
        name="حساب شام كاش (الإدارة)",
        defaults={"is_active": True, "phone": "إدارة"}
    )
    return center

@admin_required
def admin_sell_codes(request):
    """شاشة البيع الموحدة - بيع الأكواد للطلاب مباشرة وتتبع رصيد الإدارة"""
    from learning.models import Course
    from billing.models import CoursePackage, AccessCode, AccessCodeBatch
    from accounts.models import StudentProfile
    from django.contrib.auth.models import User
    from django.db import transaction
    from django.utils import timezone
    from django.db.models import Sum
    from dashboard.security import sanitize_plain_text
    from billing.services import create_codes_for_batch
    
    sham_center = _get_sham_cash_center()
    
    if request.method == "POST":
        action = request.POST.get("action", "sell_code")
        
        # NEW Feature: Create quick admin batch 
        if action == "create_admin_batch":
            course_id = request.POST.get("batch_course_id")
            package_id = request.POST.get("batch_package_id")
            prefix = request.POST.get("batch_prefix", "").upper().strip()
            quantity = int(request.POST.get("batch_quantity") or 0)
            
            if not quantity or (not course_id and not package_id):
                messages.error(request, "الرجاء إدخال الكمية واختيار مادة أو باقة.")
            else:
                course = Course.objects.filter(id=course_id).first() if course_id else None
                package = CoursePackage.objects.filter(id=package_id).first() if package_id else None
                target_name = course.title if course else (package.name if package else "كود")
                
                batch = AccessCodeBatch.objects.create(
                    name=f"إدارة: {target_name} - {timezone.now().strftime('%Y/%m/%d')}",
                    course=course,
                    package=package,
                    sales_center=sham_center,
                    allocated_count=quantity,
                    code_prefix=prefix,
                    notes="إنشاء مباشر عبر شاشة الإدارة"
                )
                created_cnt = create_codes_for_batch(batch, quantity, free_codes=False)
                messages.success(request, f"تم بنجاح إنشاء دفعة تحتوي على {created_cnt} كود للإدارة.")
            return redirect("dashboard:admin_sell_codes")

        # Default: Sell Code Action
        code_id = request.POST.get("code_id")
        student_phone = request.POST.get("student_phone")
        student_name = request.POST.get("student_name")
        price_amount = request.POST.get("price_amount")
        
        if not code_id or not student_phone or not student_name:
            messages.error(request, "الرجاء ملء كافة الحقول المطلوبة.")
            return redirect("dashboard:admin_sell_codes")
            
        try:
            with transaction.atomic():
                access_code = AccessCode.objects.select_for_update().get(id=code_id, status="active", sale_status__in=["available", "reserved"])
                
                # Normalize phone number
                student_phone = student_phone.strip()
                student_name = sanitize_plain_text(student_name.strip())
                
                # Record assignment to code itself without creating system account (student will register manually)

                access_code.assigned_student_name = student_name
                access_code.assigned_student_phone = student_phone
                access_code.sale_status = "sold"
                access_code.sold_by = request.user
                access_code.sold_at = timezone.now()
                
                # Map to Sham Cash account tracking if code belongs to no explicit external center
                if not access_code.sales_center:
                    access_code.sales_center = sham_center
                
                if price_amount:
                    access_code.sold_price_cents = int(float(price_amount) * 100)
                    access_code.price_reason = "تحديد يدوي من شاشة المبيعات"
                else:
                    base_price = 0
                    if access_code.course:
                        base_price = access_code.course.price_cents or 0
                    elif access_code.package:
                        base_price = access_code.package.price_cents or 0
                        
                    access_code.sold_price_cents = base_price
                    access_code.price_reason = "سعر كامل"
                    
                access_code.save()
                messages.success(request, f"تم تسجيل بيع الكود بنجاح للطالب {student_name}.")
                return redirect(f"/admin-dashboard/billing/sell-codes/?sold_id={access_code.id}")
                    
        except AccessCode.DoesNotExist:
            messages.error(request, "الكود المحدد غير موجود أو تم بيعه مسبقاً.")
        except Exception as e:
            messages.error(request, f"حدث خطأ أثناء بيع الكود: {str(e)}")
            
        return redirect("dashboard:admin_sell_codes")
        
    courses = Course.objects.filter(status="published").select_related("instructor").order_by("title")
    packages = CoursePackage.objects.filter(is_active=True).order_by("name")
    
    # --- Calc Sham Cash Admin Stats ---
    sold_admin_codes = AccessCode.objects.filter(sales_center=sham_center, sale_status="sold").select_related("course", "package")
    real_cents = sold_admin_codes.aggregate(s=Sum("sold_price_cents"))["s"] or 0
    
    std_cents = 0
    for c in sold_admin_codes:
        if c.course and c.course.price_cents:
            std_cents += c.course.price_cents
        elif c.package and c.package.price_cents:
            std_cents += c.package.price_cents
            
    sham_stats = {
        "total_sold": sold_admin_codes.count(),
        "balance_syp": int(real_cents / 100),
        "expected_syp": int(std_cents / 100),
        "discount_syp": int((std_cents - real_cents) / 100),
    }
    
    sold_id = request.GET.get("sold_id")
    sold_code = None
    sold_price_syp = 0
    sold_code_qr = None
    platform_qr = None
    
    if sold_id:
        sold_code = AccessCode.objects.filter(id=sold_id, sale_status="sold").select_related("course", "package", "course__instructor").first()
        if sold_code:
            if sold_code.sold_price_cents:
                sold_price_syp = int(sold_code.sold_price_cents / 100)
            
            # Generate QR codes for visual card rendering
            sold_code_qr = _generate_qr_base64(sold_code.code)
            platform_url = request.build_absolute_uri("/")
            platform_qr = _generate_qr_base64(platform_url)
        
    context = {
        "courses": courses,
        "packages": packages,
        "sold_code": sold_code,
        "sold_price_syp": sold_price_syp,
        "sold_code_qr": sold_code_qr,
        "platform_qr": platform_qr,
        "sham_stats": sham_stats,
        "sham_center_id": sham_center.id,
    }
    return render(request, "dashboard/admin_sell_codes.html", context)


@admin_required
def get_available_codes_api(request):
    from django.http import JsonResponse
    course_id = request.GET.get("course_id")
    package_id = request.GET.get("package_id")
    sales_center_id = request.GET.get("sales_center_id")
    from billing.models import AccessCode
    
    filters = {"status": "active", "sale_status__in": ["available", "reserved"]}
    if course_id:
        filters["course_id"] = course_id
    elif package_id:
        filters["package_id"] = package_id
    else:
        return JsonResponse({"codes": []})
        
    if sales_center_id:
        if sales_center_id.isdigit():
            filters["sales_center_id"] = sales_center_id
        elif sales_center_id == "none":
            filters["sales_center__isnull"] = True
            
    codes = AccessCode.objects.filter(**filters).values("id", "code")[:100]
    return JsonResponse({"codes": list(codes)})


@admin_required
@require_POST
def admin_action_migrate_thair(request):
    """Admin action to migrate legacy data into the new Sham Cash ledger."""
    from billing.models import SalesCenter, AccessCode, AccessCodeBatch
    from django.db import transaction
    from django.contrib import messages
    from django.shortcuts import redirect

    target = _get_sham_cash_center()
    pattern = "ثائر"
    
    try:
        with transaction.atomic():
            # 1. Find any batches whose NAME contains the pattern
            target_batches = AccessCodeBatch.objects.filter(name__icontains=pattern).exclude(sales_center=target)
            batch_ids = list(target_batches.values_list('id', flat=True))
            updated_batches_by_name = target_batches.update(sales_center=target)
            
            # Move codes of those specific batches
            updated_codes_by_batch_link = 0
            if batch_ids:
                updated_codes_by_batch_link = AccessCode.objects.filter(batch_id__in=batch_ids).exclude(sales_center=target).update(sales_center=target)

            # 2. Move everything linked to any center containing "ثائر"
            legacy_centers = SalesCenter.objects.filter(name__icontains=pattern).exclude(id=target.id)
            updated_batches_by_center = 0
            updated_codes_by_center = 0
            if legacy_centers.exists():
                updated_batches_by_center = AccessCodeBatch.objects.filter(sales_center__in=legacy_centers).update(sales_center=target)
                updated_codes_by_center = AccessCode.objects.filter(sales_center__in=legacy_centers).update(sales_center=target)

            final_batch_total = updated_batches_by_name + updated_batches_by_center
            final_code_total = updated_codes_by_batch_link + updated_codes_by_center

            if final_batch_total == 0 and final_code_total == 0:
                messages.info(request, f"تم التحقق، لا يوجد حالياً أي دفعات أو أكواد تحمل اسم ({pattern}) لنقلها.")
            else:
                messages.success(request, f"✅ اكتمل الدمج! تم نقل إجمالي {final_batch_total} دفعة و {final_code_total} كود إلى حساب شام كاش بنجاح.")
    except Exception as e:
        messages.error(request, f"فشلت عملية النقل: {str(e)}")

    return redirect("dashboard:admin_sell_codes")

@admin_required
def admin_whatsapp_control(request):
    import time
    import threading
    from django.contrib.auth import get_user_model
    from accounts.models import StudentProfile, AcademicBranch

    User = get_user_model()

    # Helper for background broadcasts with advanced humanized anti-spam throttle & grammar parser
    def _background_broadcast(student_list, raw_message, skip_duplicates=False):
        import random
        import hashlib
        import logging
        logger = logging.getLogger(__name__)
        from .whatsapp_utils import format_phone_to_intl, parse_gender_grammar
        from dashboard.models import WhatsAppMessageLog
        
        try:
            raw_hash = ""
            if skip_duplicates:
                normalized = " ".join(raw_message.strip().split())
                raw_hash = hashlib.sha256(normalized.encode('utf-8')).hexdigest()
                
            sent_count = 0
            for profile in student_list:
                try:
                    if profile and profile.phone:
                        # Format phone
                        clean_phone = format_phone_to_intl(profile.phone)
                        
                        # Check for duplicates if requested
                        if skip_duplicates and raw_hash:
                            if WhatsAppMessageLog.objects.filter(phone=clean_phone, raw_text_hash=raw_hash).exists():
                                continue
                                
                        student_user = profile.user
                        student_name = student_user.get_full_name() or student_user.username
                        
                        # 1. Basic personalization ({name})
                        personalized = raw_message.replace("{name}", student_name).replace("{الاسم}", student_name)
                        
                        # 2. Intelligent Grammar Restructuring ({مذكر|مؤنث})
                        student_gender = getattr(profile, 'gender', 'unknown')
                        personalized = parse_gender_grammar(personalized, student_gender)
                        
                        # Trigger immediate message push
                        send_whatsapp_message(profile.phone, personalized, raw_text=raw_message)
                        sent_count += 1
                        
                        # 3. Advanced Jitter: Simulate natural typing & thinking gap
                        natural_delay = random.uniform(4.0, 8.0)
                        time.sleep(natural_delay)
                        
                        # 4. Batch Breather: Every 12 messages, take a deep human pause
                        if sent_count % 12 == 0:
                            deep_pause = random.uniform(20.0, 45.0)
                            time.sleep(deep_pause)
                except Exception as inner_err:
                    logger.error(f"Error sending message in broadcast loop to profile {profile.id if profile else 'Unknown'}: {inner_err}", exc_info=True)
        except Exception as outer_err:
            logger.error(f"Critical error in _background_broadcast task setup: {outer_err}", exc_info=True)

    if request.method == "POST":
        # 1. Handle Logout
        if "logout" in request.POST:
            result = logout_whatsapp()
            if result.get("status") == "success":
                messages.success(request, "تم تسجيل الخروج بنجاح.")
            else:
                messages.error(request, "فشل تسجيل الخروج.")
            return redirect("dashboard:admin_whatsapp_control")
        
        # 2. Handle Single Send
        elif "send_single" in request.POST:
            target_phone = request.POST.get("phone", "").strip()
            message_body = request.POST.get("message", "").strip()
            if target_phone and message_body:
                from .whatsapp_utils import format_phone_to_intl, parse_gender_grammar
                
                clean_phone = format_phone_to_intl(target_phone)
                last_nine = clean_phone[-9:] if len(clean_phone) >= 9 else clean_phone
                
                profile = None
                if last_nine:
                    profile = StudentProfile.objects.filter(phone__contains=last_nine).first()
                
                student_name = "طالبنا العزيز"
                student_gender = "unknown"
                
                if profile:
                    student_user = profile.user
                    student_name = student_user.get_full_name() or student_user.username
                    student_gender = getattr(profile, 'gender', 'unknown')
                
                # 1. Basic personalization ({name})
                personalized = message_body.replace("{name}", student_name).replace("{الاسم}", student_name)
                
                # 2. Intelligent Grammar Restructuring ({مذكر|مؤنث})
                personalized = parse_gender_grammar(personalized, student_gender)
                
                sent = send_whatsapp_message(target_phone, personalized)
                if sent:
                    messages.success(request, f"✅ تم إرسال الرسالة بنجاح إلى الرقم {target_phone}")
                else:
                    messages.error(request, "❌ فشل الإرسال، تأكد أن البوابة متصلة.")
            else:
                messages.warning(request, "يرجى ملء الرقم والرسالة.")
            return redirect("dashboard:admin_whatsapp_control")

        # 3. Handle Advanced Segmentation Broadcast with Gender Filter
        elif "send_broadcast" in request.POST:
            target_type = request.POST.get("target_type", "subscribed")
            message_body = request.POST.get("message", "").strip()
            branch_filter = request.POST.get("branch_filter", "").strip()
            gender_filter = request.POST.get("gender_filter", "").strip()  # 'male', 'female', or empty
            skip_duplicates = request.POST.get("skip_duplicates") == "on"
            
            if not message_body:
                messages.warning(request, "لا يمكن إطلاق حملة برسالة فارغة.")
                return redirect("dashboard:admin_whatsapp_control")
 
            profiles = []
            msg_segment = ""
            
            if target_type == "subscribed":
                # Subscribed to a specific course
                course_id = request.POST.get("course_id")
                if not course_id:
                    messages.warning(request, "يرجى تحديد الدورة المستهدفة للإرسال للمشتركين.")
                    return redirect("dashboard:admin_whatsapp_control")
                
                course = get_object_or_404(Course, id=course_id)
                grants = AccessGrant.objects.filter(course=course).select_related('user__student_profile')
                for grant in grants:
                    p = getattr(grant.user, 'student_profile', None)
                    if p and p.phone:
                        # Apply manual filtering for branch & gender
                        if not branch_filter or p.track == branch_filter:
                            if not gender_filter or p.gender == gender_filter:
                                profiles.append(p)
                msg_segment = f"المشتركين في دورة ({course.title})"
 
            elif target_type == "all_registered":
                # Everyone with a student profile and phone registered
                qs = StudentProfile.objects.exclude(phone__isnull=True).exclude(phone="").select_related('user')
                if branch_filter:
                    qs = qs.filter(track=branch_filter)
                if gender_filter:
                    qs = qs.filter(gender=gender_filter)
                profiles = list(qs)
                msg_segment = "جميع المسجلين بالمنصة"
 
            elif target_type == "subscribed_any":
                # Active subscribers having at least 1 active access grant
                active_grants_user_ids = AccessGrant.objects.values_list('user_id', flat=True).distinct()
                qs = StudentProfile.objects.filter(user_id__in=active_grants_user_ids).exclude(phone__isnull=True).exclude(phone="").select_related('user')
                if branch_filter:
                    qs = qs.filter(track=branch_filter)
                if gender_filter:
                    qs = qs.filter(gender=gender_filter)
                profiles = list(qs)
                msg_segment = "المشتركون الفعّالون (لديهم اشتراك نشط)"
 
            elif target_type == "unsubscribed_any":
                # Registered students who DO NOT have any active access grants yet!
                active_grants_user_ids = AccessGrant.objects.values_list('user_id', flat=True)
                qs = StudentProfile.objects.exclude(phone__isnull=True).exclude(phone="").exclude(user_id__in=active_grants_user_ids).select_related('user')
                if branch_filter:
                    qs = qs.filter(track=branch_filter)
                if gender_filter:
                    qs = qs.filter(gender=gender_filter)
                profiles = list(qs)
                msg_segment = "المسجلين غير المشتركين بأي مادة"
            
            if branch_filter:
                msg_segment += f" - فرع ({branch_filter})"
            else:
                msg_segment += " - جميع الفروع"
                
            if gender_filter == 'male':
                msg_segment += " (ذكور فقط)"
            elif gender_filter == 'female':
                msg_segment += " (إناث فقط)"
            
            if not profiles:
                messages.info(request, f"لم يتم العثور على طلاب تطابق شريحة ({msg_segment}).")
            else:
                count = len(profiles)
                # Trigger background broadcasting
                threading.Thread(
                    target=_background_broadcast,
                    args=(profiles, message_body, skip_duplicates),
                    daemon=True
                ).start()
                
                info_msg = f"🚀 انطلقت حملة الإرسال الذكية لـ ({msg_segment})! سيتم مراسلة {count} طالب في الخلفية."
                if skip_duplicates:
                    info_msg += " (مع تجنب المراسلة المكررة)"
                messages.success(request, info_msg)
            return redirect("dashboard:admin_whatsapp_control")

        # 4. Save Custom Template
        elif "add_template" in request.POST:
            title = request.POST.get("template_title", "").strip()
            content = request.POST.get("template_content", "").strip()
            if title and content:
                WhatsAppTemplate.objects.create(title=title, content=content)
                messages.success(request, "✅ تم حفظ قالب الرسالة الجديد بنجاح.")
            else:
                messages.warning(request, "يرجى ملء عنوان ومحتوى القالب.")
            return redirect("dashboard:admin_whatsapp_control")

        # 5. Edit Existing Saved Template
        elif "edit_template" in request.POST:
            template_id = request.POST.get("template_id")
            title = request.POST.get("template_title", "").strip()
            content = request.POST.get("template_content", "").strip()
            if title and content:
                WhatsAppTemplate.objects.filter(id=template_id).update(title=title, content=content)
                messages.success(request, "✅ تم حفظ تعديلات القالب بنجاح.")
            else:
                messages.warning(request, "يرجى عدم ترك عنوان أو محتوى القالب فارغاً.")
            return redirect("dashboard:admin_whatsapp_control")

        # 6. Delete Saved Template
        elif "delete_template" in request.POST:
            template_id = request.POST.get("template_id")
            WhatsAppTemplate.objects.filter(id=template_id).delete()
            messages.success(request, "تم حذف قالب الرسالة بنجاح.")
            return redirect("dashboard:admin_whatsapp_control")

        # 7. Toggle 2FA verification
        elif "toggle_2fa" in request.POST:
            from .whatsapp_utils import set_2fa_disabled
            disable_otp = request.POST.get("disable_2fa") == "true"
            success = set_2fa_disabled(disable_otp)
            if success:
                if disable_otp:
                    messages.success(request, "🔓 تم إيقاف التحقق الثنائي بنجاح. يمكن للطلاب تسجيل الدخول وإنشاء حسابات دون إرسال رمز تحقق.")
                else:
                    messages.success(request, "🔒 تم تفعيل التحقق الثنائي بنجاح. سيتم طلب رمز تحقق عبر واتساب عند تسجيل الدخول أو إنشاء حساب.")
            else:
                messages.error(request, "❌ فشل حفظ الإعدادات. يرجى التحقق من صلاحيات المجلدات على الخادم.")
            return redirect("dashboard:admin_whatsapp_control")

    status_data = get_whatsapp_status()
    # Fetch with related instructor information to display instructor name in selection list
    courses = Course.objects.filter(status="published").select_related("instructor").order_by("title")
    db_templates = WhatsAppTemplate.objects.all()
    branches = AcademicBranch.objects.filter(is_active=True).order_by("sort_order", "name")

    # Fetch recent students from both profiles and sold codes to list under the VCF hub
    from billing.models import AccessCode
    recent_sold_codes = AccessCode.objects.filter(
        sale_status="sold"
    ).exclude(
        assigned_student_phone=""
    ).exclude(
        assigned_student_phone__isnull=True
    ).order_by("-sold_at")[:15]
    
    recent_students = []
    seen_phones = set()
    
    for code in recent_sold_codes:
        phone = code.assigned_student_phone.strip()
        name = code.assigned_student_name.strip()
        if phone and phone not in seen_phones:
            seen_phones.add(phone)
            recent_students.append({
                "name": name,
                "phone": phone,
                "source": "كود مباع",
                "date": code.sold_at
            })
            
    recent_profiles = StudentProfile.objects.exclude(phone="").exclude(phone__isnull=True).select_related("user").order_by("-id")[:15]
    for p in recent_profiles:
        phone = p.phone.strip()
        name = p.user.get_full_name() or p.user.username
        if phone and phone not in seen_phones:
            seen_phones.add(phone)
            recent_students.append({
                "name": name,
                "phone": phone,
                "source": "حساب مسجل",
                "date": p.user.date_joined
            })
            
    recent_students.sort(key=lambda x: x["date"] or timezone.now(), reverse=True)
    recent_students = recent_students[:12]

    from .whatsapp_utils import is_2fa_disabled
    return render(request, "dashboard/admin_whatsapp_control.html", {
        "status": status_data.get("status", "offline"),
        "qr": status_data.get("qr"),
        "has_qr": status_data.get("hasQr", False),
        "wa_user": status_data.get("user"),
        "courses": courses,
        "templates": db_templates,
        "branches": branches,
        "recent_students": recent_students,
        "otp_2fa_disabled": is_2fa_disabled(),
    })



@admin_required
def admin_impersonate_instructor(request, instructor_id):
    instructor_user = get_object_or_404(User, id=instructor_id)
    
    # Verify this user is an active instructor
    if not hasattr(instructor_user, 'instructor_profile'):
        messages.error(request, "هذا المستخدم ليس مدرساً.")
        return redirect("dashboard:admin_instructors")
        
    admin_id = request.user.id
    
    # Authenticate and login as the instructor
    login(request, instructor_user)
    
    # Store original admin ID in session (after login to avoid session flush issues)
    request.session["impersonator_admin_id"] = admin_id
    request.session["impersonator_type"] = "instructor"
    
    messages.success(request, f"✅ تم تسجيل الدخول بصفتك الأستاذ {instructor_user.get_full_name() or instructor_user.username} بنجاح.")
    return redirect("dashboard:instructor_dashboard")


@admin_required
def admin_impersonate_student(request, student_id):
    student_user = get_object_or_404(User, id=student_id)
    
    admin_id = request.user.id
    
    # Authenticate and login as the student
    login(request, student_user)
    
    # Store original admin ID in session (after login to avoid session flush issues)
    request.session["impersonator_admin_id"] = admin_id
    request.session["impersonator_type"] = "student"
    
    messages.success(request, f"✅ تم تسجيل الدخول بصفتك الطالب {student_user.get_full_name() or student_user.username} بنجاح.")
    return redirect("dashboard:my_courses")


@login_required
def exit_impersonate(request):
    admin_id = request.session.get("impersonator_admin_id")
    if not admin_id:
        messages.error(request, "إجراء غير مصرح به.")
        return redirect("dashboard:home")
        
    admin_user = get_object_or_404(User, id=admin_id)
    impersonator_type = request.session.pop("impersonator_type", "student")
    
    # Clear impersonation and login back as admin
    request.session.pop("impersonator_admin_id", None)
    login(request, admin_user)
    
    messages.success(request, "✅ تم العودة إلى حساب الإدارة بنجاح.")
    
    if impersonator_type == "instructor":
        return redirect("dashboard:admin_instructors")
    else:
        return redirect("dashboard:admin_students")



@admin_required
@require_POST
def admin_whatsapp_send_sold_card(request):
    from django.http import JsonResponse
    import json
    import requests
    from .whatsapp_utils import WHATSAPP_GATEWAY_URL, WHATSAPP_API_SECRET
    
    try:
        data = json.loads(request.body)
        phone = data.get("phone", "").strip()
        message = data.get("message", "").strip()
        image_base64 = data.get("image", "")  # data:image/png;base64,....
        
        if not phone or not message:
            return JsonResponse({"ok": False, "error": "رقم الهاتف ونص الرسالة مطلوبان."})
        
        payload = {
            "number": phone,
            "message": message,
        }
        if image_base64:
            payload["image"] = image_base64
            
        headers = {
            "X-API-Secret": WHATSAPP_API_SECRET,
            "Content-Type": "application/json"
        }
        
        resp = requests.post(f"{WHATSAPP_GATEWAY_URL}/send-message", json=payload, headers=headers, timeout=25)
        
        if resp.status_code == 200:
            return JsonResponse({"ok": True, "message": "تم إرسال بطاقة التفعيل آلياً بنجاح!"})
        else:
            return JsonResponse({"ok": False, "error": f"فشل إرسال الرسالة من سيرفر الواتساب: {resp.text}"})
            
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"حدث خطأ أثناء معالجة الطلب: {str(e)}"})


@admin_required
def admin_export_contacts_vcf(request):
    from django.http import HttpResponse
    from billing.models import AccessCode
    from accounts.models import StudentProfile
    
    single_phone = request.GET.get("phone", "").strip()
    single_name = request.GET.get("name", "").strip()
    
    contacts = []
    
    if single_phone:
        name = single_name or f"طالب {single_phone}"
        contacts.append((name, single_phone))
        filename = f"contact_{single_phone}.vcf"
    else:
        # Query all sold codes that have student name and phone
        sold_codes = AccessCode.objects.filter(
            sale_status="sold"
        ).exclude(
            assigned_student_phone=""
        ).exclude(
            assigned_student_phone__isnull=True
        ).order_by("-sold_at")
        
        seen_phones = set()
        
        # 1. Add sold codes contacts
        for code in sold_codes:
            phone = code.assigned_student_phone.strip()
            name = code.assigned_student_name.strip()
            if phone and phone not in seen_phones:
                seen_phones.add(phone)
                contacts.append((name, phone))
                
        # 2. Add student profiles contacts
        profiles = StudentProfile.objects.exclude(phone="").exclude(phone__isnull=True).select_related("user")
        for p in profiles:
            phone = p.phone.strip()
            name = p.user.get_full_name() or p.user.username
            if phone and phone not in seen_phones:
                seen_phones.add(phone)
                contacts.append((name, phone))
        filename = "mohtarifo_contacts.vcf"
            
    # Build the multi-contact VCF content
    vcf_content = ""
    for name, phone in contacts:
        # Robust Syrian phone normalization
        clean_phone = ''.join(filter(str.isdigit, str(phone)))
        if clean_phone.startswith('00963'):
            clean_phone = '963' + clean_phone[5:]
        elif clean_phone.startswith('09') and len(clean_phone) == 10:
            clean_phone = '963' + clean_phone[1:]
        elif len(clean_phone) == 9 and clean_phone.startswith('9'):
            clean_phone = '963' + clean_phone
            
        vcf_content += "BEGIN:VCARD\n"
        vcf_content += "VERSION:3.0\n"
        vcf_content += f"FN:{name}\n"
        vcf_content += f"TEL;TYPE=CELL;TYPE=VOICE;waid={clean_phone}:+{clean_phone}\n"
        vcf_content += "END:VCARD\n"
        
    response = HttpResponse(vcf_content, content_type="text/x-vcard")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@admin_required
def admin_send_contacts_vcf_whatsapp(request):
    from django.http import JsonResponse
    import requests
    import base64
    from billing.models import AccessCode
    from accounts.models import StudentProfile
    from .whatsapp_utils import WHATSAPP_GATEWAY_URL, WHATSAPP_API_SECRET
    
    # Target phone
    single_phone = request.GET.get("phone", "").strip()
    single_name = request.GET.get("name", "").strip()
    
    contacts = []
    seen_phones = set()
    admin_phone = "963983232446"
    
    if single_phone:
        name = single_name or f"طالب {single_phone}"
        contacts.append((name, single_phone))
        file_label = f"طالب فردي ({name})"
        fileName = f"contact_{single_phone}.vcf"
    else:
        # Query all sold codes that have student name and phone
        sold_codes = AccessCode.objects.filter(
            sale_status="sold"
        ).exclude(
            assigned_student_phone=""
        ).exclude(
            assigned_student_phone__isnull=True
        ).order_by("-sold_at")
        
        for code in sold_codes:
            phone = code.assigned_student_phone.strip()
            name = code.assigned_student_name.strip()
            if phone and phone not in seen_phones:
                seen_phones.add(phone)
                contacts.append((name, phone))
                
        profiles = StudentProfile.objects.exclude(phone="").exclude(phone__isnull=True).select_related("user")
        for p in profiles:
            phone = p.phone.strip()
            name = p.user.get_full_name() or p.user.username
            if phone and phone not in seen_phones:
                seen_phones.add(phone)
                contacts.append((name, phone))
        file_label = "جميع جهات الاتصال بالكامل"
        fileName = "mohtarifo_contacts.vcf"
        
    # Build VCF content
    vcf_content = ""
    for name, phone in contacts:
        # Robust Syrian phone normalization
        clean_phone = ''.join(filter(str.isdigit, str(phone)))
        if clean_phone.startswith('00963'):
            clean_phone = '963' + clean_phone[5:]
        elif clean_phone.startswith('09') and len(clean_phone) == 10:
            clean_phone = '963' + clean_phone[1:]
        elif len(clean_phone) == 9 and clean_phone.startswith('9'):
            clean_phone = '963' + clean_phone
            
        vcf_content += "BEGIN:VCARD\n"
        vcf_content += "VERSION:3.0\n"
        vcf_content += f"FN:{name}\n"
        vcf_content += f"TEL;TYPE=CELL;TYPE=VOICE;waid={clean_phone}:+{clean_phone}\n"
        vcf_content += "END:VCARD\n"
        
    if not vcf_content:
        return JsonResponse({"ok": False, "error": "لا توجد أسماء لتصديرها."})
        
    # Base64 encode
    vcf_bytes = vcf_content.encode('utf-8')
    vcf_base64 = "data:text/vcard;base64," + base64.b64encode(vcf_bytes).decode('utf-8')
    
    payload = {
        "number": admin_phone,
        "document": vcf_base64,
        "mimetype": "text/vcard",
        "fileName": fileName,
        "message": f"تصدير جهات اتصال منصة محترفو التعليم ⚡\nنوع التصدير: {file_label}\nالعدد: {len(contacts)} جهة اتصال."
    }
    
    headers = {
        "X-API-Secret": WHATSAPP_API_SECRET,
        "Content-Type": "application/json"
    }
    
    try:
        resp = requests.post(f"{WHATSAPP_GATEWAY_URL}/send-message", json=payload, headers=headers, timeout=25)
        if resp.status_code == 200:
            return JsonResponse({"ok": True, "message": f"تم إرسال {file_label} بنجاح كملف VCF إلى رقم الإدارة 0983232446! 🎉"})
        else:
            return JsonResponse({"ok": False, "error": f"فشل إرسال الملف عبر بوابة الواتساب: {resp.text}"})
    except Exception as e:
        return JsonResponse({"ok": False, "error": f"حدث خطأ أثناء الاتصال بالبوابة: {str(e)}"})
